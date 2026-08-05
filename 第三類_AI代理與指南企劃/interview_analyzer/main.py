import os
import sys
import json
import uuid
import socket
import itertools
from datetime import datetime
from typing import Optional, List
from fastapi import FastAPI, UploadFile, File, Form, HTTPException, Header, Query
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse, JSONResponse
from pydantic import BaseModel, Field
from google import genai
from google.genai import types
from docx_generator import generate_interview_report_docx, generate_pre_interview_report_docx
from excel_generator import generate_interview_report_excel, generate_pre_interview_report_excel

if hasattr(sys.stdout, 'reconfigure'):
    try:
        sys.stdout.reconfigure(encoding='utf-8')
    except:
        pass

app = FastAPI(title="AI Voice Interview Analyzer for Materials & Supply Chain HR")

DATA_DIR = "data"
AUDIO_DIR = os.path.join(DATA_DIR, "audios")
DB_FILE = os.path.join(DATA_DIR, "db.json")

os.makedirs(AUDIO_DIR, exist_ok=True)
os.makedirs("static", exist_ok=True)
if not os.path.exists(DB_FILE):
    with open(DB_FILE, "w", encoding="utf-8") as f:
        json.dump([], f)

FLASH_MODELS_CASCADE = [
    'gemini-3.6-flash',
    'gemini-3.5-flash',
    'gemini-3.1-flash-lite',
    'gemini-3-flash',
    'gemini-2.5-flash',
    'gemini-2.0-flash',
    'gemini-1.5-flash',
]

JOB_SPECS_FILE = os.path.join(DATA_DIR, "職缺需求與部門清單.xlsx")

def ensure_job_specs_excel():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    if os.path.exists(JOB_SPECS_FILE):
        return

    wb = Workbook()
    
    # ── 分頁 1：職缺需求與規格清單 ──────────────────────────────
    ws1 = wb.active
    ws1.title = "職缺需求與規格清單"
    ws1.views.sheetView[0].showGridLines = True

    # 標題欄位
    headers = [
        "部門類別", "職缺名稱", "104工作內容 (Job Description)", 
        "條件要求 (Requirements)", "推薦黃金 DISC 型態", "快捷標籤"
    ]
    
    ws1.row_dimensions[1].height = 28
    for col_idx, h in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col_idx, value=h)
        cell.font = Font(name="微軟正黑體", size=11, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    default_jobs = [
        [
            "🏭 勝一化工 & 廠區管理類",
            "【彰濱廠區】助理管理師",
            "1. 槽車灌裝進出貨作業與過磅\n2. 包材灌裝作業與庫存管理\n3. 產品調和與投料作業\n4. 堆高機操作與倉庫堆疊\n5. 設備保養檢查與6S環境維護\n6. 其它上級主管交辦事項",
            "工作經歷：不拘 (1年以上經驗佳)\n學歷要求：專科、大學\n科系要求：化學工程、材料工程、自然科學類相關\n具備證照：堆高機操作技術士 (單一級)",
            "CS型 (C型60%+S型40%)",
            "是"
        ],
        [
            "📦 資材與物料供應鏈類",
            "資材部 - 現場助理工程師",
            "1. 現場原料驗收、發料與盤點作業\n2. T100/ERP系統進出貨數據登打與過磅\n3. 包材批號維護與現場庫位管理\n4. 堆高機進出貨過磅與備料作業\n5. 異常庫存回報與料號核對",
            "工作經歷：1年以上\n學歷要求：專科、大學\n具備證照：堆高機操作證照佳\n擅長工具：Excel、ERP系統操作",
            "CS型 (C型60%+S型40%)",
            "是"
        ],
        [
            "📦 資材與物料供應鏈類",
            "資材部 - 資材工程師",
            "1. BOM表結構化分析與物料主檔維護\n2. 物料供需平衡規劃與安全庫存監控\n3. 採購交期追蹤與生產排程溝通\n4. 滯銷與呆滯料號分析改善",
            "工作經歷：2年以上資材/物控經驗\n學歷要求：大學以上\n科系要求：工業工程、資訊管理、企管相關\n擅長工具：Excel進階函數、ERP/T100系統",
            "CS型 (C型70%+S型30%)",
            "是"
        ],
        [
            "📦 資材與物料供應鏈類",
            "資材部 - 資材行政專員",
            "1. 資材部門行政單據審核與歸檔\n2. 跨部門溝通協調與進貨驗收單連動\n3. ERP系統數據核對與報表產出\n4. 部門會議紀錄與資材後勤支援",
            "工作經歷：1年以上行政或資材經驗\n學歷要求：專科、大學\n擅長工具：MS Office (Word, Excel), ERP",
            "SC型 (S型60%+C型40%)",
            "是"
        ],
        [
            "💼 行政、財務與人資類",
            "會計行政人員 / 財務專員",
            "1. 應收/應付款項帳務處理與傳票開立\n2. 零用金管理與費用報銷審核\n3. 月結報表編製與營業稅申報協助\n4. 銀行往來與資金調度作業",
            "工作經歷：1~2年以上會計經驗\n學歷要求：專科、大學\n科系要求：會計、財稅、財務金融相關\n具備證照：丙級會計事務技術士佳",
            "C型 (C型80%+S型20%)",
            "是"
        ],
        [
            "📦 資材與物料供應鏈類",
            "倉管 / 物料管理員",
            "1. 倉庫成品/原料進出貨收發與盤點\n2. 庫位整理、標籤貼附與6S環境維護\n3. 堆高機搬運與車輛裝卸作業",
            "工作經歷：不拘\n學歷要求：高中/職以上\n具備證照：堆高機操作技術士證照",
            "SC型 (S型60%+C型40%)",
            "是"
        ],
        [
            "🔬 研發品保與技術專案類",
            "研發 (RD) 化學工程師",
            "1. 特用化學品與電子級溶劑配方研發\n2. 實驗室分析儀器操作 (GC, HPLC, ICP-MS)\n3. 研發專利檢索與技術報告撰寫",
            "工作經歷：1~3年化學研發經驗\n學歷要求：碩士、大學\n科系要求：化學、化學工程、材料相關",
            "C型 (C型80%+D型20%)",
            "是"
        ],
        [
            "💼 行政、財務與人資類",
            "人資 (HR) 招募/訓練管理師",
            "1. 104職缺維護、履歷篩選與面談邀約\n2. 員工教育訓練課程規劃與執行\n3. 考勤管理、勞健保加退保與薪酬核算",
            "工作經歷：2年以上 HR 經驗\n學歷要求：大學以上\n科系要求：人力資源、企業管理、心理相關",
            "IS型 (I型60%+S型40%)",
            "是"
        ]
    ]

    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'), right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'), bottom=Side(style='thin', color='D9D9D9')
    )

    for row_idx, r_data in enumerate(default_jobs, 2):
        ws1.row_dimensions[row_idx].height = 65
        for col_idx, val in enumerate(r_data, 1):
            cell = ws1.cell(row=row_idx, column=col_idx, value=val)
            cell.font = Font(name="微軟正黑體", size=10)
            cell.border = thin_border
            align_h = "center" if col_idx in [1, 2, 5, 6] else "left"
            cell.alignment = Alignment(horizontal=align_h, vertical="center", wrap_text=True)

    ws1.column_dimensions['A'].width = 25
    ws1.column_dimensions['B'].width = 28
    ws1.column_dimensions['C'].width = 45
    ws1.column_dimensions['D'].width = 40
    ws1.column_dimensions['E'].width = 22
    ws1.column_dimensions['F'].width = 12

    # ── 分頁 2：DISC 人格類型參考說明 ──────────────────────────
    ws2 = wb.create_sheet(title="DISC 人格類型參考說明")
    ws2.views.sheetView[0].showGridLines = True

    headers2 = ["DISC 類型代號", "人格特徵名稱", "核心特質與行為風格", "建議適合之職務類型"]
    ws2.row_dimensions[1].height = 28
    for col_idx, h in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=col_idx, value=h)
        cell.font = Font(name="微軟正黑體", size=11, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    disc_info = [
        ["D (Dominance)", "支配 / 主導型", "果斷獨立、目標導向、重效率與結果、具領導魄力與決策力。", "高階主管、業務開發、專案 PM、生產廠務主管"],
        ["I (Influence)", "影響 / 社交型", "熱情具感染力、善於溝通、創造力高、樂觀外向與展現影響力。", "行銷企劃、人資招募、公關客服、業務專員"],
        ["S (Steadiness)", "穩健 / 支援型", "溫和耐心、重團隊和諧、高忠誠度、配合度高、重視標準流程。", "行政後勤、資材助理、倉管員、客戶服務、出納員"],
        ["C (Conscientiousness)", "謹慎 / 分析型", "嚴謹精確、邏輯導向、注重 SOP 事實與品質、細節敏感度高。", "品保檢驗工程師、資材工程師、會計財務、RD 化學研發"]
    ]

    for row_idx, r_data in enumerate(disc_info, 2):
        ws2.row_dimensions[row_idx].height = 40
        for col_idx, val in enumerate(r_data, 1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=val)
            cell.font = Font(name="微軟正黑體", size=10)
            cell.border = thin_border
            align_h = "center" if col_idx in [1, 2] else "left"
            cell.alignment = Alignment(horizontal=align_h, vertical="center", wrap_text=True)

    ws2.column_dimensions['A'].width = 20
    ws2.column_dimensions['B'].width = 18
    ws2.column_dimensions['C'].width = 45
    ws2.column_dimensions['D'].width = 40

    wb.save(JOB_SPECS_FILE)

ensure_job_specs_excel()

# 動態讀取職缺需求與規格清單 API (從 Excel 檔讀取)
@app.get("/api/job-positions")
async def get_job_positions():
    ensure_job_specs_excel()
    from openpyxl import load_workbook
    
    positions = []
    tags = []
    
    try:
        wb = load_workbook(JOB_SPECS_FILE, data_only=True)
        ws = wb["職缺需求與規格清單"]
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[1]:
                continue
            cat = str(row[0]).strip() if row[0] else "其它職缺類"
            title = str(row[1]).strip()
            jd = str(row[2]).strip() if len(row) > 2 and row[2] else ""
            req = str(row[3]).strip() if len(row) > 3 and row[3] else ""
            disc = str(row[4]).strip() if len(row) > 4 and row[4] else ""
            is_tag = str(row[5]).strip() if len(row) > 5 and row[5] else ""
            
            positions.append({
                "category": cat,
                "title": title,
                "job_description": jd,
                "requirements": req,
                "disc_benchmark": disc,
                "is_tag": (is_tag == "是")
            })
            if is_tag == "是":
                tags.append(title)
    except Exception as e:
        print(f"[Excel Read Error] 讀取職缺清單失敗: {e}")
        
    return {"positions": positions, "tags": tags}

# 連線 104 自動爬取職缺並寫入/更新 Excel 清單 API
@app.post("/api/crawl-104-jobs")
async def crawl_104_jobs(
    company_keyword: str = Query("勝一化工"),
    x_gemini_api_keys: Optional[str] = Header(None)
):
    key_pool = APIKeyPool(x_gemini_api_keys)
    ensure_job_specs_excel()

    prompt = f"""
    你是一位專業的 HR 數據爬蟲與資料結構化專家。
    請連線搜尋並爬取【{company_keyword}】在 104 人力銀行 (104.com.tw) 刊登的真實職缺資料。
    重點包含但不限於：
    1. 【彰濱廠區】助理管理師 / 倉管物料員
    2. 資材助理工程師 / 資材工程師
    3. 廠區化學工程師 / 環安衛管理師 / 會計行政

    請將爬取到的職缺，整理為以下結構化資料清單：
    - category: 部門與職務類別
    - title: 104 刊登之職缺名稱
    - job_description: 104 上顯示的完整工作內容 (列出 1. 2. 3. 點)
    - requirements: 104 上要求的經歷、學歷、科系、擅長工具與具備證照
    - disc_benchmark: 建議之黃金 DISC 型態 (例如 CS型)
    - is_tag: 是否設為快捷標籤 ("是" 或 "否")
    """

    class ScrapedJobItem(BaseModel):
        category: str
        title: str
        job_description: str
        requirements: str
        disc_benchmark: str
        is_tag: str

    class ScrapedJobsList(BaseModel):
        jobs: List[ScrapedJobItem]

    def do_crawl(client: genai.Client):
        parsed_res, used_model = call_gemini_with_model_cascade(
            client=client,
            contents=[prompt],
            response_schema=ScrapedJobsList
        )
        return parsed_res.model_dump(), used_model

    scraped_data, used_model = key_pool.execute_with_retry(do_crawl)
    jobs_list = scraped_data.get("jobs", [])

    if jobs_list:
        from openpyxl import load_workbook, Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        wb = load_workbook(JOB_SPECS_FILE)
        ws = wb["職缺需求與規格清單"]

        # 抓取目前現有職缺名稱以避免重複覆蓋
        existing_titles = set()
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row and row[1]:
                existing_titles.add(str(row[1]).strip())

        thin_border = Border(
            left=Side(style='thin', color='D9D9D9'), right=Side(style='thin', color='D9D9D9'),
            top=Side(style='thin', color='D9D9D9'), bottom=Side(style='thin', color='D9D9D9')
        )

        r_idx = ws.max_row + 1
        new_added_count = 0

        for job in jobs_list:
            t = job.get("title", "").strip()
            if not t or t in existing_titles:
                continue
            
            ws.row_dimensions[r_idx].height = 65
            r_data = [
                job.get("category", "104抓取職缺"),
                t,
                job.get("job_description", ""),
                job.get("requirements", ""),
                job.get("disc_benchmark", "CS型"),
                job.get("is_tag", "是")
            ]

            for col_idx, val in enumerate(r_data, 1):
                cell = ws.cell(row=r_idx, column=col_idx, value=val)
                cell.font = Font(name="微軟正黑體", size=10)
                cell.border = thin_border
                align_h = "center" if col_idx in [1, 2, 5, 6] else "left"
                cell.alignment = Alignment(horizontal=align_h, vertical="center", wrap_text=True)

            r_idx += 1
            new_added_count += 1

        wb.save(JOB_SPECS_FILE)

    return {"status": "ok", "added_count": len(jobs_list), "used_model": used_model}

def get_local_ip():
    try:
        s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        s.connect(("8.8.8.8", 80))
        ip = s.getsockname()[0]
        s.close()
        return ip
    except Exception:
        return "127.0.0.1"

def find_available_port(start_port: int = 8000, max_attempts: int = 50) -> int:
    for port in range(start_port, start_port + max_attempts):
        with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:
            try:
                s.bind(("0.0.0.0", port))
                return port
            except OSError:
                continue
    return start_port

class APIKeyPool:
    def __init__(self, req_keys: Optional[str] = None):
        raw_keys = req_keys or os.environ.get("GEMINI_API_KEYS", os.environ.get("GEMINI_API_KEY", ""))
        self.keys = [k.strip() for k in raw_keys.split(",") if k.strip()]
        self._iterator = itertools.cycle(self.keys) if self.keys else None

    def get_client_for_key(self, key: str) -> genai.Client:
        return genai.Client(api_key=key)

    def execute_with_retry(self, action_func):
        if not self.keys:
            raise HTTPException(
                status_code=400, 
                detail="未設定 API Key。請點擊右上角「🔑 API Key 管理」輸入 Key。"
            )
        
        last_error = None
        for _ in range(len(self.keys)):
            key = next(self._iterator)
            try:
                client = self.get_client_for_key(key)
                return action_func(client)
            except Exception as e:
                err_str = str(e)
                last_error = e
                if any(kw in err_str for kw in ["401", "UNAUTHENTICATED", "invalid authentication"]):
                    print(f"[KeyPool Warning] Key {key[:8]}... 401 Invalid: {err_str[:80]}")
                    continue
                if any(kw in err_str for kw in ["429", "RESOURCE_EXHAUSTED", "Quota", "rate limit"]):
                    print(f"[KeyPool Warning] Key {key[:8]}... 額度耗盡，自動切換下一組...")
                    continue
                raise HTTPException(status_code=500, detail=f"Gemini API 處理失敗: {err_str[:150]}")
                
        raise HTTPException(
            status_code=400, 
            detail=f"API 呼叫失敗 (錯誤: {str(last_error)[:120]})。"
        )

def call_gemini_with_model_cascade(client: genai.Client, contents: list, response_schema):
    last_exception = None
    for model_name in FLASH_MODELS_CASCADE:
        try:
            config = types.GenerateContentConfig(
                response_mime_type="application/json",
                response_schema=response_schema,
                temperature=0.2
            )
            res = client.models.generate_content(
                model=model_name,
                contents=contents,
                config=config
            )
            print(f"[Model Cascade Success] 採用模型: {model_name}")
            return res.parsed, model_name
        except Exception as e:
            last_exception = e
            print(f"[Model Cascade Warning] 模型 {model_name} 限制: {str(e)[:80]}，嘗試下一階...")
            continue
    raise last_exception

class BigFive(BaseModel):
    openness: int = Field(description="經驗開放性 (1-100)")
    conscientiousness: int = Field(description="盡責性/盡職度 (1-100)")
    extraversion: int = Field(description="外向性 (1-100)")
    agreeableness: int = Field(description="宜人性/親和力 (1-100)")
    emotional_stability: int = Field(description="情緒穩定度 (1-100)")

class JobMatch(BaseModel):
    title: str = Field(description="職缺名稱")
    score: int = Field(description="契合度評分 (1-100)")
    reason: str = Field(description="原因")

class InterviewReport(BaseModel):
    transcript: str = Field(description="完整的語音轉文字逐字稿 (Transcript)")
    candidate_summary: str = Field(description="對話核心摘要")
    communication_style: str = Field(description="溝通風格與語調表達特徵")
    acoustic_observation: str = Field(description="聲學與語速觀察（節奏、停頓、情緒波動）")
    disc_type: str = Field(description="DISC 人格類型")
    big_five: BigFive
    recommended_jobs: List[JobMatch]
    unsuitable_jobs: List[JobMatch]
    management_advice: str
    followup_questions: List[str]

class PureTranscriptResponse(BaseModel):
    transcript: str = Field(description="語音轉譯出的文字內容")

class PreInterviewQuestion(BaseModel):
    category: str = Field(description="提問類別 (例如: 歷練轉換/技能系統/溝通態度/通勤廠區配合)")
    question: str = Field(description="具體建議提問題目")
    purpose: str = Field(description="提問目的與背景分析")
    evaluation_focus: str = Field(description="面試官應觀察與評判的重點")

class PreInterviewAnalysisReport(BaseModel):
    candidate_name: str = Field(description="應徵者姓名 (若無標示請填 未提供)")
    candidate_age: str = Field(description="年齡或出生年 (例如 36歲 或 79年次)")
    education: str = Field(description="最高學歷與學校科系")
    total_experience: str = Field(description="總工作年資 (例如 14~15年)")
    recent_job: str = Field(description="最近工作職稱與公司")
    applied_position: str = Field(description="應徵目標職務與部門")
    overall_match_score: int = Field(description="整體履歷匹配綜合評分 (1-100)")
    match_summary: str = Field(description="整體履歷契合度總評 (2-3句話綜合摘要)")
    strengths: List[str] = Field(description="履歷三大核心優勢與亮點 (條列式)")
    matching_skills: List[str] = Field(description="符合應徵職務所需的技能、證照與學經歷 (條列式)")
    missing_or_gap_skills: List[str] = Field(description="履歷中較弱、未揭露或與職務需求有落差的技能/系統/項目 (條列式)")
    career_transition_analysis: str = Field(description="職涯歷練與轉折分析 (例如經歷轉換、停業與轉型經驗分析)")
    key_risks_and_concerns: List[str] = Field(description="面試前關鍵疑點與潛在風險 (例如通勤地點、職等調適、離職/停業原因等)")
    suggested_questions: List[PreInterviewQuestion] = Field(description="建議面試提問問題清單 (4-6個精準結構化題目)")



def load_records():
    with open(DB_FILE, "r", encoding="utf-8") as f:
        return json.load(f)

def save_records(records):
    with open(DB_FILE, "w", encoding="utf-8") as f:
        json.dump(records, f, ensure_ascii=False, indent=2)

app.mount("/static", StaticFiles(directory="static"), name="static")
app.mount("/audios", StaticFiles(directory=AUDIO_DIR), name="audios")

@app.get("/")
async def get_index():
    return FileResponse("static/index.html")

@app.get("/api/records")
async def get_records():
    return load_records()

# 匯出 Word 評估報告 (.docx)
@app.get("/api/export-docx/{record_id}")
async def export_docx(record_id: str):
    records = load_records()
    rec = next((r for r in records if r["id"] == record_id), None)
    if not rec:
        raise HTTPException(status_code=404, detail="找不到該筆紀錄")

    if rec.get("type") == "pre_analysis" or rec.get("pre_report"):
        if not rec.get("pre_report"):
            raise HTTPException(status_code=400, detail="該筆事前分析紀錄未完成，無法導出 Word 報告")
        cand = rec.get("candidate_name", "應徵者")
        filename = f"事前履歷分析與提問報告_{cand}_{rec['created_at'].replace(' ', '_').replace(':', '-')}.docx"
        output_path = os.path.join(DATA_DIR, f"export_pre_{record_id}.docx")
        generate_pre_interview_report_docx(rec, output_path)
    else:
        if rec.get("status") != "analyzed" or not rec.get("report"):
            raise HTTPException(status_code=400, detail="該筆紀錄尚未完成 AI 特質分析，無法匯出 Word 報告")
        filename = f"面試特質評估報告_{rec['created_at'].replace(' ', '_').replace(':', '-')}.docx"
        output_path = os.path.join(DATA_DIR, f"export_{record_id}.docx")
        generate_interview_report_docx(rec, output_path)

    return FileResponse(
        path=output_path, 
        filename=filename, 
        media_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document"
    )

# 匯出 Excel 評估總表 (.xlsx)
@app.get("/api/export-excel/{record_id}")
async def export_excel(record_id: str):
    records = load_records()
    rec = next((r for r in records if r["id"] == record_id), None)
    if not rec:
        raise HTTPException(status_code=404, detail="找不到該筆紀錄")

    if rec.get("type") == "pre_analysis" or rec.get("pre_report"):
        if not rec.get("pre_report"):
            raise HTTPException(status_code=400, detail="該筆事前分析紀錄未完成，無法導出 Excel 報告")
        cand = rec.get("candidate_name", "應徵者")
        filename = f"事前履歷分析報告_{cand}_{rec['created_at'].replace(' ', '_').replace(':', '-')}.xlsx"
        output_path = os.path.join(DATA_DIR, f"export_pre_{record_id}.xlsx")
        generate_pre_interview_report_excel(rec, output_path)
    else:
        if rec.get("status") != "analyzed" or not rec.get("report"):
            raise HTTPException(status_code=400, detail="該筆紀錄尚未完成 AI 特質分析，無法匯出 Excel 報告")
        cand = rec.get("candidate_name", "應徵者")
        filename = f"面試特質評估報告_{cand}_{rec['created_at'].replace(' ', '_').replace(':', '-')}.xlsx"
        output_path = os.path.join(DATA_DIR, f"export_{record_id}.xlsx")
        generate_interview_report_excel(rec, output_path)

    return FileResponse(
        path=output_path, 
        filename=filename, 
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

# 事前履歷 AI 分析與面試提問生成 API
@app.post("/api/pre-analyze")
async def pre_analyze_resume(
    target_position: str = Form("【彰濱廠區】助理管理師 / 資材部行政專員"),
    resume_text: Optional[str] = Form(None),
    file: Optional[UploadFile] = File(None),
    x_gemini_api_keys: Optional[str] = Header(None)
):
    key_pool = APIKeyPool(x_gemini_api_keys)
    
    contents = []
    
    # 處理上傳檔案 (PDF, Image, TXT)
    if file and file.filename:
        file_bytes = await file.read()
        filename_lower = file.filename.lower()
        mime_type = file.content_type or "application/octet-stream"
        
        if any(ext in filename_lower for ext in ['.pdf', '.png', '.jpg', '.jpeg', '.webp']):
            if 'pdf' in filename_lower or 'pdf' in mime_type:
                mime_type = "application/pdf"
            elif 'png' in filename_lower:
                mime_type = "image/png"
            elif 'jpg' in filename_lower or 'jpeg' in filename_lower:
                mime_type = "image/jpeg"
            elif 'webp' in filename_lower:
                mime_type = "image/webp"
                
            contents.append(types.Part.from_bytes(data=file_bytes, mime_type=mime_type))
        else:
            try:
                text_content = file_bytes.decode("utf-8", errors="ignore")
                if resume_text:
                    resume_text = resume_text + "\n\n[上傳檔案內容]:\n" + text_content
                else:
                    resume_text = text_content
            except Exception:
                pass

    if resume_text and resume_text.strip():
        contents.append(f"【提供之履歷文字與資料】:\n{resume_text.strip()}")

    if not contents:
        raise HTTPException(status_code=400, detail="請上傳履歷檔案 (PDF / 圖片 / 文字檔) 或輸入履歷文字。")

    # 從 Excel 中查詢該職缺之 104 工作內容與條件要求
    matched_jd = ""
    matched_req = ""
    matched_disc = ""
    try:
        from openpyxl import load_workbook
        wb = load_workbook(JOB_SPECS_FILE, data_only=True)
        ws = wb["職缺需求與規格清單"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row and row[1] and str(row[1]).strip() in target_position:
                matched_jd = str(row[2]).strip() if len(row) > 2 and row[2] else ""
                matched_req = str(row[3]).strip() if len(row) > 3 and row[3] else ""
                matched_disc = str(row[4]).strip() if len(row) > 4 and row[4] else ""
                break
    except Exception:
        pass

    prompt = f"""
    你是一位專業資深 HR 人資主管與組織面試專家。
    應徵目標職務與部門為：【{target_position}】。
    
    【公司 104 官方刊登之標準工作內容】:
    {matched_jd or "未特別載明，請依該職務常態標準評估"}

    【公司 104 官方要求之條件與證照規格】:
    {matched_req or "未特別載明，請依該職務常態標準評估"}

    【建議黃金 DISC 型態指標】:
    {matched_disc or "未特別載明"}

    請詳細分析所提供的履歷資料，對照上述 104 工作內容與條件要求，進行面試前的「事前履歷契合度與疑點分析」，並為面試官擬定 4~6 個高針對性、精準結構化的面試提問題目。

    分析重點包含：
    1. 基本背景與學經歷梳理 (包含姓名、年齡、最高學歷、總工作年資、最近工作)。
    2. 對照【104條件要求與工作內容】與履歷之技能符合項目與技能落差 (例如 堆高機證照/科系/ERP/Word/Excel/SOP 等)。
    3. 歷練轉折、空窗、公司停業轉型或異業轉換之合理性與心態調適。
    4. 工作居住地與應徵地點 (例如台中市南屯區通勤彰濱廠區) 之穩定性與潛在風險。
    5. 針對上述疑點與落差，設計 4~6 個結構化面試提問題目，並附上提問目的與面試官觀察評判重點。
    """
    contents.append(prompt)

    def do_pre_analysis(client: genai.Client):
        parsed_res, used_model = call_gemini_with_model_cascade(
            client=client,
            contents=contents,
            response_schema=PreInterviewAnalysisReport
        )
        return parsed_res.model_dump(), used_model

    report_data, used_model = key_pool.execute_with_retry(do_pre_analysis)

    rec_id = str(uuid.uuid4())
    cand_name = report_data.get("candidate_name") or "應徵者"
    new_record = {
        "id": rec_id,
        "type": "pre_analysis",
        "candidate_name": cand_name,
        "target_dept": target_position,
        "created_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "status": "pre_analyzed",
        "used_model": used_model,
        "pre_report": report_data
    }

    records = load_records()
    records.insert(0, new_record)
    save_records(records)

    return new_record


# 📄 獨立工具：將任意 PDF/圖片/履歷檔案完整的內容直接萃取轉換導出成 Excel 試算表 (.xlsx)
@app.post("/api/convert-pdf-to-excel")
async def convert_pdf_to_excel(
    file: UploadFile = File(...),
    x_gemini_api_keys: Optional[str] = Header(None)
):
    key_pool = APIKeyPool(x_gemini_api_keys)
    if not file or not file.filename:
        raise HTTPException(status_code=400, detail="請選擇要轉換的 PDF 履歷或文件檔案。")
    
    file_bytes = await file.read()
    filename_lower = file.filename.lower()
    mime_type = file.content_type or "application/pdf"
    
    if 'pdf' in filename_lower: mime_type = "application/pdf"
    elif 'png' in filename_lower: mime_type = "image/png"
    elif any(ext in filename_lower for ext in ['.jpg', '.jpeg']): mime_type = "image/jpeg"
    
    file_part = types.Part.from_bytes(data=file_bytes, mime_type=mime_type)
    
    prompt = """
    你是一位資料轉換專家。請完整且精準地閱讀這份 PDF / 圖片文件履歷的【所有原始文字、表格內容與細節】。
    請不要進行任何裁減、刪減或簡化摘要，將這份文件的全部內容整理轉換為結構化的 JSON 資料，包含：
    1. candidate_name: 個人姓名/標題
    2. basic_info: 基本資料 (性別、年齡/出生年、聯絡電話、Email、居住地、通訊地址、希望職缺等)
    3. education: 學歷紀錄 (學校名稱、科系、學位、起訖年月)
    4. experience: 完整工作經歷列表 (公司名稱、職稱、任職時間、部門、工作內容詳細條列與成就)
    5. skills_and_certificates: 專業技能、外語能力、專業證照與軟體系統工具
    6. autobiography: 完整自傳/自我介紹內容
    7. raw_full_text: 文件的完整逐字原始文字內容 (確保無遺漏)
    """

    class PDFExtractedData(BaseModel):
        candidate_name: str = Field(description="姓名")
        basic_info: str = Field(description="個人基本資料與聯絡方式")
        education: str = Field(description="完整學歷紀錄")
        experience: str = Field(description="完整工作經歷與職責內容")
        skills_and_certificates: str = Field(description="專業技能、證照與語言")
        autobiography: str = Field(description="自傳全內文")
        raw_full_text: str = Field(description="文件完整原始文字紀錄")

    def do_convert(client: genai.Client):
        parsed_res, used_model = call_gemini_with_model_cascade(
            client=client,
            contents=[file_part, prompt],
            response_schema=PDFExtractedData
        )
        return parsed_res.model_dump(), used_model

    extracted, used_model = key_pool.execute_with_retry(do_convert)

    # 建立 Excel 活頁簿 (全單一單格、絕不合併欄位)
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "PDF轉譯資料列表"
    ws.views.sheetView[0].showGridLines = True

    cand = extracted.get("candidate_name", "履歷文件")
    
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'), right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'), bottom=Side(style='thin', color='D9D9D9')
    )

    # 1. 頂部標題列 (不跨欄合併，A1放置)
    ws["A1"] = f"📄 PDF 履歷轉譯 - {cand}"
    ws["A1"].font = Font(name="微軟正黑體", size=14, bold=True, color="1F4E78")
    ws["B1"] = f"轉譯時間: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws["B1"].font = Font(name="微軟正黑體", size=10, color="595959")
    ws.row_dimensions[1].height = 30

    # 2. 表頭標籤列
    ws["A2"] = "資料區段與大綱"
    ws["B2"] = "詳細內容與條列細節 (可自由複製/篩選/刪除)"
    for col_ref in ["A2", "B2"]:
        cell = ws[col_ref]
        cell.font = Font(name="微軟正黑體", size=11, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
    ws.row_dimensions[2].height = 28

    sections = [
        ("應徵者姓名", extracted.get("candidate_name", "")),
        ("個人基本資料與聯絡資訊", extracted.get("basic_info", "")),
        ("完整學歷背景", extracted.get("education", "")),
        ("工作經歷與職掌條列", extracted.get("experience", "")),
        ("專業技能與證照", extracted.get("skills_and_certificates", "")),
        ("完整自傳內容", extracted.get("autobiography", "")),
        ("PDF 原始全文字內容備份", extracted.get("raw_full_text", ""))
    ]

    r_idx = 3
    for title, content in sections:
        cell_a = ws.cell(row=r_idx, column=1, value=title)
        cell_a.font = Font(name="微軟正黑體", size=10, bold=True, color="1F4E78")
        cell_a.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        cell_a.alignment = Alignment(horizontal="left", vertical="top")
        cell_a.border = thin_border

        cell_b = ws.cell(row=r_idx, column=2, value=content)
        cell_b.font = Font(name="微軟正黑體", size=10)
        cell_b.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        cell_b.border = thin_border

        # 依內容行數彈性調整列高，絕不合併任何列或欄
        line_count = len(str(content).split("\n"))
        calculated_height = max(28, min(200, line_count * 18))
        ws.row_dimensions[r_idx].height = calculated_height
        
        r_idx += 1

    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 90

    out_file = os.path.join(DATA_DIR, f"pdf_converted_{uuid.uuid4().hex[:8]}.xlsx")
    wb.save(out_file)

    clean_name = file.filename.replace(".pdf", "").replace(".PDF", "")
    return FileResponse(
        path=out_file,
        filename=f"{clean_name}_完整內容.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

# 刪除紀錄
@app.delete("/api/records/{record_id}")
async def delete_record(record_id: str):
    records = load_records()
    rec = next((r for r in records if r["id"] == record_id), None)
    if not rec:
        raise HTTPException(status_code=404, detail="找不到該筆紀錄")

    if "audio_path" in rec and os.path.exists(rec["audio_path"]):
        try: os.remove(rec["audio_path"])
        except: pass

    export_path = os.path.join(DATA_DIR, f"export_{record_id}.docx")
    if os.path.exists(export_path):
        try: os.remove(export_path)
        except: pass

    records = [r for r in records if r["id"] != record_id]
    save_records(records)

    return {"message": "紀錄已成功刪除", "remaining_count": len(records)}

# 1. 開始新的面試 Session (支援應徵者姓名與自訂檔名)
@app.post("/api/start-session")
async def start_session(
    target_dept: str = Query("資材部 (現場助理工程師/工程師/行政)"),
    candidate_name: str = Query("未具名")
):
    rec_id = str(uuid.uuid4())
    time_str = datetime.now().strftime("%Y%m%d_%H%M%S")
    safe_name = candidate_name.strip().replace(" ", "_") or "應徵者"
    audio_filename = f"面試錄音_{safe_name}_{time_str}.webm"
    saved_path = os.path.join(AUDIO_DIR, audio_filename)

    with open(saved_path, "wb") as f:
        pass

    new_record = {
        "id": rec_id,
        "candidate_name": safe_name,
        "target_dept": target_dept,
        "audio_url": f"/audios/{audio_filename}",
        "audio_path": saved_path,
        "created_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "transcript": "",
        "chunks_count": 0,
        "status": "recording",
        "used_model": None,
        "report": None
    }

    records = load_records()
    records.insert(0, new_record)
    save_records(records)

    return new_record

# 2. 每 5 分鐘背景拋送 Chunk 追加
@app.post("/api/append-chunk")
async def append_chunk(
    session_id: str = Form(...),
    file: UploadFile = File(...),
    x_gemini_api_keys: Optional[str] = Header(None)
):
    key_pool = APIKeyPool(x_gemini_api_keys)
    records = load_records()
    rec = next((r for r in records if r["id"] == session_id), None)
    if not rec:
        raise HTTPException(status_code=404, detail="Session 不存在")

    chunk_bytes = await file.read()
    if not chunk_bytes:
        return rec

    with open(rec["audio_path"], "ab") as f:
        f.write(chunk_bytes)

    mime_type = file.content_type or "audio/webm"

    def do_transcribe_chunk(client: genai.Client):
        audio_part = types.Part.from_bytes(data=chunk_bytes, mime_type=mime_type)
        prompt = """請精準轉譯這段錄音內容為繁體中文逐字稿。
請自動分析對話中不同人的聲音音色與語調，進行講者分離 (Speaker Diarization)，明確標註講者身分：
【面試官】: ...
【求職者】: ...
格式請一律採用「【面試官】: 內容」與「【求職者】: 內容」交替呈現。"""
        parsed_res, used_model = call_gemini_with_model_cascade(
            client=client,
            contents=[audio_part, prompt],
            response_schema=PureTranscriptResponse
        )
        return parsed_res.transcript, used_model

    try:
        chunk_text, used_model = key_pool.execute_with_retry(do_transcribe_chunk)
        if chunk_text and chunk_text.strip():
            existing = rec["transcript"].strip()
            rec["transcript"] = (existing + "\n" + chunk_text.strip()).strip()
            rec["used_model"] = used_model
    except Exception as e:
        print(f"[Chunk Warning] 分段轉寫錯誤: {e}")

    rec["chunks_count"] = rec.get("chunks_count", 0) + 1
    save_records(records)

    return rec

# 3. 結束 Session 並發起分析
@app.post("/api/finish-session/{session_id}")
async def finish_session(
    session_id: str,
    analyze: bool = Query(True),
    x_gemini_api_keys: Optional[str] = Header(None)
):
    key_pool = APIKeyPool(x_gemini_api_keys)
    records = load_records()
    rec = next((r for r in records if r["id"] == session_id), None)
    if not rec:
        raise HTTPException(status_code=404, detail="Session 不存在")

    if not analyze:
        rec["status"] = "transcribed"
        save_records(records)
        return rec

    with open(rec["audio_path"], "rb") as f:
        full_audio_bytes = f.read()

    ext = os.path.splitext(rec["audio_path"])[1]
    mime_type = "audio/webm" if "webm" in ext else "audio/wav"
    target_dept = rec.get("target_dept", "資材部 (現場助理工程師/工程師/行政)")

    def do_final_analysis(client: genai.Client):
        contents = []
        if len(full_audio_bytes) <= 20 * 1024 * 1024:
            contents.append(types.Part.from_bytes(data=full_audio_bytes, mime_type=mime_type))

        prompt = f"""
        你是一位頂尖組織心理學家與 HR 顧問。
        應徵目標部門與職務為：【{target_dept}】。

        （特別評估注意事項：若目標為資材部現場助理工程師/工程師，黃金 DISC 為 C型(分析) + S型(穩健) CS型，著重高精準度、料號與SOP細節敏感度；若為資材行政，著重 S型(支援) + C型(分析) SC型）。

        這段長時間面試對話的完整逐字稿為：
        "{rec['transcript']}"

        請綜合對話內容與聲音風格，進行全方位特質與【{target_dept}】職務適性評估報告。
        """
        contents.append(prompt)

        parsed_res, used_model = call_gemini_with_model_cascade(
            client=client,
            contents=contents,
            response_schema=InterviewReport
        )
        return parsed_res.model_dump(), used_model

    report_data, used_model = key_pool.execute_with_retry(do_final_analysis)

    rec["status"] = "analyzed"
    rec["used_model"] = used_model
    if not report_data.get("transcript"):
        report_data["transcript"] = rec["transcript"]
    rec["report"] = report_data

    save_records(records)
    return rec

@app.post("/api/analyze-record/{record_id}")
async def analyze_existing_record(
    record_id: str,
    target_dept: Optional[str] = Query(None),
    x_gemini_api_keys: Optional[str] = Header(None)
):
    key_pool = APIKeyPool(x_gemini_api_keys)
    records = load_records()
    rec = next((r for r in records if r["id"] == record_id), None)
    if not rec:
        raise HTTPException(status_code=404, detail="找不到該筆面試紀錄")

    if target_dept:
        rec["target_dept"] = target_dept
    dept_name = rec.get("target_dept", "資材部 (現場助理工程師/工程師/行政)")

    saved_path = rec["audio_path"]
    if not os.path.exists(saved_path):
        raise HTTPException(status_code=404, detail="音訊檔案已遺失")

    with open(saved_path, "rb") as f:
        audio_bytes = f.read()

    ext = os.path.splitext(saved_path)[1]
    mime_type = "audio/webm" if "webm" in ext else "audio/wav"

    def do_analyze(client: genai.Client):
        contents = []
        if len(audio_bytes) <= 20 * 1024 * 1024:
            contents.append(types.Part.from_bytes(data=audio_bytes, mime_type=mime_type))

        prompt = f"""
        你是一位頂尖組織心理學家與 HR 顧問。
        應徵目標部門與職務為：【{dept_name}】。

        （特別評估注意事項：若目標為資材部現場助理工程師/工程師，黃金 DISC 為 C型(分析) + S型(穩健) CS型，著重高精準度、料號與SOP細節敏感度；若為資材行政，著重 S型(支援) + C型(分析) SC型）。

        這段錄音的語音轉寫文字為：
        "{rec['transcript']}"

        請結合聲音風格與對話內容，針對【{dept_name}】進行深度特質與職務契合度評估。
        """
        contents.append(prompt)

        parsed_res, used_model = call_gemini_with_model_cascade(
            client=client,
            contents=contents,
            response_schema=InterviewReport
        )
        return parsed_res.model_dump(), used_model

    report_data, used_model = key_pool.execute_with_retry(do_analyze)

    rec["status"] = "analyzed"
    rec["used_model"] = used_model
    rec["report"] = report_data
    save_records(records)

    return rec

# 4. 斷網補救：讀取本機音檔重新呼叫 AI 進行轉寫與特質分析 API
@app.post("/api/re-transcribe/{record_id}")
async def re_transcribe_record(
    record_id: str,
    candidate_name: Optional[str] = Query(None),
    x_gemini_api_keys: Optional[str] = Header(None)
):
    key_pool = APIKeyPool(x_gemini_api_keys)
    records = load_records()
    rec = next((r for r in records if r["id"] == record_id), None)
    if not rec:
        raise HTTPException(status_code=404, detail="找不到該筆紀錄")

    if candidate_name and candidate_name.strip():
        rec["candidate_name"] = candidate_name.strip()

    saved_path = rec.get("audio_path")
    if not saved_path or not os.path.exists(saved_path):
        raise HTTPException(status_code=404, detail="本機音訊檔案不存在，無法進行重新辨識")

    with open(saved_path, "rb") as f:
        full_audio_bytes = f.read()

    if not full_audio_bytes or len(full_audio_bytes) < 100:
        raise HTTPException(status_code=400, detail="本機音檔為空白檔或容量過小，無有效聲音")

    ext = os.path.splitext(saved_path)[1].lower()
    mime_type = "audio/webm" if "webm" in ext else ("audio/mp3" if "mp3" in ext else "audio/wav")
    target_dept = rec.get("target_dept", "資材部 (現場助理工程師/工程師/行政)")

    # Step 1: 重新語音轉文字 (Transcript)
    def do_re_transcribe(client: genai.Client):
        audio_part = types.Part.from_bytes(data=full_audio_bytes, mime_type=mime_type)
        prompt = """請詳細且完整地將這整段面試錄音內容轉換為繁體中文逐字稿 (Transcript)。
特別注意事項：
這是雙人對話錄音，請發揮 AI 的音色與對話脈絡辨識能力，進行「講者分離 (Speaker Diarization)」，將對話明確區分為：
【面試官】: ...
【求職者】: ...
格式請一律採用「【面試官】: 內容」與「【求職者】: 內容」分行呈現，切勿將兩人聲音混在一起。"""
        parsed_res, used_model = call_gemini_with_model_cascade(
            client=client,
            contents=[audio_part, prompt],
            response_schema=PureTranscriptResponse
        )
        return parsed_res.transcript, used_model

    transcript_text, used_model = key_pool.execute_with_retry(do_re_transcribe)
    rec["transcript"] = transcript_text.strip() if transcript_text else "無對話"
    rec["used_model"] = used_model

    # Step 2: 進行 AI 面試特質評估
    def do_re_analysis(client: genai.Client):
        contents = [
            types.Part.from_bytes(data=full_audio_bytes, mime_type=mime_type),
            f"""
            你是一位頂尖組織心理學家與 HR 顧問。
            應徵目標部門與職務為：【{target_dept}】。

            請結合對話聲音風格與文字內容（逐字稿：{rec['transcript'][:2000]}...），
            進行全方位特質與【{target_dept}】職務適性評估報告。
            """
        ]
        parsed_res, used_model = call_gemini_with_model_cascade(
            client=client,
            contents=contents,
            response_schema=InterviewReport
        )
        return parsed_res.model_dump(), used_model

    report_data, used_model = key_pool.execute_with_retry(do_re_analysis)
    rec["status"] = "analyzed"
    rec["used_model"] = used_model
    if not report_data.get("transcript"):
        report_data["transcript"] = rec["transcript"]
    rec["report"] = report_data

    save_records(records)
    return rec

if __name__ == "__main__":
    import uvicorn
    import webbrowser
    local_ip = get_local_ip()
    port = find_available_port(8000)
    url = f"http://localhost:{port}"
    lan_url = f"http://{local_ip}:{port}"
    
    print("========================================================")
    print("  AI 面試語音與資材特質分析系統 - 伺服器啟動成功！")
    print(f"  -> 本機網址: {url}")
    print(f"  -> 區網網址: {lan_url}")
    print("========================================================")
    
    try:
        webbrowser.open(url)
    except:
        pass
        
    uvicorn.run(app, host="0.0.0.0", port=port)
