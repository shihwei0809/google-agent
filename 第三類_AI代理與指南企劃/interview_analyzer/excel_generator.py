import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def create_styled_cell(ws, cell_ref, value, font_size=11, bold=False, color="000000", fill_color=None, align_h="left", align_v="center", wrap_text=True):
    cell = ws[cell_ref]
    cell.value = value
    cell.font = Font(name="微軟正黑體", size=font_size, bold=bold, color=color)
    
    alignment_kwargs = {"horizontal": align_h, "vertical": align_v, "wrap_text": wrap_text}
    cell.alignment = Alignment(**alignment_kwargs)
    
    if fill_color:
        cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
        
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )
    cell.border = thin_border
    return cell

def generate_pre_interview_report_excel(record_data: dict, output_path: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "事前履歷分析報告"
    ws.views.sheetView[0].showGridLines = True
    
    pre = record_data.get("pre_report", {})
    cand_name = pre.get("candidate_name", record_data.get("candidate_name", "應徵者"))
    dept = record_data.get("target_dept", "未指定職務")
    
    # 標題欄
    ws.merge_cells("A1:F1")
    title_cell = ws["A1"]
    title_cell.value = f"📋 應徵者事前履歷評估與提問建議總表 - {cand_name}"
    title_cell.font = Font(name="微軟正黑體", size=16, bold=True, color="FFFFFF")
    title_cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 40

    # 副標題資訊
    headers = [
        ("應徵姓名", cand_name),
        ("應徵目標職務", dept),
        ("評估時間", record_data.get("created_at", "")),
        ("整體契合度評分", f"{pre.get('overall_match_score', 0)} 分")
    ]
    
    ws.merge_cells("A2:B2")
    ws["A2"] = "個人基本資訊與契合度總評"
    ws["A2"].font = Font(name="微軟正黑體", size=12, bold=True, color="1F4E78")
    ws["A2"].fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    ws.row_dimensions[2].height = 25

    # 基本資料表格
    info_table = [
        ["應徵者姓名", cand_name, "年齡/出生年", pre.get("candidate_age", "未標示"), "最高學歷", pre.get("education", "未標示")],
        ["總工作年資", pre.get("total_experience", "未標示"), "最近工作職稱", pre.get("recent_job", "未標示"), "契合度評分", f"{pre.get('overall_match_score', 0)} / 100 分"]
    ]
    
    current_row = 3
    for row_idx, row_data in enumerate(info_table):
        ws.row_dimensions[current_row].height = 24
        cols = ["A", "B", "C", "D", "E", "F"]
        for c_idx in range(6):
            cell_ref = f"{cols[c_idx]}{current_row}"
            val = row_data[c_idx]
            is_header = (c_idx % 2 == 0)
            fill = "F2F2F2" if is_header else "FFFFFF"
            font_color = "1F4E78" if is_header else "000000"
            create_styled_cell(ws, cell_ref, val, font_size=10, bold=is_header, color=font_color, fill_color=fill, align_h="center" if is_header else "left")
        current_row += 1

    current_row += 1
    # 綜合摘要
    ws.merge_cells(f"A{current_row}:F{current_row}")
    create_styled_cell(ws, f"A{current_row}", "💡 整體履歷契合度綜合總評", font_size=11, bold=True, color="FFFFFF", fill_color="2F5597", align_h="left")
    ws.row_dimensions[current_row].height = 24
    current_row += 1
    
    ws.merge_cells(f"A{current_row}:F{current_row+1}")
    create_styled_cell(ws, f"A{current_row}", pre.get("match_summary", "無"), font_size=10, align_v="top")
    ws.row_dimensions[current_row].height = 22
    ws.row_dimensions[current_row+1].height = 22
    current_row += 2

    current_row += 1
    # 核心亮點與技能分析 (分欄)
    ws.merge_cells(f"A{current_row}:C{current_row}")
    create_styled_cell(ws, f"A{current_row}", "🌟 核心優勢與亮點", font_size=11, bold=True, color="FFFFFF", fill_color="385723", align_h="center")
    ws.merge_cells(f"D{current_row}:F{current_row}")
    create_styled_cell(ws, f"D{current_row}", "⚠️ 技能落差或疑點標註", font_size=11, bold=True, color="FFFFFF", fill_color="C65911", align_h="center")
    ws.row_dimensions[current_row].height = 24
    current_row += 1

    strengths = pre.get("strengths", [])
    gaps = pre.get("missing_or_gap_skills", [])
    max_len = max(len(strengths), len(gaps), 3)

    for i in range(max_len):
        s_text = f"• {strengths[i]}" if i < len(strengths) else ""
        g_text = f"• {gaps[i]}" if i < len(gaps) else ""
        
        ws.merge_cells(f"A{current_row}:C{current_row}")
        create_styled_cell(ws, f"A{current_row}", s_text, font_size=10, fill_color="F4F9F1", align_v="top")
        
        ws.merge_cells(f"D{current_row}:F{current_row}")
        create_styled_cell(ws, f"D{current_row}", g_text, font_size=10, fill_color="FCEFEE", align_v="top")
        
        ws.row_dimensions[current_row].height = 28
        current_row += 1

    current_row += 1
    # 結構化面試建議提問表
    ws.merge_cells(f"A{current_row}:F{current_row}")
    create_styled_cell(ws, f"A{current_row}", "🎯 建議結構化面試提問清單 (針對缺口與疑點)", font_size=11, bold=True, color="FFFFFF", fill_color="1F4E78", align_h="left")
    ws.row_dimensions[current_row].height = 26
    current_row += 1

    headers_q = [("類別", "A"), ("具體建議提問題目", "B:C"), ("提問目的與背景", "D"), ("面試官評判觀察重點", "E:F")]
    ws.row_dimensions[current_row].height = 24
    create_styled_cell(ws, f"A{current_row}", "項次/類別", font_size=10, bold=True, color="1F4E78", fill_color="D9E1F2", align_h="center")
    ws.merge_cells(f"B{current_row}:C{current_row}")
    create_styled_cell(ws, f"B{current_row}", "具體建議面試提問題目", font_size=10, bold=True, color="1F4E78", fill_color="D9E1F2", align_h="center")
    create_styled_cell(ws, f"D{current_row}", "提問目的與背景分析", font_size=10, bold=True, color="1F4E78", fill_color="D9E1F2", align_h="center")
    ws.merge_cells(f"E{current_row}:F{current_row}")
    create_styled_cell(ws, f"E{current_row}", "面試官觀察與評判重點", font_size=10, bold=True, color="1F4E78", fill_color="D9E1F2", align_h="center")
    current_row += 1

    questions = pre.get("suggested_questions", [])
    for idx, q in enumerate(questions, 1):
        ws.row_dimensions[current_row].height = 42
        cat = q.get("category", "")
        q_text = q.get("question", "")
        purpose = q.get("purpose", "")
        eval_focus = q.get("evaluation_focus", "")
        
        create_styled_cell(ws, f"A{current_row}", f"{idx}. {cat}", font_size=10, bold=True, align_h="center", align_v="center")
        
        ws.merge_cells(f"B{current_row}:C{current_row}")
        create_styled_cell(ws, f"B{current_row}", q_text, font_size=10, align_v="center")
        
        create_styled_cell(ws, f"D{current_row}", purpose, font_size=9, color="595959", align_v="center")
        
        ws.merge_cells(f"E{current_row}:F{current_row}")
        create_styled_cell(ws, f"E{current_row}", eval_focus, font_size=9, color="333333", align_v="center")
        
        current_row += 1

    # 設定欄寬
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 28
    ws.column_dimensions['C'].width = 28
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 25
    ws.column_dimensions['F'].width = 25

    wb.save(output_path)
    return output_path

def generate_interview_report_excel(record_data: dict, output_path: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "面試特質評估報告"
    ws.views.sheetView[0].showGridLines = True

    report = record_data.get("report", {})
    cand_name = record_data.get("candidate_name", "應徵者")
    dept = record_data.get("target_dept", "資材部")
    created_at = record_data.get("created_at", "")

    # 標題
    ws.merge_cells("A1:E1")
    title_cell = ws["A1"]
    title_cell.value = f"🎙️ AI 面試對話與特質評估總表 - {cand_name}"
    title_cell.font = Font(name="微軟正黑體", size=16, bold=True, color="FFFFFF")
    title_cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 40

    # 摘要區
    summary_data = [
        ["應徵者姓名", cand_name, "目標部門/職務", dept],
        ["面試對話時間", created_at, "DISC 人格類型", report.get("disc_type", "未標示")]
    ]
    
    current_row = 2
    for r in summary_data:
        ws.row_dimensions[current_row].height = 24
        create_styled_cell(ws, f"A{current_row}", r[0], font_size=10, bold=True, color="1F4E78", fill_color="F2F2F2", align_h="center")
        create_styled_cell(ws, f"B{current_row}", r[1], font_size=10, align_h="left")
        create_styled_cell(ws, f"C{current_row}", r[2], font_size=10, bold=True, color="1F4E78", fill_color="F2F2F2", align_h="center")
        ws.merge_cells(f"D{current_row}:E{current_row}")
        create_styled_cell(ws, f"D{current_row}", r[3], font_size=10, bold=True if r[2]=="DISC 人格類型" else False, color="C65911" if r[2]=="DISC 人格類型" else "000000", align_h="left")
        current_row += 1

    current_row += 1
    # 逐字稿與對話摘要
    ws.merge_cells(f"A{current_row}:E{current_row}")
    create_styled_cell(ws, f"A{current_row}", "💬 對話摘要與溝通特徵分析", font_size=11, bold=True, color="FFFFFF", fill_color="2F5597", align_h="left")
    ws.row_dimensions[current_row].height = 24
    current_row += 1

    sections = [
        ("對話核心摘要", report.get("candidate_summary", "無")),
        ("溝通風格與語調表達", report.get("communication_style", "無")),
        ("聲學與語速觀察", report.get("acoustic_observation", "無")),
        ("管理與帶領建議", report.get("management_advice", "無"))
    ]

    for sec_title, sec_content in sections:
        create_styled_cell(ws, f"A{current_row}", sec_title, font_size=10, bold=True, color="1F4E78", fill_color="D9E1F2", align_h="center", align_v="center")
        ws.merge_cells(f"B{current_row}:E{current_row}")
        create_styled_cell(ws, f"B{current_row}", sec_content, font_size=10, align_v="center")
        ws.row_dimensions[current_row].height = 36
        current_row += 1

    current_row += 1
    # 逐字稿完整呈現
    ws.merge_cells(f"A{current_row}:E{current_row}")
    create_styled_cell(ws, f"A{current_row}", "📝 語音轉文字完整逐字稿 (Transcript)", font_size=11, bold=True, color="FFFFFF", fill_color="595959", align_h="left")
    ws.row_dimensions[current_row].height = 24
    current_row += 1

    transcript_text = record_data.get("transcript") or report.get("transcript") or "無逐字稿"
    ws.merge_cells(f"A{current_row}:E{current_row+2}")
    create_styled_cell(ws, f"A{current_row}", transcript_text, font_size=9, color="333333", align_v="top")
    for r_idx in range(current_row, current_row+3):
        ws.row_dimensions[r_idx].height = 25

    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 25

    wb.save(output_path)
    return output_path
