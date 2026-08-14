import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.utils.units import pixels_to_EMU
from io import BytesIO
import qrcode
import re
from collections import defaultdict
import os
import shutil
import sys

excel_file = r"d:\GOOGLE ANGET\isotank bacode.xlsx"
backup_file = r"d:\GOOGLE ANGET\isotank bacode_分頁整理版.xlsx"

QR_SIZE_PX = 114  # 3 cm
ROW_HEIGHT_PT = 115
COL_OFFSET_EMU = pixels_to_EMU(36)
ROW_OFFSET_EMU = pixels_to_EMU(20)
SIZE_EMU = pixels_to_EMU(QR_SIZE_PX)

def classify_code(code):
    c = code.upper().strip()
    m = re.match(r'E(\d+)', c)
    if m:
        num = int(m.group(1))
        if num < 20: return 'E0X~E1X (13個)'
        elif num < 40: return 'E2X~E3X (12個)'
        elif num < 50: return 'E4X (8個)'
        elif num < 60: return 'E5X (10個)'
        elif num < 70: return 'E6X (10個)'
        elif num < 80: return 'E7X (10個)'
        elif num < 90: return 'E8X (10個)'
        elif num < 100: return 'E9X (10個)'
        elif num < 200: return 'E10X (10個)'
        elif 300 <= num < 310: return 'E30X (9個)'
        elif 310 <= num < 320: return 'E31X (9個)'
        elif 320 <= num <= 329: return 'E32X (10個)'
        elif 330 <= num < 340: return 'E33X與S系列 (11個)'
        elif 340 <= num < 350: return 'E34X (8個)'
        elif 350 <= num < 360: return 'E35X (8個)'
        elif 360 <= num < 370: return 'E36X (8個)'
        else: return f'E{num//10}X'
    elif c.startswith('0') or c.startswith('1'):
        return 'E0X~E1X (13個)'
    elif c.startswith('S'):
        return 'E33X與S系列 (11個)'
    else:
        return '新增標籤'

def create_centered_qr_image(text_data, target_row, target_col):
    qr = qrcode.QRCode(
        version=1,
        error_correction=qrcode.constants.ERROR_CORRECT_M,
        box_size=10,
        border=2,
    )
    qr.add_data(text_data)
    qr.make(fit=True)
    img_pil = qr.make_image(fill_color="black", back_color="white").resize((QR_SIZE_PX, QR_SIZE_PX))
    
    buf = BytesIO()
    img_pil.save(buf, format="PNG")
    buf.seek(0)
    
    img = Image(buf)
    img.width = QR_SIZE_PX
    img.height = QR_SIZE_PX
    
    marker = AnchorMarker(
        col=target_col - 1,
        colOff=COL_OFFSET_EMU,
        row=target_row - 1,
        rowOff=ROW_OFFSET_EMU
    )
    img.anchor = OneCellAnchor(_from=marker, ext=XDRPositiveSize2D(SIZE_EMU, SIZE_EMU))
    return img

def parse_input_sources(inputs):
    extracted_items = []
    
    for inp in inputs:
        inp_str = str(inp).strip().strip('"').strip("'")
        if not inp_str: continue
        
        # Check if input is a text/csv file path
        if os.path.exists(inp_str) and (inp_str.endswith('.txt') or inp_str.endswith('.csv')):
            print(f"[批次檔案] 正在讀取文字檔/CSV: {inp_str}")
            with open(inp_str, 'r', encoding='utf-8', errors='ignore') as f:
                content = f.read()
                codes = re.split(r'[\s,\t\r\n]+', content)
                for c in codes:
                    if c.strip() and c.strip().upper() != '編號':
                        extracted_items.append({'code': c.strip().upper(), 'serial': '', 'cap': ''})
        # Check if input is an Excel file path
        elif os.path.exists(inp_str) and (inp_str.endswith('.xlsx') or inp_str.endswith('.xls')):
            print(f"[批次檔案] 正在讀取 Excel 檔案: {inp_str}")
            wb_inp = openpyxl.load_workbook(inp_str, data_only=True)
            for sheet_name in wb_inp.sheetnames:
                ws_inp = wb_inp[sheet_name]
                for r in range(1, ws_inp.max_row + 1):
                    # Check if there is code in Col 3 (like 上品新槽車) or any col
                    code_val = ws_inp.cell(r, 3).value or ws_inp.cell(r, 1).value
                    serial_val = ws_inp.cell(r, 4).value if ws_inp.max_column >= 4 else None
                    cap_val = ws_inp.cell(r, 5).value if ws_inp.max_column >= 5 else None
                    
                    for c in range(1, ws_inp.max_column + 1):
                        val = ws_inp.cell(r, c).value
                        if val:
                            val_s = str(val).strip()
                            if re.match(r'^[E0-9S][0-9A-Z]{2,9}$', val_s, re.IGNORECASE) and val_s.upper() != '編號':
                                serial_s = str(serial_val).strip() if (serial_val and c == 3) else ''
                                cap_s = str(cap_val).strip() if (cap_val and c == 3) else ''
                                extracted_items.append({'code': val_s.upper(), 'serial': serial_s, 'cap': cap_s})
        else:
            codes = re.split(r'[\s,\t\r\n]+', inp_str)
            for c in codes:
                if c.strip() and c.strip().upper() != '編號':
                    extracted_items.append({'code': c.strip().upper(), 'serial': '', 'cap': ''})
                    
    return extracted_items

def batch_add_codes(new_items_list):
    if not os.path.exists(excel_file):
        print(f"[錯誤] 找不到主 Excel 檔案: {excel_file}")
        return

    try:
        wb = openpyxl.load_workbook(excel_file)
    except PermissionError:
        print("\n[錯誤] Excel 檔案目前正被其他程式開啟中，請先關閉 Excel 檔後再按 Enter 重新執行！")
        return

    # Read existing codes in oracle_sync_data (Sheet 1)
    ws_sync = wb.worksheets[0]
    sync_codes = set()
    for r in range(1, ws_sync.max_row + 1):
        val = ws_sync.cell(r, 1).value
        if val: sync_codes.add(str(val).strip().upper())

    # Read existing codes across category sheets
    category_codes = set()
    for sheet_name in wb.sheetnames:
        if sheet_name.startswith('oracle'): continue
        ws = wb[sheet_name]
        for r in range(3, ws.max_row + 1):
            valA = ws.cell(r, 1).value
            valC = ws.cell(r, 3).value
            if valA: category_codes.add(str(valA).strip().upper())
            if valC: category_codes.add(str(valC).strip().upper())

    added_summary = defaultdict(list)
    added_to_sync_count = 0
    skipped_count = 0
    
    # Process each item
    for item_info in new_items_list:
        clean_code = item_info['code']
        serial_info = item_info.get('serial', '')
        cap_info = item_info.get('cap', '')
        
        if not clean_code: continue

        # 1. Check & Add to oracle_sync_data (Sheet 1) if missing
        if clean_code not in sync_codes:
            new_r = ws_sync.max_row + 1
            ws_sync.cell(new_r, 1, value=clean_code) # Col A: 槽號
            if serial_info:
                ws_sync.cell(new_r, 3, value=serial_info) # Col C: 罐號
            if cap_info:
                ws_sync.cell(new_r, 4, value=cap_info) # Col D: 容量
            sync_codes.add(clean_code)
            added_to_sync_count += 1

        # 2. Check & Add to Category Sheet
        if clean_code in category_codes:
            skipped_count += 1
            continue
            
        g_name = classify_code(clean_code)
        
        # If sheet doesn't exist, create it
        if g_name not in wb.sheetnames:
            ws = wb.create_sheet(title=g_name)
            ws.views.sheetView[0].showGridLines = True
            
            title_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
            header_fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
            font_title = Font(name="微軟正黑體", size=14, bold=True, color="FFFFFF")
            font_header = Font(name="微軟正黑體", size=11, bold=True, color="FFFFFF")
            center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
            ws.merge_cells("A1:D1")
            ws.row_dimensions[1].height = 36
            c_t = ws.cell(1, 1, value=f"ISOTANK 條碼標籤清單 (3cm大條碼置中版) — 【{g_name}】")
            c_t.font = font_title; c_t.fill = title_fill; c_t.alignment = center_align
            for col in range(1, 5): ws.cell(1, col).fill = title_fill
            
            ws.row_dimensions[2].height = 28
            headers = ["ISOTANK 編號", "QR Code 條碼 (3cm)", "ISOTANK 編號", "QR Code 條碼 (3cm)"]
            for c_idx, h_text in enumerate(headers, 1):
                ch = ws.cell(2, c_idx, value=h_text)
                ch.font = font_header; ch.fill = header_fill; ch.alignment = center_align
                
            ws.column_dimensions['A'].width = 18
            ws.column_dimensions['B'].width = 24
            ws.column_dimensions['C'].width = 18
            ws.column_dimensions['D'].width = 24

        ws = wb[g_name]
        
        # Count existing items in sheet
        item_count = 0
        for r in range(3, ws.max_row + 1):
            if ws.cell(r, 1).value: item_count += 1
            if ws.cell(r, 3).value: item_count += 1
            
        # Determine position
        if item_count % 2 == 0:
            target_row = 3 + (item_count // 2)
            target_col_text = 1
            target_col_qr = 2
        else:
            target_row = 3 + (item_count // 2)
            target_col_text = 3
            target_col_qr = 4

        font_code = Font(name="微軟正黑體", size=13, bold=True, color="1F4E78")
        thin_border = Border(left=Side(style='thin', color='D9D9D9'), right=Side(style='thin', color='D9D9D9'), top=Side(style='thin', color='D9D9D9'), bottom=Side(style='thin', color='D9D9D9'))
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        alt_row_fill = PatternFill(start_color="F2F5F9", end_color="F2F5F9", fill_type="solid")
        white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        current_fill = alt_row_fill if (target_row % 2 == 1) else white_fill

        ws.row_dimensions[target_row].height = ROW_HEIGHT_PT
        
        c_text = ws.cell(target_row, target_col_text, value=clean_code)
        c_text.font = font_code; c_text.fill = current_fill; c_text.border = thin_border; c_text.alignment = center_align
        
        c_qr = ws.cell(target_row, target_col_qr)
        c_qr.fill = current_fill; c_qr.border = thin_border
        
        qr_img = create_centered_qr_image(clean_code, target_row=target_row, target_col=target_col_qr)
        ws.add_image(qr_img)
        
        category_codes.add(clean_code)
        added_summary[g_name].append(clean_code)

    total_added_cat = sum(len(v) for v in added_summary.values())

    if total_added_cat > 0 or added_to_sync_count > 0:
        try:
            wb.save(excel_file)
            shutil.copyfile(excel_file, backup_file)
            
            print("\n========================================================")
            print(f" [成功] 處理完成！成功新增 {total_added_cat} 個分類條碼，同步寫入第1分頁 {added_to_sync_count} 個編號")
            print("========================================================")
            for sheet_cat, codes in added_summary.items():
                print(f"  [分頁] [{sheet_cat}] -> 新增 {len(codes)} 個: {', '.join(codes)}")
            print("========================================================")
            print(f"已成功更新存檔: {excel_file}")
            
        except PermissionError:
            print("\n[警告] Excel 檔案正在被開啟中，請先關閉 Excel 檔後再次執行！")
    else:
        print(f"\n[提示] 所有編號均已存在於檔案中。")

if __name__ == "__main__":
    if len(sys.argv) > 1:
        raw_inputs = sys.argv[1:]
        parsed_items = parse_input_sources(raw_inputs)
        batch_add_codes(parsed_items)
    else:
        print("========================================================")
        print("  ISOTANK 槽車編號【一鍵拖曳/貼上】3cm條碼產生器")
        print("========================================================")
        print(" 提示：")
        print(" 1. 可直接【拖曳 Excel / TXT 檔案】進來此視窗並按 Enter")
        print(" 2. 或直接【貼上/輸入槽車編號】(如 E110 E111) 並按 Enter")
        print("========================================================\n")
        
        inp = input("請拖入檔案或貼上槽車編號後按 [Enter] 執行：\n> ").strip()
        parsed_items = parse_input_sources([inp])
        
        if parsed_items:
            print(f"\n[系統] 偵測到 {len(parsed_items)} 個待處理編號，正在進行雙向寫入與條碼繪製...")
            batch_add_codes(parsed_items)
        else:
            print("[提示] 未輸入有效編號或檔案。")
            
        input("\n處理完成！按 Enter 鍵結束...")
