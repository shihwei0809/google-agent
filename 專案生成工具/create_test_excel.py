import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def build_wb(is_interviewer_version=False):
    wb = openpyxl.Workbook()

    title_font = Font(name="Microsoft JhengHei", size=14, bold=True, color="1F4E78")
    subtitle_font = Font(name="Microsoft JhengHei", size=11, italic=True, color="595959")
    header_font = Font(name="Microsoft JhengHei", size=11, bold=True, color="FFFFFF")
    bold_font = Font(name="Microsoft JhengHei", size=11, bold=True)
    regular_font = Font(name="Microsoft JhengHei", size=11)
    error_font = Font(name="Microsoft JhengHei", size=11, bold=True, color="C00000")

    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    sub_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )

    # -------------------------------------------------------------
    # Sheet 1: 測驗說明
    # -------------------------------------------------------------
    ws_info = wb.active
    ws_info.title = "01_測驗說明"
    ws_info.views.sheetView[0].showGridLines = True

    role_str = "【考生用】" if not is_interviewer_version else "【面試官解答版】"
    ws_info.append([f"勝一化工【彰濱廠區】助理管理師 - Excel 上機實作測驗說明 {role_str}"])
    ws_info.append(["測驗時間：15 分鐘 | 總分：100 分"])
    ws_info.append([])
    ws_info.append(["測驗大綱與說明："])
    ws_info.append(["1. 本測驗共分為三個考題頁籤（Tab）："])
    ws_info.append(["   - 【02_過磅數據對帳】（40分）：考查過磅計算、數字對帳與抓出異常資料能力。"])
    ws_info.append(["   - 【03_T100料號VLOOKUP】（30分）：考查是否會使用 VLOOKUP/XLOOKUP 函數串接料號。"])
    ws_info.append(["   - 【04_包材領用統計】（30分）：考查能否建立樞紐分析表（Pivot Table）做日報統計。"])
    ws_info.append(["2. 測驗完成後，請儲存檔案並通知面試官。"])

    ws_info['A1'].font = title_font
    ws_info['A2'].font = subtitle_font
    ws_info['A4'].font = bold_font

    for row in range(5, 10):
        ws_info[f'A{row}'].font = regular_font

    # -------------------------------------------------------------
    # Sheet 2: 題目1_過磅與數據對帳
    # -------------------------------------------------------------
    ws1 = wb.create_sheet(title="02_過磅數據對帳")
    ws1.views.sheetView[0].showGridLines = True

    ws1.append(["題目 1：進出貨過磅數據計算與異常抓漏（40分）"])
    ws1.append(["任務：1. 請在『淨重(kg)』欄位使用 Excel 公式計算（淨重 = 毛重 - 空重）。"])
    ws1.append(["      2. 檢視下方過磅紀錄，找出 2 筆數據邏輯或計算異常的項目，並在『異常檢查標註』欄位寫下原因。"])
    ws1.append([])

    headers1 = ["單號", "過磅日期", "槽車/車號", "客戶/廠商名稱", "品名", "毛重 (kg)", "空重 (kg)", "淨重 (kg) [請寫公式]", "異常檢查標註"]
    ws1.append(headers1)

    data1 = [
        ["WB2026070101", "2026-07-01", "KAA-1234", "長興材料", "異丙醇(IPA)", 35000, 15000, "", ""],
        ["WB2026070102", "2026-07-01", "KAB-5678", "台積電", "醋酸正丁酯(NBA)", 32000, 14500, "", ""],
        ["WB2026070103", "2026-07-01", "KAC-9012", "聯電", "丙二醇單甲基醚(PM)", 13000, 14800, "", ""],
        ["WB2026070104", "2026-07-01", "KAD-3456", "南亞塑膠", "異丙醇(IPA)", 36500, 15200, "", ""],
        ["WB2026070105", "2026-07-01", "KAE-7890", "日月光", "環己酮(CYC)", 28000, 14000, "", ""],
        ["WB2026070106", "2026-07-01", "KAF-2345", "群創光電", "丙二醇單甲基醚醋酸酯(PMA)", 34000, 41000, "", ""],
        ["WB2026070107", "2026-07-01", "KAG-6789", "友達光電", "異丙醇(IPA)", 33500, 14200, "", ""],
        ["WB2026070108", "2026-07-01", "KAH-0123", "力積電", "醋酸正丁酯(NBA)", 31000, 14100, "", ""],
    ]

    for row in data1:
        ws1.append(row)

    ws1['A1'].font = title_font
    ws1['A2'].font = regular_font
    ws1['A3'].font = regular_font

    for col in range(1, 10):
        cell = ws1.cell(row=5, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for r in range(6, 6 + len(data1)):
        for c in range(1, 10):
            cell = ws1.cell(row=r, column=c)
            cell.font = regular_font
            cell.border = thin_border
            if c in [6, 7]:
                cell.number_format = '#,##0'
                cell.alignment = Alignment(horizontal="right")

    # -------------------------------------------------------------
    # Sheet 3: 題目2_T100料號VLOOKUP
    # -------------------------------------------------------------
    ws2 = wb.create_sheet(title="03_T100料號VLOOKUP")
    ws2.views.sheetView[0].showGridLines = True

    ws2.append(["題目 2：T100 系統料號自動對照與填入（30分）"])
    ws2.append(["任務：請在【每日驗收紀錄表】的『T100系統料號』欄位（Column C），使用 VLOOKUP 或 XLOOKUP 函數帶出右側對照表之標準料號。"])
    ws2.append([])

    headers2_left = ["驗收單號", "包材品名", "T100系統料號 [請用公式]", "驗收數量", "檢驗狀態"]
    headers2_right = ["T100標準料號", "包材品名對照"]

    ws2.cell(row=4, column=1, value="【每日驗收紀錄表】").font = bold_font
    ws2.cell(row=4, column=7, value="【T100料號主檔對照表】").font = bold_font

    for c, h in enumerate(headers2_left, start=1):
        cell = ws2.cell(row=5, column=c, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for c, h in enumerate(headers2_right, start=7):
        cell = ws2.cell(row=5, column=c, value=h)
        cell.font = header_font
        cell.fill = PatternFill(start_color="333399", end_color="333399", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    rec_data = [
        ["RC-20260701", "50加侖鍍鋅鐵桶", "", 120, "合格"],
        ["RC-20260702", "1000L IBC耐酸鹼桶", "", 15, "合格"],
        ["RC-20260703", "20L PE塑膠小桶", "", 300, "合格"],
        ["RC-20260704", "200L 烤漆鋼桶", "", 80, "待檢"],
        ["RC-20260705", "50加侖鍍鋅鐵桶", "", 200, "合格"],
        ["RC-20260706", "20L PE塑膠小桶", "", 150, "合格"],
    ]

    master_data = [
        ["MAT-DRUM-50G-GI", "50加侖鍍鋅鐵桶"],
        ["MAT-IBC-1000L", "1000L IBC耐酸鹼桶"],
        ["MAT-CAN-20L-PE", "20L PE塑膠小桶"],
        ["MAT-DRUM-200L-P", "200L 烤漆鋼桶"],
        ["MAT-PALLET-WOOD", "120x100木製棧板"],
    ]

    for r_idx, row in enumerate(rec_data, start=6):
        for c_idx, val in enumerate(row, start=1):
            cell = ws2.cell(row=r_idx, column=c_idx, value=val)
            cell.font = regular_font
            cell.border = thin_border

    for r_idx, row in enumerate(master_data, start=6):
        for c_idx, val in enumerate(row, start=7):
            cell = ws2.cell(row=r_idx, column=c_idx, value=val)
            cell.font = regular_font
            cell.border = thin_border
            cell.fill = sub_fill

    ws2['A1'].font = title_font
    ws2['A2'].font = regular_font

    # -------------------------------------------------------------
    # Sheet 4: 題目3_包材領用統計
    # -------------------------------------------------------------
    ws3 = wb.create_sheet(title="04_包材領用統計")
    ws3.views.sheetView[0].showGridLines = True

    ws3.append(["題目 3：包材日領用量樞紐分析統計（30分）"])
    ws3.append(["任務：請利用下方的【包材領用明細表】，在此頁面右側或下方插入一個『樞紐分析表（Pivot Table）』，統計各『領用部門』針對各『包材名稱』的領用總數量。"])
    ws3.append([])

    headers3 = ["領用日期", "領用單號", "領用部門", "包材名稱", "批號", "領用數量", "發料人"]
    ws3.append(headers3)

    data3 = [
        ["2026-07-01", "ISS-001", "彰濱一廠-IPA產線", "50加侖鍍鋅鐵桶", "LOT-20260615-A", 50, "張助理"],
        ["2026-07-01", "ISS-002", "彰濱二廠-電子級PMA", "1000L IBC耐酸鹼桶", "LOT-20260620-B", 4, "張助理"],
        ["2026-07-01", "ISS-003", "彰濱一廠-IPA產線", "20L PE塑膠小桶", "LOT-20260618-C", 100, "李管理師"],
        ["2026-07-02", "ISS-004", "彰濱三廠-化學品分裝", "50加侖鍍鋅鐵桶", "LOT-20260615-A", 40, "張助理"],
        ["2026-07-02", "ISS-005", "彰濱二廠-電子級PMA", "200L 烤漆鋼桶", "LOT-20260622-D", 30, "李管理師"],
        ["2026-07-02", "ISS-006", "彰濱一廠-IPA產線", "50加侖鍍鋅鐵桶", "LOT-20260625-E", 60, "張助理"],
        ["2026-07-03", "ISS-007", "彰濱三廠-化學品分裝", "20L PE塑膠小桶", "LOT-20260618-C", 80, "李管理師"],
    ]

    for row in data3:
        ws3.append(row)

    ws3['A1'].font = title_font
    ws3['A2'].font = regular_font

    for col in range(1, 8):
        cell = ws3.cell(row=4, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for r in range(5, 5 + len(data3)):
        for c in range(1, 8):
            cell = ws3.cell(row=r, column=c)
            cell.font = regular_font
            cell.border = thin_border

    # -------------------------------------------------------------
    # Sheet 5: 面試官解答與評分標準 (僅面試官版包含)
    # -------------------------------------------------------------
    if is_interviewer_version:
        ws4 = wb.create_sheet(title="05_面試官解答與評分標準")
        ws4.views.sheetView[0].showGridLines = True

        ws4.append(["勝一化工【彰濱廠區】助理管理師 - 應徵者實機測驗解答與評分標準"])
        ws4.append(["(本分頁僅供面試官參考評分使用)"])
        ws4.append([])

        ws4.append(["【題目 1 解答與評核點】（40分）"])
        ws4.append(["- 公式範例：H6 輸入 `=F6-G6` 並向下複製。"])
        ws4.append(["- 異常 1（WB2026070103）：毛重 13,000 kg 小於空重 14,800 kg，淨重算出來為負數 (-1,800 kg)。"])
        ws4.append(["- 異常 2（WB2026070106）：空重 41,000 kg 輸入錯誤（多打一個 0），導致淨重為負數 (-7,000 kg)。"])
        ws4.append(["- 評分基準：公式正確得 20 分；能抓出 2 筆異常並說明合理原因得 20 分。"])
        ws4.append([])

        ws4.append(["【題目 2 解答與評核點】（30分）"])
        ws4.append(["- 公式範例：C6 輸入 `=VLOOKUP(B6, $G$6:$H$10, 1, FALSE)` 或 `=XLOOKUP(B6, $H$6:$H$10, $G$6:$G$10)`"])
        ws4.append(["- 評分基準：能正確使用 VLOOKUP/XLOOKUP 並鎖定欄位 ($) 計算得 30 分；若手動貼上得 0 分。"])
        ws4.append([])

        ws4.append(["【題目 3 解答與評核點】（30分）"])
        ws4.append(["- 成果要求：建立樞紐分析表，列為「領用部門」，欄為「包材名稱」，值為「領用數量 之 加總」。"])
        ws4.append(["- 評分基準：正確拉出樞紐分析表得 30 分。"])

        ws4['A1'].font = title_font
        ws4['A2'].font = error_font
        ws4['A4'].font = bold_font
        ws4['A10'].font = bold_font
        ws4['A15'].font = bold_font

    # Adjust column widths
    for sheet in wb.worksheets:
        for col in sheet.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value:
                    val_str = str(cell.value)
                    byte_len = len(val_str.encode('utf-8'))
                    max_len = max(max_len, byte_len)
            sheet.column_dimensions[col_letter].width = max(max_len * 0.85 + 4, 12)

    return wb

# Build and save Examinee version
wb_examinee = build_wb(is_interviewer_version=False)
wb_examinee.save(r"d:\GOOGLE ANGET\勝一化工_彰濱廠區_助理管理師_實機測驗題庫_考生用.xlsx")

# Build and save Interviewer version
wb_interviewer = build_wb(is_interviewer_version=True)
wb_interviewer.save(r"d:\GOOGLE ANGET\勝一化工_彰濱廠區_助理管理師_實機測驗題庫_面試官解答版.xlsx")

print("Successfully generated candidate & interviewer excel files!")
