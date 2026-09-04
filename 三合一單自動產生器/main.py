import pytesseract
from pytesseract import Output
from tkinter import filedialog
from io import BytesIO
from openpyxl.drawing.image import Image as OpenpyxlImage
import tkinter as tk
from tkinter import ttk, messagebox
import openpyxl
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.cell.rich_text import TextBlock, CellRichText
from openpyxl.cell.text import InlineFont
import os
import re
import glob
import json
import csv
import webbrowser
from datetime import datetime, timedelta, date
from openpyxl.utils import get_column_letter
import calendar
import qrcode
from PIL import Image as PILImage, Image
from io import BytesIO
from copy import copy

# ================= 浮動日曆選擇器 =================

class CalendarDialog(tk.Toplevel):
    def __init__(self, parent, target_var):
        super().__init__(parent)
        self.title("選擇日期")
        self.resizable(False, False)
        self.target_var = target_var
        self.transient(parent)
        self.grab_set()

        now = datetime.now()
        self.year = now.year
        self.month = now.month

        val = target_var.get().strip()
        if val:
            for fmt in ("%Y/%m/%d", "%Y-%m-%d", "%Y/%M/%d", "%Y.%m.%d"):
                try:
                    dt = datetime.strptime(val, fmt)
                    self.year = dt.year
                    self.month = dt.month
                    break
                except ValueError:
                    pass

        try:
            x = parent.winfo_pointerx()
            y = parent.winfo_pointery()
            self.geometry(f"+{x}+{y}")
        except Exception:
            pass

        self.setup_ui()
        
    def setup_ui(self):
        header = tk.Frame(self, pady=5, bg="#F5F5F5")
        header.pack(fill="x")
        
        tk.Button(header, text="◄", command=self.prev_month, width=3, relief="groove").pack(side="left", padx=5)
        self.lbl_month = tk.Label(header, text="", font=("Arial", 11, "bold"), bg="#F5F5F5", width=12)
        self.lbl_month.pack(side="left", expand=True)
        tk.Button(header, text="►", command=self.next_month, width=3, relief="groove").pack(side="right", padx=5)

        week_frame = tk.Frame(self, bg="#FFFFFF")
        week_frame.pack(fill="x", padx=5, pady=(5, 0))
        for w in ["一", "二", "三", "四", "五", "六", "日"]:
            fg_col = "#D32F2F" if w in ("六", "日") else "#333333"
            tk.Label(week_frame, text=w, width=4, font=("Arial", 9, "bold"), fg=fg_col, bg="#FFFFFF").pack(side="left")

        self.grid_frame = tk.Frame(self, bg="#FFFFFF", padx=5, pady=5)
        self.grid_frame.pack()
        
        btn_frame = tk.Frame(self, pady=5, bg="#F5F5F5")
        btn_frame.pack(fill="x")
        tk.Button(btn_frame, text="今天", command=self.select_today, bg="#E3F2FD", relief="groove", font=("Arial", 9)).pack(side="left", padx=5)
        tk.Button(btn_frame, text="明天", command=self.select_tomorrow, bg="#E8F5E9", relief="groove", font=("Arial", 9)).pack(side="left", padx=5)
        tk.Button(btn_frame, text="清除", command=self.clear_date, bg="#FFEBEE", relief="groove", font=("Arial", 9)).pack(side="right", padx=5)
        
        self.render_calendar()

    def render_calendar(self):
        for widget in self.grid_frame.winfo_children():
            widget.destroy()
            
        self.lbl_month.config(text=f"{self.year} 年 {self.month} 月")
        cal = calendar.monthcalendar(self.year, self.month)
        
        for r_idx, week in enumerate(cal):
            for c_idx, day in enumerate(week):
                if day == 0:
                    tk.Label(self.grid_frame, text="", width=4, bg="#FFFFFF").grid(row=r_idx, column=c_idx)
                else:
                    date_str = f"{self.year}/{self.month:02d}/{day:02d}"
                    fg_col = "#D32F2F" if c_idx >= 5 else "#000000"
                    btn = tk.Button(
                        self.grid_frame, 
                        text=str(day), 
                        width=4, 
                        bg="#F9F9F9",
                        fg=fg_col,
                        relief="flat",
                        command=lambda d=date_str: self.set_date(d)
                    )
                    btn.grid(row=r_idx, column=c_idx, padx=1, pady=1)

    def prev_month(self):
        if self.month == 1:
            self.month = 12
            self.year -= 1
        else:
            self.month -= 1
        self.render_calendar()

    def next_month(self):
        if self.month == 12:
            self.month = 1
            self.year += 1
        else:
            self.month += 1
        self.render_calendar()

    def select_today(self):
        t = datetime.now()
        self.set_date(f"{t.year}/{t.month:02d}/{t.day:02d}")

    def select_tomorrow(self):
        t = datetime.now() + timedelta(days=1)
        self.set_date(f"{t.year}/{t.month:02d}/{t.day:02d}")

    def clear_date(self):
        self.target_var.set("")
        self.destroy()

    def set_date(self, date_str):
        self.target_var.set(date_str)
        self.destroy()

def split_date_and_time(raw_val):
    """
    將 Excel / 貼上讀取的 datetime 或日期時間字串拆分為 (純日期, 純時間)
    例如:
    - datetime.datetime(2026, 8, 14, 0, 0) -> ("2026-08-14", "")
    - "2026-08-14 00:00:00" -> ("2026-08-14", "")
    - "2026-08-14 08:30:00" -> ("2026-08-14", "08:30")
    """
    if raw_val is None:
        return "", ""
    
    if isinstance(raw_val, datetime):
        d_str = f"{raw_val.year:04d}-{raw_val.month:02d}-{raw_val.day:02d}"
        if raw_val.hour == 0 and raw_val.minute == 0 and raw_val.second == 0:
            t_str = ""
        else:
            t_str = f"{raw_val.hour:02d}:{raw_val.minute:02d}"
        return d_str, t_str

    val_str = str(raw_val).strip()
    if not val_str:
        return "", ""

    parts = val_str.split()
    if len(parts) >= 2:
        d_part = parts[0]
        t_part = parts[1]
        if t_part in ("00:00:00", "00:00", "0:00", "00:00:00.000"):
            return d_part, ""
        else:
            t_sub = t_part.split(':')
            if len(t_sub) >= 2:
                return d_part, f"{t_sub[0]:0>2}:{t_sub[1]:0>2}"
            return d_part, t_part

    return val_str, ""

def normalize_date_str(raw):
    if not raw:
        return ""
    if hasattr(raw, "strftime"):
        return raw.strftime("%Y-%m-%d")
    s = str(raw).strip()
    if not s:
        return ""
    try:
        f = float(s)
        if 20000 < f < 60000:
            dt = datetime(1899, 12, 30) + timedelta(days=int(f))
            return dt.strftime("%Y-%m-%d")
    except ValueError:
        pass
    s_part = s.split()[0]
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d", "%Y%m%d", "%m/%d/%Y", "%d/%m/%Y"):
        try:
            return datetime.strptime(s_part, fmt).strftime("%Y-%m-%d")
        except ValueError:
            pass
    m = re.match(r"^(\d{2,3})[-/\.](\d{1,2})[-/\.](\d{1,2})", s_part)
    if m and int(m.group(1)) < 1900:
        y = int(m.group(1)) + 1911
        m_val = int(m.group(2))
        d_val = int(m.group(3))
        return f"{y:04d}-{m_val:02d}-{d_val:02d}"
    return s_part

def clean_location_str(loc, mapping_dict=None):
    if not loc:
        return ""
    s = str(loc).strip().upper()
    cleaned = re.sub(r'台積電?|新竹|台中|台南|廠|[-_\s]', '', s)
    if mapping_dict and cleaned in mapping_dict:
        return cleaned
    if mapping_dict and s in mapping_dict:
        return s
    m_ap = re.search(r'(AP\d+[A-Z0-9]*)', s)
    if m_ap:
        return m_ap.group(1)
    m_loc = re.search(r'(\d+[A-Z]\d+[A-Z0-9]*|\d+[A-Z0-9]+)', s)
    if m_loc:
        val = m_loc.group(1)
        if (mapping_dict and val in mapping_dict) or any(x in val for x in ["18P", "15P", "12P", "14P"]):
            return val
    return cleaned or s

class ImportRangeDialog(tk.Toplevel):
    """
    Excel 匯入筆數與範圍選擇對話框 (支援全分頁跨頁統計、今天~後天智慧過濾與倒數擷取)
    """
    def __init__(self, parent, records, filename="", sheet_count=1):
        super().__init__(parent)
        self.parent_app = parent
        self.title("📥 Excel 匯入筆數與範圍選擇 (跨分頁智慧過濾)")
        self.geometry("1080x700")
        self.minsize(920, 560)
        self.configure(padx=15, pady=15)
        self.transient(parent)
        self.grab_set()

        # 視窗居中於父視窗
        try:
            self.update_idletasks()
            pw = parent.winfo_width()
            ph = parent.winfo_height()
            px = parent.winfo_rootx()
            py = parent.winfo_rooty()
            w, h = 1080, 700
            x = max(20, px + (pw - w) // 2)
            y = max(20, py + (ph - h) // 2)
            self.geometry(f"{w}x{h}+{x}+{y}")
        except Exception:
            pass

        self.all_records = records
        self.sheet_count = max(1, sheet_count)
        self.total_records_count = len(records)
        self.selected_records = None

        # 計算 今天、明天、後天 日期
        today = datetime.now().date()
        self.d0 = today.strftime("%Y-%m-%d")
        self.d1 = (today + timedelta(days=1)).strftime("%Y-%m-%d")
        self.d2 = (today + timedelta(days=2)).strftime("%Y-%m-%d")
        self.d0_d2 = [self.d0, self.d1, self.d2]

        # 預設過濾邏輯：優先過濾「今天~後天 (3天)」
        window_matches = [r for r in self.all_records if r.get("date") in self.d0_d2]
        if window_matches:
            self.current_filtered_records = window_matches
            self.active_mode = "d0_d2"
            self.summary_text = f"已優先過濾「今天~後天 ({self.d0} ~ {self.d2})」共 {len(window_matches)} 筆 (全 {self.sheet_count} 個分頁總計 {self.total_records_count} 筆)"
        else:
            self.current_filtered_records = list(self.all_records)
            self.active_mode = "all"
            self.summary_text = f"全 {self.sheet_count} 個分頁總計 {self.total_records_count} 筆排程 (無今天~後天資料，顯示全部)"

        self.count_var = tk.IntVar(value=len(self.current_filtered_records) if len(self.current_filtered_records) <= 10 else 10)
        self.info_var = tk.StringVar(value=self.summary_text)

        self.date_btns = {}
        self.setup_ui(filename)
        self.update_preview()

    def setup_ui(self, filename):
        # 頂部提示資訊
        info_frame = tk.LabelFrame(self, text="檔案偵測結果", font=("Arial", 10, "bold"), padx=10, pady=8, fg="#002060")
        info_frame.pack(fill="x", pady=(0, 10))

        tk.Label(info_frame, text=f"📄 檔案名稱：{filename}", font=("Arial", 10, "bold")).pack(anchor="w")
        lbl_sum = tk.Label(info_frame, textvariable=self.info_var, fg="#2E7D32", font=("Arial", 10, "bold"), wraplength=1020, justify="left")
        lbl_sum.pack(anchor="w", pady=(2, 0))

        # 篩選控制區
        ctrl_frame = tk.LabelFrame(self, text="🎯 篩選與筆數設定（支援日期過濾與倒數擷取）", font=("Arial", 10, "bold"), padx=10, pady=8, fg="#C00000")
        ctrl_frame.pack(fill="x", pady=(0, 10))

        # 1. 篩選日期按鈕列
        date_bar = tk.Frame(ctrl_frame)
        date_bar.pack(fill="x", pady=(0, 6))

        tk.Label(date_bar, text="篩選日期：", font=("Arial", 10, "bold")).pack(side="left")

        btn_3days = tk.Button(
            date_bar, 
            text="📅 優先抓今天~後天 (3天)", 
            command=lambda: self.filter_by_date_mode("d0_d2"), 
            font=("Arial", 9, "bold"), 
            padx=8, 
            cursor="hand2"
        )
        btn_3days.pack(side="left", padx=3)
        self.date_btns["d0_d2"] = btn_3days

        btn_d0 = tk.Button(
            date_bar, 
            text="今天", 
            command=lambda: self.filter_by_date_mode("d0"), 
            font=("Arial", 9, "bold"), 
            padx=8, 
            cursor="hand2"
        )
        btn_d0.pack(side="left", padx=3)
        self.date_btns["d0"] = btn_d0

        btn_d1 = tk.Button(
            date_bar, 
            text="明天", 
            command=lambda: self.filter_by_date_mode("d1"), 
            font=("Arial", 9, "bold"), 
            padx=8, 
            cursor="hand2"
        )
        btn_d1.pack(side="left", padx=3)
        self.date_btns["d1"] = btn_d1

        btn_d2 = tk.Button(
            date_bar, 
            text="後天", 
            command=lambda: self.filter_by_date_mode("d2"), 
            font=("Arial", 9, "bold"), 
            padx=8, 
            cursor="hand2"
        )
        btn_d2.pack(side="left", padx=3)
        self.date_btns["d2"] = btn_d2

        btn_all_date = tk.Button(
            date_bar, 
            text="全部日期", 
            command=lambda: self.filter_by_date_mode("all"), 
            font=("Arial", 9, "bold"), 
            padx=8, 
            cursor="hand2"
        )
        btn_all_date.pack(side="left", padx=3)
        self.date_btns["all"] = btn_all_date

        self.refresh_date_buttons()

        # 2. 筆數快速按鈕列
        count_bar = tk.Frame(ctrl_frame)
        count_bar.pack(fill="x")

        tk.Label(count_bar, text="選擇筆數：", font=("Arial", 10, "bold")).pack(side="left")

        for num, text in [(5, "5 筆"), (10, "10 筆"), (20, "20 筆"), ("all", "全部")]:
            tk.Button(
                count_bar, 
                text=text, 
                command=lambda n=num: self.set_count(n), 
                font=("Arial", 9, "bold"), 
                bg="#E3F2FD", 
                fg="#0D47A1", 
                padx=6, 
                cursor="hand2"
            ).pack(side="left", padx=3)

        tk.Label(count_bar, text="自訂筆數：", font=("Arial", 9)).pack(side="left", padx=(10, 0))
        self.spin = tk.Spinbox(count_bar, from_=1, to=max(1, len(self.current_filtered_records)), textvariable=self.count_var, width=6, command=self.update_preview, font=("Arial", 10, "bold"))
        self.spin.pack(side="left", padx=5)
        self.spin.bind("<KeyRelease>", lambda e: self.update_preview())
        tk.Label(count_bar, text="筆", font=("Arial", 9)).pack(side="left")

        # 預覽表格區 (採用專業 Treeview 多欄呈現，欄寬自由拖拉，全資料完整展示)
        preview_frame = tk.LabelFrame(self, text="📋 即將匯入資料即時預覽 (欄寬可自由拉動，全欄位完整呈現)", font=("Arial", 10, "bold"), padx=6, pady=6, fg="#002060")
        preview_frame.pack(fill="both", expand=True, pady=(0, 10))

        # 定義 Treeview 欄位
        columns = ("idx", "sheet", "date", "time", "batch", "tank", "loc", "long_code")
        self.tree = ttk.Treeview(preview_frame, columns=columns, show="headings", selectmode="none")

        col_defs = [
            ("idx", "項次", 50, "center"),
            ("sheet", "來源分頁", 150, "w"),
            ("date", "出貨日期 📅", 105, "center"),
            ("time", "到廠時間", 85, "center"),
            ("batch", "批號 (10碼)", 125, "center"),
            ("tank", "槽號", 75, "center"),
            ("loc", "指送地點", 95, "center"),
            ("long_code", "地點長代號 (全稱)", 250, "w")
        ]

        for col_id, heading_text, width, anchor in col_defs:
            self.tree.heading(col_id, text=heading_text, anchor=anchor)
            self.tree.column(col_id, width=width, minwidth=40, anchor=anchor, stretch=True)

        scroll_y = ttk.Scrollbar(preview_frame, orient="vertical", command=self.tree.yview)
        scroll_x = ttk.Scrollbar(preview_frame, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=scroll_y.set, xscrollcommand=scroll_x.set)

        scroll_y.pack(side="right", fill="y")
        scroll_x.pack(side="bottom", fill="x")
        self.tree.pack(side="left", fill="both", expand=True)

        # 斑馬紋樣式設定
        self.tree.tag_configure("evenrow", background="#FFFFFF")
        self.tree.tag_configure("oddrow", background="#F2F7FA")

        # 底部確定按鈕區
        action_frame = tk.Frame(self)
        action_frame.pack(fill="x")

        self.btn_confirm = tk.Button(
            action_frame, 
            text="🚀 確認匯入資料", 
            command=self.confirm_import, 
            bg="#4CAF50", 
            fg="white", 
            font=("Arial", 11, "bold"), 
            pady=7,
            cursor="hand2"
        )
        self.btn_confirm.pack(side="left", fill="x", expand=True, padx=(0, 5))

        tk.Button(
            action_frame, 
            text="❌ 取消", 
            command=self.destroy, 
            bg="#9E9E9E", 
            fg="white", 
            font=("Arial", 10), 
            pady=7,
            width=10,
            cursor="hand2"
        ).pack(side="right", padx=5)

    def refresh_date_buttons(self):
        for mode, btn in self.date_btns.items():
            if mode == self.active_mode:
                if mode == "d0_d2":
                    btn.config(bg="#FF9800", fg="white")
                elif mode == "all":
                    btn.config(bg="#7B1FA2", fg="white")
                else:
                    btn.config(bg="#1976D2", fg="white")
            else:
                btn.config(bg="#EEEEEE", fg="#424242")

    def filter_by_date_mode(self, mode):
        if mode == "d0_d2":
            filtered = [r for r in self.all_records if r.get("date") in self.d0_d2]
            label = f"「今天~後天 ({self.d0} ~ {self.d2})」"
        elif mode == "d0":
            filtered = [r for r in self.all_records if r.get("date") == self.d0]
            label = f"「今天 ({self.d0})」"
        elif mode == "d1":
            filtered = [r for r in self.all_records if r.get("date") == self.d1]
            label = f"「明天 ({self.d1})」"
        elif mode == "d2":
            filtered = [r for r in self.all_records if r.get("date") == self.d2]
            label = f"「後天 ({self.d2})」"
        else:
            filtered = list(self.all_records)
            label = "「全部日期」"

        if not filtered:
            messagebox.showinfo("無排程資料", f"在所有 {self.sheet_count} 個分頁中，找不到符合 {label} 的出貨排程！")
            return

        self.active_mode = mode
        self.current_filtered_records = filtered
        self.refresh_date_buttons()
        self.info_var.set(f"已篩選 {label} 共 {len(filtered)} 筆 (全 {self.sheet_count} 個分頁總計 {self.total_records_count} 筆)")
        
        self.spin.config(to=max(1, len(filtered)))
        self.count_var.set(len(filtered))
        self.update_preview()

    def set_count(self, num):
        if num == "all":
            self.count_var.set(len(self.current_filtered_records))
        else:
            self.count_var.set(min(int(num), len(self.current_filtered_records)))
        self.update_preview()

    def update_preview(self):
        total = len(self.current_filtered_records)
        try:
            cnt = self.count_var.get()
        except Exception:
            cnt = 1
        cnt = max(1, min(cnt, total))
        self.count_var.set(cnt)

        slice_records = self.current_filtered_records[-cnt:]

        # 清空 Treeview
        for item in self.tree.get_children():
            self.tree.delete(item)

        for idx, rec in enumerate(slice_records):
            sheet_name = rec.get("sheet", "分頁")
            d_str = rec.get("date") or ""
            t_str = rec.get("time") or ""
            b_str = rec.get("batch") or ""
            tank_str = rec.get("tank") or ""
            loc_str = rec.get("loc") or ""
            long_code_str = rec.get("long_code") or getattr(self.parent_app, "mapping_dict", {}).get(loc_str, "")

            tag = "evenrow" if idx % 2 == 0 else "oddrow"
            self.tree.insert(
                "",
                "end",
                values=(
                    f"[{idx+1:02d}]",
                    sheet_name,
                    d_str,
                    t_str,
                    b_str,
                    tank_str,
                    loc_str,
                    long_code_str
                ),
                tags=(tag,)
            )

        self.btn_confirm.config(text=f"🚀 確認由下往上擷取最新的 {cnt} 筆匯入系統")

    def confirm_import(self):
        cnt = self.count_var.get()
        cnt = max(1, min(cnt, len(self.current_filtered_records)))
        self.selected_records = self.current_filtered_records[-cnt:]
        self.destroy()

# ================= 核心邏輯 =================

def get_tank_from_batch(batch):
    batch = batch.strip()
    if not batch:
        return ""
    if len(batch) != 10:
        return "長度錯誤"
    if batch.endswith('J1'):
        return batch[5:8]
    return batch[5:9]

def find_row_by_label(ws, labels):
    for r in range(1, 20):
        val = ws.cell(row=r, column=2).value
        if val and isinstance(val, str):
            for label in labels:
                if label in val:
                    return r
    return None

def generate_transport_notice_file(output_path, items, mat_no="L12C53161"):
    is_append = os.path.exists(output_path)
    if is_append:
        wb = openpyxl.load_workbook(output_path)
        ws = wb.active
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "運輸通知表"
        ws.views.sheetView[0].showGridLines = True
        
        # 欄寬設定 (Left Card: A~F, Spacer: G, Right Card: H~M)
        col_widths = {
            'A': 8, 'B': 18, 'C': 18, 'D': 16, 'E': 14, 'F': 16,
            'G': 4,
            'H': 8, 'I': 18, 'J': 18, 'K': 16, 'L': 14, 'M': 16
        }
        for col, width in col_widths.items():
            ws.column_dimensions[col].width = width
        
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    
    fill_yellow = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    fill_green = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    fill_bright_yellow = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    
    openpyxl_font_dark_blue = Font(name="Microsoft JhengHei", color="002060", size=11)
    openpyxl_font_dark_blue_b13 = Font(name="Microsoft JhengHei", color="002060", size=13, bold=True)
    openpyxl_font_strike_blue_b13 = Font(name="Microsoft JhengHei", color="002060", size=13, bold=True, strike=True)
    openpyxl_font_red_b13 = Font(name="Microsoft JhengHei", color="C00000", size=13, bold=True)
    openpyxl_font_dark_blue_b14 = Font(name="Microsoft JhengHei", color="002060", size=14, bold=True)
    
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    weekdays = ["星期一", "星期二", "星期三", "星期四", "星期五", "星期六", "星期日"]

    def render_notice_card(start_r, start_c, is_modified_card, item):
        r1 = start_r
        r2 = start_r + 1
        r3 = start_r + 2
        r4 = start_r + 3
        r5 = start_r + 4
        r6 = start_r + 5
        
        c1 = start_c
        c2 = start_c + 1
        c3 = start_c + 2
        c4 = start_c + 3
        c5 = start_c + 4
        c6 = start_c + 5
        
        for r in range(r1, r6 + 1):
            ws.row_dimensions[r].height = 24 if r >= r3 else 22
            for c in range(c1, c6 + 1):
                cell = ws.cell(row=r, column=c)
                cell.border = thin_border
                cell.font = openpyxl_font_dark_blue
                cell.alignment = align_center

        # --- Title ---
        ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c3)
        title_suffix = "出貨排程修正通知" if is_modified_card else "出貨排程通知"
        title_rt = CellRichText([
            TextBlock(InlineFont(color="002060", b=True, sz=12, rFont="Microsoft JhengHei"), "Shiny IPA Lorry\n"),
            TextBlock(InlineFont(color="002060", sz=10, rFont="Microsoft JhengHei"), f"(料號：{mat_no}) {title_suffix}")
        ])
        cell_a1 = ws.cell(row=r1, column=c1)
        cell_a1.value = title_rt
        cell_a1.fill = fill_yellow
        cell_a1.alignment = align_center
        
        cell_d1 = ws.cell(row=r1, column=c4, value="廠區")
        cell_d1.fill = fill_yellow
        cell_d1.font = Font(name="Microsoft JhengHei", color="002060", size=11, bold=True)
        
        cell_e1 = ws.cell(row=r1, column=c5, value="槽號")
        cell_e1.fill = fill_yellow
        cell_e1.font = Font(name="Microsoft JhengHei", color="002060", size=11, bold=True)
        
        date_raw = item.get("date", "").strip()
        formatted_date = ""
        weekday_str = ""
        if date_raw:
            dt = None
            for fmt in ("%Y/%m/%d", "%Y-%m-%d", "%Y/%M/%d", "%Y.%m.%d"):
                try:
                    dt = datetime.strptime(date_raw, fmt)
                    break
                except ValueError:
                    pass
            if dt:
                weekday_str = weekdays[dt.weekday()]
                formatted_date = f"{dt.year}/{dt.month}/{dt.day}"
            else:
                formatted_date = date_raw
        
        cell_f1 = ws.cell(row=r1, column=c6, value=formatted_date)
        cell_f1.fill = fill_green
        cell_f1.font = Font(name="Microsoft JhengHei", color="002060", size=11, bold=True)
        
        cell_f2 = ws.cell(row=r2, column=c6, value=weekday_str)
        cell_f2.fill = fill_green
        cell_f2.font = Font(name="Microsoft JhengHei", color="002060", size=11, bold=True)

        # --- Body ---
        ws.merge_cells(start_row=r3, start_column=c1, end_row=r6, end_column=c1)
        cell_a3 = ws.cell(row=r3, column=c1, value="IPA")
        cell_a3.font = openpyxl_font_dark_blue_b14
        
        time_val = item.get("time", "").strip()
        mod_time_val = item.get("mod_time", "").strip()
        
        ws.merge_cells(start_row=r3, start_column=c2, end_row=r3, end_column=c3)
        ws.cell(row=r3, column=c2, value="預計到廠時間")
        cell_f3 = ws.cell(row=r3, column=c6, value=time_val)
        
        ws.merge_cells(start_row=r4, start_column=c2, end_row=r4, end_column=c3)
        ws.cell(row=r4, column=c2, value="修正到廠時間")
        cell_f4 = ws.cell(row=r4, column=c6)
        
        if is_modified_card:
            cell_f3.font = openpyxl_font_strike_blue_b13
            cell_f4.value = mod_time_val
            cell_f4.font = openpyxl_font_red_b13
        else:
            cell_f3.font = openpyxl_font_dark_blue_b13
            cell_f4.value = ""
        
        ws.merge_cells(start_row=r5, start_column=c2, end_row=r5, end_column=c3)
        ws.cell(row=r5, column=c2, value="充填數量(KG)")
        ws.cell(row=r5, column=c6, value="4300")
        
        ws.merge_cells(start_row=r2, start_column=c4, end_row=r5, end_column=c4)
        full_loc = item.get("loc", "").strip()
        
        d_rt = CellRichText([
            TextBlock(InlineFont(color="002060", b=True, sz=14, rFont="Microsoft JhengHei"), "台積\n"),
            TextBlock(InlineFont(color="C00000", b=True, sz=14, rFont="Microsoft JhengHei"), full_loc)
        ])
        cell_d2 = ws.cell(row=r2, column=c4)
        cell_d2.value = d_rt
        
        ws.merge_cells(start_row=r2, start_column=c5, end_row=r5, end_column=c5)
        cell_e2 = ws.cell(row=r2, column=c5, value=item.get("tank", "").strip())
        cell_e2.font = openpyxl_font_dark_blue_b14
        cell_e2.fill = fill_bright_yellow
        
        ws.merge_cells(start_row=r6, start_column=c2, end_row=r6, end_column=c5)
        cell_b6 = ws.cell(row=r6, column=c2)
        cell_b6.value = CellRichText([
            TextBlock(InlineFont(color="C00000", b=True, sz=11, rFont="Microsoft JhengHei"), "PFA 500ml"),
            TextBlock(InlineFont(color="002060", sz=11, rFont="Microsoft JhengHei"), " 取樣瓶裝原液 8 分滿放置工具箱內")
        ])
        
        cell_f6 = ws.cell(row=r6, column=c6, value="6 支")
        cell_f6.font = Font(name="Microsoft JhengHei", color="002060", size=11, bold=True)

    curr_row = ws.max_row + 2 if is_append and ws.max_row > 1 else 1
    for item in items:
        # 左側：原始出貨排程通知卡片 (Columns A~F)
        render_notice_card(curr_row, 1, False, item)
        
        # 若該筆有輸入「修正到廠時間」，在右側旁 (Columns H~M) 併排產生「出貨排程修正通知」卡片
        if item.get("mod_time", "").strip():
            render_notice_card(curr_row, 8, True, item)
            
        curr_row += 7

    wb.save(output_path)
    wb.close()

def build_single_row_lorry_workbook(src_ws, target_row, max_cols=30):
    new_wb = openpyxl.Workbook()
    new_ws = new_wb.active
    new_ws.title = src_ws.title
    
    # 複製欄寬
    for col_letter, col_dim in src_ws.column_dimensions.items():
        if col_dim.width:
            new_ws.column_dimensions[col_letter].width = col_dim.width
            
    def copy_cell(s_cell, d_cell):
        d_cell.value = s_cell.value
        if s_cell.has_style:
            d_cell.font = copy(s_cell.font)
            d_cell.border = copy(s_cell.border)
            d_cell.fill = copy(s_cell.fill)
            d_cell.number_format = copy(s_cell.number_format)
            d_cell.protection = copy(s_cell.protection)
            d_cell.alignment = copy(s_cell.alignment)

    # 複製列高 (保持第 6 列表頭與資料列原始高度)
    for r in range(1, 8):
        src_r = r if r <= 6 else target_row
        if src_ws.row_dimensions[src_r].height:
            new_ws.row_dimensions[r].height = src_ws.row_dimensions[src_r].height

    # 複製 1~6 列表頭
    for r in range(1, 7):
        for c in range(1, max_cols + 1):
            copy_cell(src_ws.cell(r, c), new_ws.cell(r, c))
            
    # 複製目標資料列至第 7 列
    for c in range(1, max_cols + 1):
        copy_cell(src_ws.cell(target_row, c), new_ws.cell(7, c))
        
    # 自動智慧調整欄寬，確保所有欄位表頭與資料完整顯示不被遮擋
    for col_idx in range(1, max_cols + 1):
        col_letter = get_column_letter(col_idx)
        max_len = 0
        for r in range(1, 8):
            val = new_ws.cell(r, col_idx).value
            if val is not None:
                if isinstance(val, (datetime, date)):
                    s = val.strftime('%Y/%m/%d')
                else:
                    s = str(val).strip()
                lines = s.split('\n')
                for line in lines:
                    line_len = sum(2.0 if ord(ch) > 127 else 1.15 for ch in line)
                    if line_len > max_len:
                        max_len = line_len
        if max_len > 0:
            # 依最長文字加上留白邊界 (+4.0)，至少 13.0
            adjusted_width = max(max_len + 4.0, 13.0)
            orig_w = new_ws.column_dimensions[col_letter].width
            if orig_w and orig_w > adjusted_width:
                adjusted_width = orig_w
            new_ws.column_dimensions[col_letter].width = round(adjusted_width, 1)

    # 確保第 6 列表頭高度 (28.0) 與第 7 列資料列高度 (22.0) 呼吸空間
    new_ws.row_dimensions[6].height = 28.0
    new_ws.row_dimensions[7].height = 22.0

    # 保持正常從第 1 列 (A1) 完整顯示表頭與第 7 列資料 (如圖二)，不捲動遮蔽
    new_ws.freeze_panes = 'A7'
            
    return new_wb

# ================= 介面與操作 =================

class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("三合一單自動產生器 & 運輸通知表產生器")
        self.geometry("1260x820")
        self.minsize(1080, 620)
        self.configure(padx=15, pady=15)
        
        self.base_dir = os.path.dirname(os.path.abspath(__file__))
        self.template_path = os.path.join(self.base_dir, "台積電槽車barcode三合一單-範本.xlsx")
        self.mapping_path = os.path.join(self.base_dir, "地點代號對照表.xlsx")
        
        self.mapping_dict = {}
        self.imported_lorry_files = []
        self.load_mapping()
        
        self.setup_ui()
        self.entries = []
        self.coa_paths = []
        self.add_input_rows(20)

    def load_mapping(self):
        self.mapping_dict = {}
        if os.path.exists(self.mapping_path):
            try:
                map_wb = openpyxl.load_workbook(self.mapping_path, data_only=True)
                map_ws = map_wb.active
                for row in map_ws.iter_rows(values_only=True):
                    if row and len(row) >= 2 and row[0] and row[1]:
                        loc_key = str(row[0]).strip().upper()
                        loc_val = str(row[1]).strip()
                        # 過濾標題列 (例如 短地點, 長代號, 地點)
                        if any(kw in loc_key for kw in ("地點", "代號", "SHORT", "LOCATION", "KEY", "HEADER")):
                            continue
                        self.mapping_dict[loc_key] = loc_val
                map_wb.close()
            except Exception as e:
                pass

        if hasattr(self, "lbl_mapping_status"):
            m_color = "green" if os.path.exists(self.mapping_path) else "red"
            m_text = f"對照表檔案 (地點代號對照表.xlsx): ✅ 已找到 (載入 {len(self.mapping_dict)} 筆代號)" if os.path.exists(self.mapping_path) else "對照表檔案 (地點代號對照表.xlsx): ❌ 未找到"
            self.lbl_mapping_status.config(text=m_text, fg=m_color)

        if hasattr(self, "entries"):
            for entry in self.entries:
                loc_val = entry["loc_var"].get().strip().upper()
                if loc_val:
                    self.on_loc_change(entry["loc_var"], entry["long_code_var"])

    def update_lorry_status(self):
        if not hasattr(self, "imported_lorry_files"):
            self.imported_lorry_files = []
        if not self.imported_lorry_files:
            defaults = glob.glob(os.path.join(self.base_dir, "Chemical_Lorry*.xlsx"))
            if defaults:
                self.imported_lorry_files = [defaults[0]]
        
        if hasattr(self, "lbl_lorry_status"):
            if self.imported_lorry_files:
                fname = os.path.basename(self.imported_lorry_files[0])
                self.lbl_lorry_status.config(
                    text=f"生產履歷檔案 (Chemical_Lorry*.xlsx): ✅ 已就緒 ({fname})",
                    fg="#2E7D32"
                )
            else:
                self.lbl_lorry_status.config(
                    text="生產履歷檔案 (Chemical_Lorry*.xlsx): ⚠️ 尚未載入 (點擊右側按鈕或下方工具列『📋 載入生產履歷』)",
                    fg="#D32F2F"
                )

    def load_chemical_lorry_file(self):
        filepath = filedialog.askopenfilename(
            title="選擇生產履歷檔案 (Chemical_Lorry)",
            filetypes=[("Excel 活頁簿", "*.xlsx *.xls"), ("所有檔案", "*.*")]
        )
        if not filepath:
            return
        if not hasattr(self, "imported_lorry_files"):
            self.imported_lorry_files = []
        self.imported_lorry_files = [filepath]
        self.update_lorry_status()
        self.gen_lorry_var.set(True)
        fname = os.path.basename(filepath)
        messagebox.showinfo(
            "生產履歷已載入", 
            f"已成功載入生產履歷檔案：\n{fname}\n\n已為您自動勾選【產生單列生產履歷】！\n稍後點擊【開始批次產生】時，系統會自動比對每筆排程批號並單列輸出。"
        )

    def reload_mapping_with_msg(self):
        """點擊『🔄 重新載入對照表』時執行"""
        self.load_mapping()
        messagebox.showinfo("對照表已更新", f"已重新讀取『地點代號對照表.xlsx』！\n目前共載入 {len(self.mapping_dict)} 筆地點對照碼。\n表格中的地點長代號已同步更新！")

    def open_calendar_dialog(self, target_var):
        CalendarDialog(self, target_var)

    def setup_ui(self):
        # 1. 系統檔案狀態區
        status_frame = tk.LabelFrame(self, text="系統狀態", font=("Microsoft JhengHei", 10, "bold"), padx=10, pady=8)
        status_frame.pack(fill="x", pady=(0, 8))
        
        # 範本狀態
        t_color = "green" if os.path.exists(self.template_path) else "red"
        t_text = "✅ 已找到" if os.path.exists(self.template_path) else "❌ 未找到 (請將檔案放入資料夾)"
        tk.Label(status_frame, text=f"範本檔案 (台積電槽車barcode三合一單-範本.xlsx): {t_text}", fg=t_color).pack(anchor="w")
        
        # 對照表狀態列 + 重新載入按鈕
        map_status_frame = tk.Frame(status_frame)
        map_status_frame.pack(anchor="w", pady=(2, 0))

        m_color = "green" if os.path.exists(self.mapping_path) else "red"
        m_text = f"對照表檔案 (地點代號對照表.xlsx): ✅ 已找到 (載入 {len(self.mapping_dict)} 筆代號)" if os.path.exists(self.mapping_path) else "對照表檔案 (地點代號對照表.xlsx): ❌ 未找到"
        
        self.lbl_mapping_status = tk.Label(map_status_frame, text=m_text, fg=m_color)
        self.lbl_mapping_status.pack(side="left")

        tk.Button(
            map_status_frame, 
            text="🔄 重新載入對照表", 
            command=self.reload_mapping_with_msg, 
            bg="#607D8B", 
            fg="white", 
            font=("Microsoft JhengHei", 8, "bold"), 
            padx=6,
            cursor="hand2"
        ).pack(side="left", padx=10)

        # 生產履歷狀態列 + 選擇檔案按鈕
        lorry_status_frame = tk.Frame(status_frame)
        lorry_status_frame.pack(anchor="w", pady=(2, 0))

        self.lbl_lorry_status = tk.Label(lorry_status_frame, text="", fg="green")
        self.lbl_lorry_status.pack(side="left")

        tk.Button(
            lorry_status_frame, 
            text="📋 選擇生產履歷檔", 
            command=self.load_chemical_lorry_file, 
            bg="#E65100", 
            fg="white", 
            font=("Microsoft JhengHei", 8, "bold"), 
            padx=6,
            cursor="hand2"
        ).pack(side="left", padx=10)
        self.update_lorry_status()

        # 2. 頂部快捷功能與檔案載入工具列
        top_ctrl_frame = tk.Frame(self)
        top_ctrl_frame.pack(fill="x", pady=(0, 6))
        
        # 左側：資料載入與匯入按鈕群組
        left_btn_frame = tk.Frame(top_ctrl_frame)
        left_btn_frame.pack(side="left")
        
        tk.Button(left_btn_frame, text="📥 從 Excel 匯入排程", command=self.import_from_excel, bg="#1976D2", fg="white", font=("Microsoft JhengHei", 9, "bold"), padx=8, pady=2, cursor="hand2").pack(side="left", padx=(0, 4))
        tk.Button(left_btn_frame, text="📋 載入生產履歷 (Chemical_Lorry)", command=self.load_chemical_lorry_file, bg="#E65100", fg="white", font=("Microsoft JhengHei", 9, "bold"), padx=8, pady=2, cursor="hand2").pack(side="left", padx=4)
        tk.Button(left_btn_frame, text="📂 載入既有通知表修訂", command=self.load_existing_transport_notice, bg="#7B1FA2", fg="white", font=("Microsoft JhengHei", 9, "bold"), padx=8, pady=2, cursor="hand2").pack(side="left", padx=4)
        tk.Button(left_btn_frame, text="🖼️ 上傳 COA 截圖", command=self.upload_coa, bg="#FF9800", fg="white", font=("Microsoft JhengHei", 9, "bold"), padx=8, pady=2, cursor="hand2").pack(side="left", padx=4)

        # 右側：表格操作與日期快捷按鈕群組
        right_btn_frame = tk.Frame(top_ctrl_frame)
        right_btn_frame.pack(side="right")
        
        tk.Button(right_btn_frame, text="📅 帶入今天日期", command=self.set_today_all_dates, bg="#546E7A", fg="white", font=("Microsoft JhengHei", 9, "bold"), padx=8, pady=2, cursor="hand2").pack(side="left", padx=4)
        tk.Button(right_btn_frame, text="➕ 新增 10 列", command=lambda: self.add_input_rows(10), bg="#00897B", fg="white", font=("Microsoft JhengHei", 9, "bold"), padx=8, pady=2, cursor="hand2").pack(side="left", padx=4)
        tk.Button(right_btn_frame, text="🗑️ 清除全部資料", command=self.clear_all_rows, bg="#D32F2F", fg="white", font=("Microsoft JhengHei", 9, "bold"), padx=8, pady=2, cursor="hand2").pack(side="left", padx=(4, 0))

        # 3. 一鍵批次設定列 (純淨獨立，日期與時間欄位寬裕舒適)
        batch_setting_frame = tk.LabelFrame(self, text="一鍵批次設定 (出貨日期 / 預計到廠時間 / 修正到廠時間)", font=("Microsoft JhengHei", 9, "bold"), padx=10, pady=5)
        batch_setting_frame.pack(fill="x", pady=(0, 6))
        
        # 全選
        self.select_all_var = tk.BooleanVar(value=True)
        tk.Checkbutton(batch_setting_frame, text="☑ 全選所有列", variable=self.select_all_var, command=self.toggle_select_all, font=("Microsoft JhengHei", 9, "bold")).pack(side="left", padx=(0, 15))
        
        # 日期
        tk.Label(batch_setting_frame, text="批次出貨日期:").pack(side="left")
        self.default_date_var = tk.StringVar(value="")
        date_batch_entry = tk.Entry(batch_setting_frame, textvariable=self.default_date_var, width=12)
        date_batch_entry.pack(side="left", padx=2)
        tk.Button(batch_setting_frame, text="📅", command=lambda: self.open_calendar_dialog(self.default_date_var), font=("Arial", 8), width=3).pack(side="left", padx=(0, 2))
        tk.Button(batch_setting_frame, text="套用至全列", command=self.apply_default_date, bg="#607D8B", fg="white", font=("Microsoft JhengHei", 8)).pack(side="left", padx=(2, 16))
        
        # 預計時間
        tk.Label(batch_setting_frame, text="預計到廠時間:").pack(side="left")
        self.default_time_var = tk.StringVar(value="")
        tk.Entry(batch_setting_frame, textvariable=self.default_time_var, width=8).pack(side="left", padx=2)
        tk.Button(batch_setting_frame, text="套用至全列", command=self.apply_default_time, bg="#607D8B", fg="white", font=("Microsoft JhengHei", 8)).pack(side="left", padx=(2, 16))

        # 修正時間
        tk.Label(batch_setting_frame, text="修正到廠時間:").pack(side="left")
        self.default_mod_time_var = tk.StringVar(value="")
        tk.Entry(batch_setting_frame, textvariable=self.default_mod_time_var, width=8).pack(side="left", padx=2)
        tk.Button(batch_setting_frame, text="套用至全列", command=self.apply_default_mod_time, bg="#607D8B", fg="white", font=("Microsoft JhengHei", 8)).pack(side="left", padx=(2, 10))

        # 4. 報表產出勾選專屬區塊 (獨立整列，寬度充裕，文字絕不截斷)
        report_opt_frame = tk.LabelFrame(self, text="📦 欲產生的報表勾選 (可多選，點擊開始產生時將自動產出所勾選項目)", font=("Microsoft JhengHei", 9, "bold"), padx=10, pady=5)
        report_opt_frame.pack(fill="x", pady=(0, 8))

        self.gen_3in1_var = tk.BooleanVar(value=True)
        self.gen_transport_var = tk.BooleanVar(value=True)
        self.gen_lorry_var = tk.BooleanVar(value=True)

        tk.Checkbutton(report_opt_frame, text="✅ 產生三合一單 Excel (含 Barcode 與 COA)", variable=self.gen_3in1_var, font=("Microsoft JhengHei", 9, "bold"), fg="#1B5E20").pack(side="left", padx=(0, 20))
        tk.Checkbutton(report_opt_frame, text="✅ 產生運輸通知表 Excel (出貨與修正排程通知卡片)", variable=self.gen_transport_var, font=("Microsoft JhengHei", 9, "bold"), fg="#0D47A1").pack(side="left", padx=20)
        tk.Checkbutton(report_opt_frame, text="✅ 產生單列生產履歷 Excel (Chemical_Lorry 槽車充填表)", variable=self.gen_lorry_var, font=("Microsoft JhengHei", 9, "bold"), fg="#E65100").pack(side="left", padx=20)

        # 5. 表格滾動容器
        table_container = tk.Frame(self)
        table_container.pack(fill="both", expand=True)

        self.canvas = tk.Canvas(table_container, borderwidth=0, highlightthickness=0)
        self.scrollbar = ttk.Scrollbar(table_container, orient="vertical", command=self.canvas.yview)
        self.scrollbar_x = ttk.Scrollbar(table_container, orient="horizontal", command=self.canvas.xview)
        
        # 單一 Grid 容器 (scrollable_frame)
        self.scrollable_frame = tk.Frame(self.canvas, padx=5, pady=5)

        self.scrollable_frame.bind(
            "<Configure>",
            lambda e: self.canvas.configure(scrollregion=self.canvas.bbox("all"))
        )

        self.canvas.create_window((0, 0), window=self.scrollable_frame, anchor="nw")
        self.canvas.configure(yscrollcommand=self.scrollbar.set, xscrollcommand=self.scrollbar_x.set)

        self.scrollbar.pack(side="right", fill="y")
        self.scrollbar_x.pack(side="bottom", fill="x")
        self.canvas.pack(side="left", fill="both", expand=True)

        # 支援滑鼠滾輪滾動
        self.bind_all("<MouseWheel>", lambda event: self.canvas.yview_scroll(int(-1 * (event.delta / 120)), "units"))

        # --- Row 0: 表格標題列 (置於滾動區最上方，共享 100% 同一 Grid 欄位規格) ---
        headers = [
            (0, "產生"),
            (1, "項次"),
            (2, "批號 (請輸入10碼)"),
            (3, "槽號 (自動)"),
            (4, "地點 (如 15P5)"),
            (5, "長代號 (自動)"),
            (6, "出貨日期 📅"),
            (7, "預計到廠時間"),
            (8, "修正到廠時間"),
            (9, "單列清空")
        ]
        
        for col_idx, title in headers:
            lbl = tk.Label(
                self.scrollable_frame, 
                text=title, 
                font=("Arial", 10, "bold"), 
                bg="#EAEAEA", 
                fg="#333333",
                padx=6, 
                pady=6,
                relief="groove"
            )
            lbl.grid(row=0, column=col_idx, sticky="ew", padx=1, pady=(0, 6))

        # 產生按鈕
        btn_frame = tk.Frame(self)
        btn_frame.pack(fill="x", pady=10)
        tk.Button(btn_frame, text="🚀 開始批次產生 Excel 報表", command=self.generate_files, bg="#4CAF50", fg="white", font=("Arial", 12, "bold"), pady=8).pack(fill="x")

    def toggle_select_all(self):
        state = self.select_all_var.get()
        for entry in self.entries:
            entry["chk_var"].set(state)

    def apply_default_date(self):
        val = self.default_date_var.get().strip()
        for entry in self.entries:
            if entry["batch_var"].get().strip():
                entry["date_var"].set(val)

    def apply_default_time(self):
        val = self.default_time_var.get().strip()
        for entry in self.entries:
            if entry["batch_var"].get().strip():
                entry["time_var"].set(val)

    def apply_default_mod_time(self):
        val = self.default_mod_time_var.get().strip()
        for entry in self.entries:
            if entry["batch_var"].get().strip():
                entry["mod_time_var"].set(val)

    def add_input_rows(self, count):
        for i in range(count):
            row_idx = len(self.entries) + 1
            row_grid_idx = row_idx
            
            # Col 0: 勾選
            chk_var = tk.BooleanVar(value=True)
            chk = tk.Checkbutton(self.scrollable_frame, variable=chk_var)
            chk.grid(row=row_grid_idx, column=0, padx=2, pady=2)

            # Col 1: 項次
            lbl_num = tk.Label(self.scrollable_frame, text=str(row_idx), font=("Arial", 10))
            lbl_num.grid(row=row_grid_idx, column=1, padx=2, pady=2)
            
            # Col 2: 批號
            batch_var = tk.StringVar()
            batch_entry = tk.Entry(self.scrollable_frame, textvariable=batch_var, width=16, font=("Arial", 10))
            batch_entry.grid(row=row_grid_idx, column=2, padx=2, pady=2, sticky="ew")
            
            # Col 3: 槽號
            tank_var = tk.StringVar()
            tank_entry = tk.Entry(self.scrollable_frame, textvariable=tank_var, state="readonly", width=10, font=("Arial", 10), fg="blue")
            tank_entry.grid(row=row_grid_idx, column=3, padx=2, pady=2, sticky="ew")
            
            # Col 4: 地點
            loc_var = tk.StringVar()
            loc_entry = tk.Entry(self.scrollable_frame, textvariable=loc_var, width=12, font=("Arial", 10))
            loc_entry.grid(row=row_grid_idx, column=4, padx=2, pady=2, sticky="ew")
            
            # Col 5: 長代號
            long_code_var = tk.StringVar()
            long_code_entry = tk.Entry(self.scrollable_frame, textvariable=long_code_var, state="readonly", width=16, font=("Arial", 10), fg="purple")
            long_code_entry.grid(row=row_grid_idx, column=5, padx=2, pady=2, sticky="ew")
            
            # Col 6: 出貨日期 (Entry + 📅 日曆按鈕)
            date_frame = tk.Frame(self.scrollable_frame)
            date_frame.grid(row=row_grid_idx, column=6, padx=2, pady=2, sticky="ew")
            
            date_var = tk.StringVar(value="")
            date_entry = tk.Entry(date_frame, textvariable=date_var, width=11, font=("Arial", 10))
            date_entry.pack(side="left", fill="x", expand=True)
            
            btn_cal = tk.Button(date_frame, text="📅", command=lambda dv=date_var: self.open_calendar_dialog(dv), font=("Arial", 8), cursor="hand2")
            btn_cal.pack(side="right", padx=(2, 0))
            
            # Col 7: 預計到廠時間
            time_var = tk.StringVar(value="")
            time_entry = tk.Entry(self.scrollable_frame, textvariable=time_var, width=10, font=("Arial", 10))
            time_entry.grid(row=row_grid_idx, column=7, padx=2, pady=2, sticky="ew")

            # Col 8: 修正到廠時間
            mod_time_var = tk.StringVar(value="")
            mod_time_entry = tk.Entry(self.scrollable_frame, textvariable=mod_time_var, width=10, font=("Arial", 10), fg="red")
            mod_time_entry.grid(row=row_grid_idx, column=8, padx=2, pady=2, sticky="ew")

            # Col 9: 單列清空按鈕
            btn_clear_row = tk.Button(
                self.scrollable_frame, 
                text="清空", 
                command=lambda r=row_idx-1: self.clear_single_row(r), 
                font=("Microsoft JhengHei", 9, "bold"), 
                bg="#FFEBEE", 
                fg="#C62828", 
                cursor="hand2", 
                width=6,
                pady=1
            )
            btn_clear_row.grid(row=row_grid_idx, column=9, padx=4, pady=2)

            # 綁定事件
            batch_var.trace_add("write", lambda name, index, mode, bv=batch_var, tv=tank_var: self.on_batch_change(bv, tv))
            loc_var.trace_add("write", lambda name, index, mode, lv=loc_var, lcv=long_code_var: self.on_loc_change(lv, lcv))
            
            for widget in (batch_entry, loc_entry, date_entry):
                widget.bind("<<Paste>>", lambda e, r=row_idx-1, w=widget: self.on_paste(e, r, w))
                widget.bind("<Control-v>", lambda e, r=row_idx-1, w=widget: self.on_paste(e, r, w))
                widget.bind("<Control-V>", lambda e, r=row_idx-1, w=widget: self.on_paste(e, r, w))
            
            self.entries.append({
                "chk_var": chk_var,
                "batch_var": batch_var,
                "tank_var": tank_var,
                "loc_var": loc_var,
                "long_code_var": long_code_var,
                "date_var": date_var,
                "time_var": time_var,
                "mod_time_var": mod_time_var
            })

    def clear_all_rows(self):
        """一鍵清除全部輸入欄位 (全清)"""
        if messagebox.askyesno("確認清空", "確定要清空所有已填寫的批號、地點與時間資料嗎？"):
            for entry in self.entries:
                entry["batch_var"].set("")
                entry["loc_var"].set("")
                entry["long_code_var"].set("")
                entry["tank_var"].set("")
                entry["date_var"].set("")
                entry["time_var"].set("")
                entry["mod_time_var"].set("")

    def set_today_all_dates(self):
        from datetime import datetime
        today_str = datetime.now().strftime("%Y/%m/%d")
        for entry in self.entries:
            if not entry["date_var"].get().strip():
                entry["date_var"].set(today_str)

    def clear_single_row(self, r_idx):
        """清空單一列的資料 (單個清)"""
        if 0 <= r_idx < len(self.entries):
            entry = self.entries[r_idx]
            entry["batch_var"].set("")
            entry["loc_var"].set("")
            entry["long_code_var"].set("")
            entry["tank_var"].set("")
            entry["date_var"].set("")
            entry["time_var"].set("")
            entry["mod_time_var"].set("")

    def on_batch_change(self, batch_var, tank_var):
        batch = batch_var.get().upper().strip()
        tank = get_tank_from_batch(batch)
        tank_var.set(tank)

    def on_loc_change(self, loc_var, long_code_var):
        loc = loc_var.get().strip().upper()
        if not loc:
            long_code_var.set("")
            return
            
        if loc in self.mapping_dict:
            long_code_var.set(self.mapping_dict[loc])
        else:
            long_code_var.set("❌ 未知代號")

    def parse_pasted_row_items(self, parts):
        """
        智慧解析貼上行中的元素，自動辨識：出貨日期、批號、槽號、地點、時間。
        支援 Image 1 範例（到貨 + 批號 + 槽號 + 地點）及各式自訂順序！
        """
        res = {"batch": "", "loc": "", "date": "", "time": "", "mod_time": ""}
        clean_parts = [p.strip() for p in parts if p.strip()]
        if not clean_parts:
            return res

        unassigned = []
        for p in clean_parts:
            p_upper = p.upper()
            
            # 1. 日期 (包含 /, -, ., 或月/日，長度 <= 12)
            if not res["date"] and (any(char in p for char in ("/", "-", ".")) or "月" in p or "日" in p):
                if len(p) <= 12 and not p.isalnum():
                    res["date"] = p
                    continue

            # 2. 批號 (8~12 碼英數混合，例如 26817E3051)
            if not res["batch"] and (8 <= len(p_upper) <= 12 and any(c.isdigit() for c in p_upper) and any(c.isalpha() for c in p_upper)):
                res["batch"] = p_upper
                continue

            # 3. 地點 (在 mapping_dict 內或含 18P, 15P, P, 廠等)
            if not res["loc"] and (p_upper in self.mapping_dict or "18P" in p_upper or "15P" in p_upper or "P" in p_upper or "廠" in p):
                res["loc"] = p_upper
                continue

            # 4. 槽號 (3~5 碼以 E/T/P 開頭，例如 E305) -> 可略過，因為寫入批號時微服務會自動算槽號！
            if len(p_upper) <= 5 and p_upper.startswith(("E", "T", "P")):
                continue

            # 其它填入未指派
            unassigned.append(p)

        # 針對缺額自動後補填入
        for item in unassigned:
            if not res["batch"] and len(item) >= 6:
                res["batch"] = item.upper()
            elif not res["loc"] and len(item) <= 10:
                res["loc"] = item.upper()
            elif not res["date"] and ("/" in item or "-" in item or "." in item):
                res["date"] = item
            elif not res["time"]:
                res["time"] = item
            elif not res["mod_time"]:
                res["mod_time"] = item

        return res

    def on_paste(self, event, start_row_idx, widget=None):
        try:
            clipboard = self.clipboard_get()
            lines = clipboard.split('\n')
            
            valid_lines = [l for l in lines if l.strip()]
            if not valid_lines:
                return "break"
                
            needed_rows = start_row_idx + len(valid_lines)
            if needed_rows > len(self.entries):
                self.add_input_rows(needed_rows - len(self.entries))
            
            curr_row = start_row_idx
            for line in lines:
                if not line.strip(): continue
                parts = line.split('\t')
                if len(parts) == 1:
                    parts = line.split()
                    
                if len(parts) >= 2:
                    # 使用智慧元素剖析器！
                    parsed = self.parse_pasted_row_items(parts)
                    
                    if parsed["batch"]:
                        self.entries[curr_row]["batch_var"].set(parsed["batch"])
                    if parsed["loc"]:
                        self.entries[curr_row]["loc_var"].set(parsed["loc"])
                    if parsed["date"]:
                        self.entries[curr_row]["date_var"].set(parsed["date"])
                    if parsed["time"]:
                        self.entries[curr_row]["time_var"].set(parsed["time"])
                    if parsed["mod_time"]:
                        self.entries[curr_row]["mod_time_var"].set(parsed["mod_time"])
                    
                    curr_row += 1
                elif len(parts) == 1:
                    val = parts[0].strip()
                    # 單一欄位貼上：依當前焦點與內容自動指派
                    if val:
                        parsed = self.parse_pasted_row_items([val])
                        if parsed["batch"]:
                            self.entries[curr_row]["batch_var"].set(parsed["batch"])
                        elif parsed["loc"]:
                            self.entries[curr_row]["loc_var"].set(parsed["loc"])
                        elif parsed["date"]:
                            self.entries[curr_row]["date_var"].set(parsed["date"])
                        else:
                            self.entries[curr_row]["batch_var"].set(val)
                    curr_row += 1
                    
            return "break"
        except Exception as e:
            messagebox.showerror("貼上失敗", f"解析貼上內容時發生錯誤:\n{e}")
            return "break"


    def upload_coa(self):
        if not os.path.exists(r'C:\Program Files\Tesseract-OCR\tesseract.exe') and not os.path.exists(r'C:\Program Files (x86)\Tesseract-OCR\tesseract.exe'):
            ans = messagebox.askyesno(
                "缺少 OCR 引擎", 
                "系統偵測到您尚未安裝『Tesseract OCR 引擎』，無法使用截圖辨識功能！\n\n"
                "是否要立即開啟官方下載網頁？\n\n"
                "(請下載最新的 64 bit 安裝檔，並【一直按下一步】安裝在預設路徑即可)"
            )
            if ans:
                webbrowser.open("https://github.com/UB-Mannheim/tesseract/wiki")
            return
            
        filepaths = filedialog.askopenfilenames(title="選擇 COA 截圖", filetypes=[("Image files", "*.jpg *.jpeg *.png")])
        if filepaths:
            self.coa_paths.extend(filepaths)
            messagebox.showinfo("上傳成功", f"成功上傳 {len(filepaths)} 張截圖！\n目前共 {len(self.coa_paths)} 張待處理。")

    def import_from_excel(self):
        filepaths = filedialog.askopenfilenames(
            title="選擇要匯入的 Excel 或 CSV 檔案 (可多選)",
            filetypes=[("Excel & CSV files", "*.xlsx *.xls *.csv")]
        )
        if not filepaths:
            return
            
        if not hasattr(self, "imported_lorry_files"):
            self.imported_lorry_files = []
        for fp in filepaths:
            fn = os.path.basename(fp).lower()
            if fp.lower().endswith(('.xlsx', '.xls')) and ('chemical' in fn or 'lorry' in fn or '勝一' in fn):
                if fp not in self.imported_lorry_files:
                    self.imported_lorry_files.append(fp)
        self.update_lorry_status()

        total_imported = 0
        for filepath in filepaths:
            try:
                records = []
                sheet_count = 1

                if filepath.lower().endswith('.csv'):
                    rows = []
                    try:
                        with open(filepath, 'r', encoding='utf-8-sig') as f:
                            reader = csv.reader(f)
                            rows = list(reader)
                    except UnicodeDecodeError:
                        with open(filepath, 'r', encoding='cp950') as f:
                            reader = csv.reader(f)
                            rows = list(reader)

                    # Check for vertical CSV format (e.g. RawLotId in first column)
                    is_vertical_csv = False
                    for r_idx, row_vals in enumerate(rows[:20]):
                        if len(row_vals) >= 2 and str(row_vals[0]).strip() == "RawLotId":
                            is_vertical_csv = True
                            break

                    if is_vertical_csv:
                        batch_val, date_val, loc_val = "", "", ""
                        for r_idx, row_vals in enumerate(rows[:30]):
                            if len(row_vals) >= 2:
                                key = str(row_vals[0]).strip()
                                val = str(row_vals[1]).strip()
                                if key == "RawLotId": batch_val = val
                                elif key == "DeliverDate": date_val = val
                        fname = os.path.basename(filepath)
                        m = re.search(r'\s([A-Za-z0-9]+)_\d+\.csv$', fname, re.IGNORECASE)
                        if m:
                            loc_val = m.group(1)
                        else:
                            parts = fname.split('_')
                            if len(parts) >= 2:
                                loc_val = parts[-2].split(' ')[-1]
                        if batch_val:
                            norm_d = normalize_date_str(date_val)
                            records.append({
                                "sheet": "CSV",
                                "batch": batch_val,
                                "loc": clean_location_str(loc_val, self.mapping_dict),
                                "date": norm_d,
                                "time": "",
                                "mod_time": "",
                                "tank": get_tank_from_batch(batch_val)
                            })
                    else:
                        # Standard horizontal CSV
                        batch_col, loc_col, date_col, tank_col, time_col, mod_time_col = -1, -1, -1, -1, -1, -1
                        start_row = 0
                        for r_idx in range(min(15, len(rows))):
                            row = rows[r_idx]
                            if not row: continue
                            for c_idx, val in enumerate(row):
                                v = str(val or "").strip().upper()
                                if batch_col == -1 and any(k in v for k in ["批號", "BATCH", "LOT"]): batch_col = c_idx
                                if loc_col == -1 and any(k in v for k in ["地點", "指送", "交貨", "到貨地", "送達", "廠區", "LOCATION", "DEST"]): loc_col = c_idx
                                if date_col == -1 and any(k in v for k in ["到貨日", "出貨日", "出車日", "日期", "DATE"]) and "地" not in v and "點" not in v: date_col = c_idx
                                if tank_col == -1 and any(k in v for k in ["槽號", "槽車", "TANK"]) and not any(k in v for k in ["日", "期", "時間", "TIME", "DATE", "到廠", "出車", "出貨"]): tank_col = c_idx
                                if time_col == -1 and any(k in v for k in ["到貨時間", "預計", "時間", "TIME"]) and "修正" not in v: time_col = c_idx
                                if mod_time_col == -1 and "修正" in v and ("時間" in v or "TIME" in v): mod_time_col = c_idx
                            if batch_col != -1 and (loc_col != -1 or date_col != -1):
                                start_row = r_idx + 1
                                break
                        if batch_col == -1 or loc_col == -1:
                            batch_col, date_col, tank_col, loc_col = 2, 1, 3, 4
                            start_row = 2
                        for r_idx in range(start_row, len(rows)):
                            row = rows[r_idx]
                            if not row: continue
                            def get_c(c): return row[c] if c != -1 and c < len(row) else None
                            b_val = str(get_c(batch_col) or "").strip().upper()
                            l_val = str(get_c(loc_col) or "").strip().upper()
                            d_val = get_c(date_col)
                            t_val = str(get_c(tank_col) or "").strip()
                            tm_val = str(get_c(time_col) or "").strip()
                            mt_val = str(get_c(mod_time_col) or "").strip()
                            if len(b_val) != 10 or not re.search(r'[0-9]', b_val):
                                for cell in row:
                                    cs = str(cell or "").strip().upper()
                                    if len(cs) == 10 and re.search(r'[0-9]', cs) and re.search(r'[A-Z]', cs) and "/" not in cs and "-" not in cs:
                                        b_val = cs
                                        break
                            if len(b_val) == 10 and re.search(r'[0-9]', b_val):
                                is_valid_tank = (
                                    t_val and 
                                    len(t_val) <= 6 and 
                                    not any(c in t_val for c in ["-", "/", ":", " "]) and
                                    not (len(t_val) > 4 and t_val.isdigit())
                                )
                                tank_final = t_val if is_valid_tank else get_tank_from_batch(b_val)
                                clean_l = clean_location_str(l_val, self.mapping_dict)
                                records.append({
                                    "sheet": "CSV",
                                    "batch": b_val,
                                    "tank": tank_final,
                                    "loc": clean_l,
                                    "long_code": self.mapping_dict.get(clean_l, ""),
                                    "date": normalize_date_str(d_val),
                                    "time": tm_val,
                                    "mod_time": mt_val
                                })
                else:
                    # 遍歷 Excel 所有分頁 (跨分頁抓取所有有效排程)
                    wb = openpyxl.load_workbook(filepath, data_only=True)
                    sheet_count = len(wb.worksheets)

                    for ws in wb.worksheets:
                        sheet_name = ws.title
                        rows = list(ws.iter_rows(values_only=True))
                        if not rows or len(rows) == 0:
                            continue

                        # 檢查分頁全域文字是否標記為台積電
                        sheet_has_tsmc = False
                        for r_idx in range(min(5, len(rows))):
                            row_str = " ".join(str(cell or "") for cell in rows[r_idx]).upper()
                            if "TSMC" in row_str or "台積" in row_str:
                                sheet_has_tsmc = True
                                break

                        # 動態掃描前 15 列尋找標題欄位
                        batch_col = -1
                        loc_col = -1
                        date_col = -1
                        tank_col = -1
                        time_col = -1
                        mod_time_col = -1
                        cust_col = -1
                        start_row = 0

                        for r_idx in range(min(15, len(rows))):
                            row = rows[r_idx]
                            if not row: continue
                            for c_idx, val in enumerate(row):
                                v = str(val or "").strip().upper()
                                if not v: continue
                                if batch_col == -1 and any(k in v for k in ["批號", "BATCH", "LOT"]): batch_col = c_idx
                                if loc_col == -1 and any(k in v for k in ["地點", "指送", "交貨", "到貨地", "送達", "廠區", "LOCATION", "DEST"]): loc_col = c_idx
                                if date_col == -1 and (v in ["到貨", "到貨日", "日期", "出車"] or any(k in v for k in ["到貨日", "出貨日", "出車日", "日期", "DATE"])) and "地" not in v and "點" not in v: date_col = c_idx
                                if tank_col == -1 and any(k in v for k in ["槽號", "槽車", "TANK"]) and not any(k in v for k in ["日", "期", "時間", "TIME", "DATE", "到廠", "出車", "出貨"]): tank_col = c_idx
                                if time_col == -1 and any(k in v for k in ["到貨時間", "預計", "時間", "TIME"]) and "修正" not in v: time_col = c_idx
                                if mod_time_col == -1 and "修正" in v and ("時間" in v or "TIME" in v): mod_time_col = c_idx
                                if cust_col == -1 and any(k in v for k in ["對象", "客戶", "廠商", "CUSTOMER"]): cust_col = c_idx

                            if batch_col != -1 and (loc_col != -1 or date_col != -1):
                                start_row = r_idx + 1
                                break

                        if batch_col == -1 or loc_col == -1:
                            batch_col = 2
                            date_col = 1
                            tank_col = 3
                            loc_col = 4
                            start_row = 2

                        for r_idx in range(start_row, len(rows)):
                            row = rows[r_idx]
                            if not row or len(row) == 0: continue

                            def get_cell_val(c):
                                return row[c] if c != -1 and c < len(row) else None

                            b_val = str(get_cell_val(batch_col) or "").strip().upper()
                            l_val = str(get_cell_val(loc_col) or "").strip().upper()
                            d_val = get_cell_val(date_col)
                            t_val = str(get_cell_val(tank_col) or "").strip()
                            time_val = str(get_cell_val(time_col) or "").strip()
                            mt_val = str(get_cell_val(mod_time_col) or "").strip()
                            cust_val = str(get_cell_val(cust_col) or "").strip()

                            # 若預設欄位非 10 碼批號，全列搜尋 10 碼英數混合批號
                            if len(b_val) != 10 or not any(c.isdigit() for c in b_val):
                                for cell in row:
                                    cell_str = str(cell or "").strip().upper()
                                    if len(cell_str) == 10 and any(c.isdigit() for c in cell_str) and any(c.isalpha() for c in cell_str) and "/" not in cell_str and "-" not in cell_str:
                                        b_val = cell_str
                                        break

                            if len(b_val) != 10 or not any(c.isdigit() for c in b_val):
                                continue

                            # 排除非台積電客戶 (例如南亞、長春、聯電、日月光)
                            if cust_val and any(non in cust_val for non in ["南亞", "長春", "聯電", "日月光"]) and "台積" not in cust_val:
                                continue

                            clean_loc = clean_location_str(l_val, self.mapping_dict)
                            is_tsmc = sheet_has_tsmc or ("台積" in cust_val if cust_val else False) or \
                                      any(k in l_val for k in ["台積", "18P", "15P", "12P", "14P", "AP", "F"]) or \
                                      (clean_loc in self.mapping_dict)
                            if not is_tsmc and cust_val and "台積" not in cust_val:
                                continue

                            norm_date = normalize_date_str(d_val)
                            
                            # 槽號驗證：槽號為 3~6 碼英數代號（如 E319、E308、E29J），排除誤抓之日期或時間字串
                            is_valid_tank = (
                                t_val and 
                                len(t_val) <= 6 and 
                                not any(sep in t_val for sep in ["-", "/", ":", " "]) and 
                                not (len(t_val) > 4 and t_val.isdigit())
                            )
                            if not is_valid_tank:
                                t_val = get_tank_from_batch(b_val)

                            # 格式化時間 (如 0830 -> 08:30)
                            t_final = time_val
                            if isinstance(d_val, datetime) and not t_final:
                                hm = d_val.strftime("%H:%M")
                                if hm != "00:00": t_final = hm
                            elif t_final and len(t_final) == 4 and t_final.isdigit():
                                t_final = f"{t_final[:2]}:{t_final[2:]}"

                            records.append({
                                "sheet": sheet_name,
                                "batch": b_val,
                                "tank": t_val,
                                "loc": clean_loc,
                                "long_code": self.mapping_dict.get(clean_loc, ""),
                                "date": norm_date,
                                "time": t_final,
                                "mod_time": mt_val
                            })
                    wb.close()

                if not records:
                    messagebox.showinfo("匯入提示", f"在檔案 {os.path.basename(filepath)} 的所有 {sheet_count} 個分頁中，找不到任何有效的台積電排程資料！")
                    continue

                dialog = ImportRangeDialog(self, records, os.path.basename(filepath), sheet_count=sheet_count)
                self.wait_window(dialog)

                if not dialog.selected_records:
                    continue

                target_records = dialog.selected_records

                start_idx = 0
                for idx, entry in enumerate(self.entries):
                    if not entry["batch_var"].get().strip():
                        start_idx = idx
                        break
                else:
                    start_idx = len(self.entries)
                    
                needed_rows = start_idx + len(target_records)
                if needed_rows > len(self.entries):
                    self.add_input_rows(needed_rows - len(self.entries))
                    
                for i, rec in enumerate(target_records):
                    row_e = self.entries[start_idx + i]
                    row_e["batch_var"].set(rec.get("batch", ""))
                    if "tank_var" in row_e and rec.get("tank"):
                        row_e["tank_var"].set(rec["tank"])
                    row_e["loc_var"].set(rec.get("loc", ""))
                    if rec.get("date"): row_e["date_var"].set(rec["date"])
                    if rec.get("time"): row_e["time_var"].set(rec["time"])
                    if rec.get("mod_time"): row_e["mod_time_var"].set(rec["mod_time"])
                    
                total_imported += len(target_records)
                
            except Exception as e:
                messagebox.showerror("匯入錯誤", f"解析檔案 {os.path.basename(filepath)} 時發生錯誤:\n{e}")
                
        if total_imported > 0:
            messagebox.showinfo("匯入成功", f"成功從所選檔案匯入 {total_imported} 筆！")

    def load_existing_transport_notice(self):
        """
        讓人員手動選擇要修訂的『運輸通知表.xlsx』或資料夾，預設開啟當天的輸出資料夾。
        還原原本的所有位置 (第1~N列)，讓人員直接於指定列輸入『修正到廠時間』！
        """
        # 預設開啟今天的輸出資料夾
        today_dir_name = f"三合一單輸出_{datetime.now().strftime('%Y%m%d')}"
        today_dir = os.path.join(self.base_dir, today_dir_name)
        initial_dir = today_dir if os.path.exists(today_dir) else self.base_dir

        filepath = filedialog.askopenfilename(
            initialdir=initial_dir,
            title="選擇要修訂的運輸通知表 Excel 檔案",
            filetypes=[("Excel 運輸通知表", "*.xlsx *.xls")]
        )
        if not filepath:
            return

        records = []
        target_dir = os.path.dirname(filepath)
        folder_name = os.path.basename(target_dir)

        # 1. 優先嘗試讀取該輸出資料夾中的 session.json 快取（包含完整 10 碼批號）
        session_file = os.path.join(target_dir, "session.json")
        if os.path.exists(session_file):
            try:
                with open(session_file, "r", encoding="utf-8") as f:
                    records = json.load(f)
            except Exception:
                records = []

        # 2. 若該目錄無 session.json 快取，則直接精準解析選取的 Excel 運輸通知表卡片
        if not records:
            try:
                wb = openpyxl.load_workbook(filepath, data_only=True)
                ws = wb.active

                curr_r = 1
                while curr_r <= ws.max_row:
                    d_val = ws.cell(row=curr_r, column=6).value
                    t_val = ws.cell(row=curr_r+2, column=6).value
                    mt_val = ws.cell(row=curr_r+3, column=6).value or ws.cell(row=curr_r+3, column=13).value
                    l_val = ws.cell(row=curr_r+1, column=4).value
                    tank_val = ws.cell(row=curr_r+1, column=5).value

                    l_str = str(l_val).replace("台積電", "").replace("台積", "").replace("新竹", "").replace("台中", "").replace("台南", "").replace("廠", "").replace("\n", "").replace("\r", "").strip() if l_val else ""
                    d_pure, t_pure = split_date_and_time(d_val)
                    t_final = str(t_val).strip() if t_val else t_pure

                    clean_tank = str(tank_val).replace("None", "").strip() if tank_val else ""
                    if clean_tank in ("槽號", "4300", "6 支"):
                        clean_tank = ""

                    if d_pure or l_str or clean_tank:
                        clean_t = t_final if t_final not in ("4300", "6 支", "None") else ""
                        clean_mt = str(mt_val).strip() if mt_val and str(mt_val).strip() not in ("4300", "6 支", "None") else ""
                        records.append({
                            "batch": "",
                            "tank": clean_tank,
                            "loc": l_str,
                            "date": d_pure,
                            "time": clean_t,
                            "mod_time": clean_mt
                        })
                    curr_r += 7
                wb.close()
            except Exception as e:
                messagebox.showerror("載入失敗", f"讀取既有運輸通知表失敗:\n{e}")
                return

        if not records:
            messagebox.showwarning("提示", "選取的檔案中無任何可還原的運輸通知表卡片！")
            return

        # 清空目前表格
        for entry in self.entries:
            entry["batch_var"].set("")
            entry["loc_var"].set("")
            entry["long_code_var"].set("")
            entry["tank_var"].set("")
            entry["date_var"].set("")
            entry["time_var"].set("")
            entry["mod_time_var"].set("")

        needed_rows = len(records)
        if needed_rows > len(self.entries):
            self.add_input_rows(needed_rows - len(self.entries))

        for idx, rec in enumerate(records):
            row_e = self.entries[idx]
            if rec.get("batch"): row_e["batch_var"].set(rec["batch"])
            if rec.get("loc"): row_e["loc_var"].set(rec["loc"])
            if rec.get("tank"): row_e["tank_var"].set(rec["tank"])
            if rec.get("date"): row_e["date_var"].set(rec["date"])
            if rec.get("time"): row_e["time_var"].set(rec["time"])
            if rec.get("mod_time"): row_e["mod_time_var"].set(rec["mod_time"])

        messagebox.showinfo(
            "還原成功", 
            f"已成功從『{os.path.basename(filepath)}』載入 {len(records)} 筆既有排程紀錄！\n\n"
            f"所有排程已按原始位置 (第 1~{len(records)} 列) 精準對齊。\n"
            f"請直接在需要修正的排程列填寫【修正到廠時間】即可！"
        )

    def generate_files(self):
        if not os.path.exists(self.template_path):
            messagebox.showerror("錯誤", f"找不到範本檔案:\n{self.template_path}")
            return
        if not os.path.exists(self.mapping_path):
            messagebox.showerror("錯誤", f"找不到對照表檔案:\n{self.mapping_path}")
            return
            
        do_3in1 = self.gen_3in1_var.get()
        do_transport = self.gen_transport_var.get()
        
        if not do_3in1 and not do_transport:
            messagebox.showwarning("提示", "請至少勾選一種報表類型（三合一單 或 運輸通知表）！")
            return

        valid_data = []
        for idx, row in enumerate(self.entries):
            if not row["chk_var"].get():
                continue
                
            batch = row["batch_var"].get().strip().upper()
            loc = row["loc_var"].get().strip().upper()
            tank = row["tank_var"].get().strip()
            date_str = row["date_var"].get().strip()
            time_str = row["time_var"].get().strip()
            mod_time_str = row["mod_time_var"].get().strip()
            
            if not batch and not loc:
                continue
            if not batch or not loc:
                messagebox.showerror("錯誤", f"第 {idx+1} 項資料不齊全！")
                return
            if len(batch) != 10:
                messagebox.showerror("錯誤", f"第 {idx+1} 項的批號長度錯誤！\n批號必須剛好 10 碼，目前輸入: {batch} (長度 {len(batch)})")
                return
                
            valid_data.append({
                "batch": batch,
                "tank": tank,
                "loc": loc,
                "date": date_str,
                "time": time_str,
                "mod_time": mod_time_str
            })
            
        if not valid_data:
            messagebox.showinfo("提示", "請至少勾選並輸入一筆有效資料！")
            return

        missing_locs = [data["loc"] for data in valid_data if data["loc"] not in self.mapping_dict]
        if missing_locs:
            missing_str = ", ".join(set(missing_locs))
            messagebox.showerror("錯誤", f"地點代號對照表中找不到以下地點：\n{missing_str}\n\n請先更新對照表後再試！")
            return

        output_dir = os.path.join(self.base_dir, f"三合一單輸出_{datetime.now().strftime('%Y%m%d')}")
        os.makedirs(output_dir, exist_ok=True)
        
        success_3in1 = 0
        error_msgs = []
        mat_no = "L12C53161"

        if do_3in1:

            coa_crops = {}
            if os.path.exists(r'C:\Program Files\Tesseract-OCR\tesseract.exe'):
                pytesseract.pytesseract.tesseract_cmd = r'C:\Program Files\Tesseract-OCR\tesseract.exe'
            else:
                pytesseract.pytesseract.tesseract_cmd = r'C:\Program Files (x86)\Tesseract-OCR\tesseract.exe' 
            if self.coa_paths:
                for img_path in self.coa_paths:
                    try:
                        orig_img = PILImage.open(img_path)
                        img = orig_img.resize((orig_img.width * 2, orig_img.height * 2), PILImage.Resampling.LANCZOS)
                        d = pytesseract.image_to_data(img, output_type=Output.DICT)
                        
                        header_bottom = int(img.height * 0.4)
                        for i in range(len(d['text'])):
                            if 'Batch' in d['text'][i] or 'ID' in d['text'][i] or 'No' in d['text'][i]:
                                header_bottom = d['top'][i] + d['height'][i] + 12
                                break
                                
                        img_top = img.crop((0, 0, img.width, header_bottom))
                        
                        for i in range(len(d['text'])):
                            text = d['text'][i].strip()
                            digits = ''.join(c for c in text if c.isdigit())
                            if len(digits) >= 8:
                                batch_digits = digits
                                row_top = max(0, d['top'][i] - 14)
                                row_bottom = min(img.height, d['top'][i] + d['height'][i] + 16)
                                
                                img_row = img.crop((0, row_top, img.width, row_bottom))
                                new_img = PILImage.new('RGB', (img.width, img_top.height + img_row.height))
                                new_img.paste(img_top, (0, 0))
                                new_img.paste(img_row, (0, img_top.height))
                                
                                coa_crops[batch_digits] = new_img
                    except Exception as e:
                        print("OCR error:", e)

            for data in valid_data:
                batch_no = data["batch"]
                tank_no = data["tank"]
                loc = data["loc"]
                loc_code = self.mapping_dict[loc]
                
                try:
                    wb = openpyxl.load_workbook(self.template_path)
                    ws = wb.worksheets[0]
                    
                    tank_row = find_row_by_label(ws, ['槽號']) or 5
                    batch_row = find_row_by_label(ws, ['批號']) or 7
                    loc_row = find_row_by_label(ws, ['送達地點', '地點']) or 11
                    mat_row = find_row_by_label(ws, ['料號']) or 3
                    sup_row = find_row_by_label(ws, ['供應商']) or 9
                    
                    raw_mat = str(ws.cell(row=mat_row, column=3).value or "").strip()
                    if raw_mat.startswith("4"):
                        mat_no = raw_mat[1:]
                    elif raw_mat:
                        mat_no = raw_mat
                    
                    final_tank_no = "5" + tank_no
                    final_batch_no = "6" + batch_no
                    
                    ws.cell(row=tank_row, column=3).value = final_tank_no
                    ws.cell(row=batch_row, column=3).value = final_batch_no
                    ws.cell(row=loc_row, column=3).value = loc_code
                    
                    images_to_keep = []
                    for img in ws._images:
                        if img.height < 150 and img.width > 200:
                            images_to_keep.append(img)
                    ws._images = images_to_keep
                    
                    c3_val = ws.cell(row=mat_row, column=3).value or ""
                    c6_val = ws.cell(row=sup_row, column=3).value or ""
                    qr_str = f"||{c3_val}||{final_tank_no}||{final_batch_no}||{c6_val}||{loc_code}"
                    
                    qr = qrcode.QRCode(box_size=4, border=2)
                    qr.add_data(qr_str)
                    qr.make(fit=True)
                    raw_img = qr.make_image(fill_color="black", back_color="white").convert('RGB')
                    
                    offset_x = 35
                    offset_y = 45
                    new_width = raw_img.width + offset_x
                    new_height = raw_img.height + offset_y
                    img_qr = Image.new('RGBA', (new_width, new_height), (255,255,255,0))
                    img_qr.paste(raw_img, (offset_x, offset_y))
                    
                    img_byte_arr = BytesIO()
                    img_qr.save(img_byte_arr, format='PNG')
                    img_byte_arr.seek(0)
                    
                    new_qr = OpenpyxlImage(img_byte_arr)
                    new_qr.anchor = 'F2'
                    ws.add_image(new_qr)
                    
                    safe_loc = "".join(c for c in loc if c.isalnum() or c in (' ', '_', '-')).rstrip()
                    if not safe_loc:
                        safe_loc = "未命名地點"
                    
                    # 產生檔名規格：[出貨日期]. [地點]台積電槽車barcode三合一單.xlsx (例如: 2026.8.18. 18P3B台積電槽車barcode三合一單.xlsx)
                    date_raw = data.get("date", "").strip()
                    dt_file = None
                    if date_raw:
                        date_part = date_raw.split()[0]
                        for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d", "%Y%m%d", "%m/%d/%Y", "%d/%m/%Y"):
                            try:
                                dt_file = datetime.strptime(date_part, fmt)
                                break
                            except ValueError:
                                pass
                    if not dt_file:
                        dt_file = datetime.now()
                        

                    # Insert COA Crop
                    found_coa = False
                    user_digits = ''.join(c for c in batch_no if c.isdigit())
                    for k_batch, crop_img in coa_crops.items():
                        if user_digits in k_batch or k_batch in user_digits:
                            img_byte_arr2 = BytesIO()
                            crop_img.save(img_byte_arr2, format='PNG')
                            img_byte_arr2.seek(0)
                            xl_img = OpenpyxlImage(img_byte_arr2)
                            xl_img.width = 867
                            xl_img.height = 450
                            xl_img.anchor = 'F5'
                            ws.add_image(xl_img)
                            found_coa = True
                            break
                    
                    if not found_coa and self.coa_paths:
                        error_msgs.append(f"⚠️ 警告: 批號 {batch_no} 未在截圖找到，已留白處理！")

                    date_prefix = f"{dt_file.year}.{dt_file.month}.{dt_file.day}. "
                    tank_part = f"{tank_no} " if tank_no else ""
                    base_filename = f"{date_prefix}{tank_part}{safe_loc}台積電槽車barcode三合一單.xlsx"
                    output_path = os.path.join(output_dir, base_filename)
                    counter = 1
                    while os.path.exists(output_path):
                        base_filename = f"{date_prefix}{tank_part}{safe_loc}_{counter}台積電槽車barcode三合一單.xlsx"
                        output_path = os.path.join(output_dir, base_filename)
                        counter += 1
                        
                    output_filename = base_filename
                    wb.save(output_path)
                    wb.close()
                    success_3in1 += 1
                except Exception as e:
                    error_msgs.append(f"處理三合一單 {loc}_{batch_no} 失敗: {e}")

        success_transport = False
        if do_transport:
            try:
                transport_path = os.path.join(output_dir, "運輸通知表.xlsx")
                generate_transport_notice_file(transport_path, valid_data, mat_no=mat_no)
                success_transport = True
            except Exception as e:
                error_msgs.append(f"產生運輸通知表失敗: {e}")

        success_lorry = 0
        if getattr(self, "gen_lorry_var", None) and self.gen_lorry_var.get():
            lorry_sources = list(getattr(self, "imported_lorry_files", []))
            if not lorry_sources:
                lorry_sources = glob.glob(os.path.join(self.base_dir, "Chemical_Lorry*.xlsx"))
            
            for l_path in lorry_sources:
                orig_filename = os.path.splitext(os.path.basename(l_path))[0]
                orig_ext = os.path.splitext(l_path)[1]
                base_lorry_name = orig_filename.rsplit('-', 1)[0] if '-' in orig_filename else orig_filename
                
                try:
                    src_wb_l = openpyxl.load_workbook(l_path, data_only=False)
                    src_ws_l = src_wb_l.active
                    
                    batch_row_map = {}
                    for r in range(7, src_ws_l.max_row + 1):
                        val = str(src_ws_l.cell(row=r, column=1).value or "").strip().upper()
                        if val and val not in batch_row_map:
                            batch_row_map[val] = r
                        
                    for item in valid_data:
                        b_no = item["batch"]
                        l_loc = item["loc"]
                        t_no = item["tank"]
                        d_str = item["date"]
                        
                        matched_r = batch_row_map.get(b_no)
                        if matched_r:
                            wb_l = build_single_row_lorry_workbook(src_ws_l, matched_r)
                            
                            mmdd = "0000"
                            if d_str:
                                for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d", "%Y%m%d", "%m/%d/%Y", "%d/%m/%Y"):
                                    try:
                                        dt_l = datetime.strptime(d_str.split()[0], fmt)
                                        mmdd = f"{dt_l.month:02d}{dt_l.day:02d}"
                                        break
                                    except ValueError:
                                        pass
                            if mmdd == "0000":
                                now_l = datetime.now()
                                mmdd = f"{now_l.month:02d}{now_l.day:02d}"
                                
                            t_part = f"{t_no} " if t_no else ""
                            lorry_out_name = f"{base_lorry_name}-{mmdd} {t_part}{l_loc}{orig_ext}"
                            out_l_path = os.path.join(output_dir, lorry_out_name)
                            wb_l.save(out_l_path)
                            wb_l.close()
                            success_lorry += 1
                    src_wb_l.close()
                except Exception as le:
                    error_msgs.append(f"產生 Chemical_Lorry 失敗: {le}")

        # 自動快取當前 Session 資料至輸出資料夾與根目錄，供往後一鍵精準還原修訂
        try:
            for target_path in (os.path.join(output_dir, "session.json"), os.path.join(self.base_dir, "last_generated_session.json")):
                with open(target_path, "w", encoding="utf-8") as f:
                    json.dump(valid_data, f, ensure_ascii=False, indent=2)
        except Exception:
            pass

        msg_parts = []
        if do_3in1:
            msg_parts.append(f"• 三合一單：成功產生 {success_3in1} 份")
        if do_transport:
            status_str = "成功" if success_transport else "失敗"
            msg_parts.append(f"• 運輸通知表：{status_str} (共 {len(valid_data)} 筆排程卡片)")
        if getattr(self, "gen_lorry_var", None) and self.gen_lorry_var.get() and success_lorry > 0:
            msg_parts.append(f"• 單列 Chemical_Lorry：成功產生 {success_lorry} 份 (已自動對齊第 7 列)")
            
        msg = "\n".join(msg_parts) + f"\n\n檔案已儲存於資料夾：\n{output_dir}"
        
        if error_msgs:
            msg += "\n\n部分錯誤:\n" + "\n".join(error_msgs[:5])
            messagebox.showwarning("完成 (但有部分錯誤)", msg)
        else:
            messagebox.showinfo("成功", msg)
            
        os.startfile(output_dir)

if __name__ == "__main__":
    app = App()
    app.mainloop()
