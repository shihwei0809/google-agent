import os
import sys
import openpyxl
from openpyxl.drawing.image import Image as OpenpyxlImage
import qrcode
from PIL import Image
from datetime import datetime

def load_location_mapping(mapping_path):
    # Returns list of dicts: [{shortName, fullName, code}, ...]
    # Fallback defaults
    defaults = [
        {"shortName": "15P5",  "fullName": "", "code": "E1550155A"},
        {"shortName": "15P6",  "fullName": "", "code": "E1550156A"},
        {"shortName": "18P3B", "fullName": "", "code": "EF180183B"},
        {"shortName": "12P7",  "fullName": "", "code": "E00700001"},
    ]
    result = []
    if os.path.exists(mapping_path):
        try:
            wb = openpyxl.load_workbook(mapping_path, data_only=True)
            ws = wb.active
            for row in ws.iter_rows(values_only=True):
                if not row or not row[0]:
                    continue
                col0 = str(row[0]).strip()
                # Skip header row
                if col0 in ("送達地簡稱", "短地點", "shortName", "SHORT"):
                    continue
                # 3-column format: shortName | fullName | code
                if len(row) >= 3 and row[2]:
                    result.append({
                        "shortName": col0.upper(),
                        "fullName":  str(row[1]).strip() if row[1] else "",
                        "code":      str(row[2]).strip(),
                    })
                # 2-column format (legacy): shortName | code
                elif len(row) >= 2 and row[1]:
                    result.append({
                        "shortName": col0.upper(),
                        "fullName":  "",
                        "code":      str(row[1]).strip(),
                    })
            wb.close()
        except Exception as e:
            print(f"Warning: failed to read location mapping: {e}", file=sys.stderr)
    return result if result else defaults

def extract_tank_from_batch(batch_no: str) -> str:
    batch = batch_no.strip().upper()
    if len(batch) < 10:
        return ""
    if batch.endswith("J1"):
        return batch[5:8]
    return batch[5:9]

def find_matched_loc(destination: str, mappings: list):
    """Return the matched mapping dict or None.
    Checks 送達地簡稱 (shortName), 送達地全名 (fullName), 送達地點代號 (code).
    Any one match is sufficient.
    """
    dest_norm = destination.replace(" ", "").upper()

    # Pass 1: exact/contains match on any of the three fields
    for m in mappings:
        short = m["shortName"].replace(" ", "").upper()
        full  = m["fullName"].replace(" ", "").upper()
        code  = m["code"].replace(" ", "").upper()

        if (short and short in dest_norm) or \
           (full  and full  in dest_norm) or \
           (code  and code  in dest_norm):
            return m

    # Pass 2: partial short-name match (e.g. "18P3" typed as "18廠P3")
    for m in mappings:
        short = m["shortName"].replace(" ", "").upper()
        parts = [short]
        if short.endswith("B"):
            parts.append(short[:-1])
        for p in parts:
            if p in dest_norm or p.replace("P", "廠P") in dest_norm:
                return m

    # Pass 3: loose digit+suffix match
    for m in mappings:
        short = m["shortName"].replace(" ", "").upper()
        if len(short) >= 4:
            prefix = short[:2]
            suffix = short[2:]
            if prefix in dest_norm and (suffix in dest_norm or suffix.replace("P", "") in dest_norm):
                return m

    return None

def main():
    if len(sys.argv) < 3:
        print("Usage: python generate_3in1.py <batch_no> <base64_destination>", file=sys.stderr)
        sys.exit(1)
        
    batch_no = sys.argv[1].strip().upper()
    
    # Decode base64 destination to prevent Windows command line encoding corruption
    import base64
    try:
        destination = base64.b64decode(sys.argv[2].strip()).decode('utf-8')
    except Exception:
        destination = sys.argv[2].strip()
        
    # Auto-generate output path in system temp dir
    import tempfile
    safe_batch = batch_no.replace('/', '-').replace('\\', '-')[:30]
    output_path = os.path.join(tempfile.gettempdir(), f"3in1_{safe_batch}.xlsx")
    
    base_dir = os.path.dirname(os.path.abspath(__file__))
    template_path = os.path.join(base_dir, "台積電槽車barcode三合一單-範本.xlsx")
    mapping_path = os.path.join(base_dir, "地點代號對照表.xlsx")
    
    if not os.path.exists(template_path):
        print(f"Error: Template file not found at {template_path}", file=sys.stderr)
        sys.exit(2)
        
    mapping = load_location_mapping(mapping_path)
    matched_loc = find_matched_loc(destination, mapping)
    
    if not matched_loc:
        print(f"Error: Could not match destination '{destination}' to any TSMC location code in config.", file=sys.stderr)
        sys.exit(3)
        
    loc_code = matched_loc["code"]
    tank_no = extract_tank_from_batch(batch_no)
    tank_with_prefix = "5" + tank_no
    batch_with_prefix = "6" + batch_no
    
    try:
        wb = openpyxl.load_workbook(template_path)
        ws = wb["barcode"] if "barcode" in wb.sheetnames else wb.worksheets[0]
        
        ws['C5'] = tank_with_prefix
        ws['C7'] = batch_with_prefix
        ws['C11'] = loc_code
        
        mat_no = str(ws['C3'].value or "4L12C53161").strip()
        sup_no = str(ws['C9'].value or "375970680").strip()
        qr_str = f"||{mat_no}||{tank_with_prefix}||{batch_with_prefix}||{sup_no}||{loc_code}"
        ws['B20'] = qr_str
        
        # Clear old QR Code images
        images_to_keep = []
        for img in ws._images:
            ratio = (img.width/img.height if img.height>0 else 1)
            if 0.8 < ratio < 1.2 and img.width < 300:
                continue
            images_to_keep.append(img)
        ws._images = images_to_keep
        
        # Generate QR Code
        qr = qrcode.QRCode(box_size=4, border=2)
        qr.add_data(qr_str)
        qr.make(fit=True)
        raw_img = qr.make_image(fill_color="black", back_color="white").convert('RGB')
        
        # Add padding offset
        offset_x, offset_y = 35, 25
        img_qr = Image.new('RGB', (raw_img.width + offset_x, raw_img.height + offset_y), 'white')
        img_qr.paste(raw_img, (offset_x, offset_y))
        
        from io import BytesIO
        img_io = BytesIO()
        img_qr.save(img_io, format='PNG')
        img_io.seek(0)
        
        new_qr = OpenpyxlImage(img_io)
        new_qr.anchor = 'F2'
        ws.add_image(new_qr)
        
        wb.save(output_path)
        wb.close()
        # Print the actual file path so Node.js can read and serve it
        print(output_path)
        
    except Exception as e:
        print(f"Error generating Excel: {e}", file=sys.stderr)
        sys.exit(4)

if __name__ == '__main__':
    main()
