import os
import sys
import re
import json
import csv
import webbrowser
import calendar
from datetime import datetime, timedelta
from io import BytesIO
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import openpyxl
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.cell.rich_text import TextBlock, CellRichText
from openpyxl.cell.text import InlineFont
from PIL import Image as PILImage, Image
import qrcode
import pytesseract
from pytesseract import Output

# ================= 浮動日曆選擇器 =================


# ================= 產品品名與固定充填重量對照表 =================
PRODUCT_WEIGHT_MAP = {
    "IPA": "4300",
    "IPA HQ": "4300",
    "IPAHQ": "4300",
    "SEP73E5": "4300",
    "SEP73E4": "4300",
    "PMAHQ": "4300",
    "PMA": "4300",
    "CPNE4R": "4300",
    "CPNE4": "4300",
    "EDG": "4300",
    "BDGE": "4300",
    "SET100": "4300",
}

def get_product_weight(product_name):
    if not product_name:
        return "4300"
    p = str(product_name).strip()
    return PRODUCT_WEIGHT_MAP.get(p, PRODUCT_WEIGHT_MAP.get(p.upper(), "4300"))

class CalendarDialog(tk.Toplevel):
    def __init__(self, parent, target_var):
        super().__init__(parent)
        self.title("選擇日期")
        self.resizable(False, False)
        self.target_var = target_var
        self.transient(parent)
        self.grab_set()

        now = datetime.now()
        self.year = now.year
        self.month = now.month

        val = target_var.get().strip()
        if val:
            for fmt in ("%Y/%m/%d", "%Y-%m-%d", "%Y/%M/%d", "%Y.%m.%d"):
                try:
                    dt = datetime.strptime(val, fmt)
                    self.year = dt.year
                    self.month = dt.month
                    break
                except ValueError:
                    pass

        try:
            x = parent.winfo_pointerx()
            y = parent.winfo_pointery()
            self.geometry(f"+{x}+{y}")
        except Exception:
            pass

        self.setup_ui()
        
    def setup_ui(self):
        header = tk.Frame(self, pady=5, bg="#F5F5F5")
        header.pack(fill="x")
        
        tk.Button(header, text="◄", command=self.prev_month, width=3, relief="groove").pack(side="left", padx=5)
        self.lbl_month = tk.Label(header, text="", font=("Arial", 11, "bold"), bg="#F5F5F5", width=12)
        self.lbl_month.pack(side="left", expand=True)
        tk.Button(header, text="►", command=self.next_month, width=3, relief="groove").pack(side="right", padx=5)

        week_frame = tk.Frame(self, bg="#FFFFFF")
        week_frame.pack(fill="x", padx=5, pady=(5, 0))
        for w in ["一", "二", "三", "四", "五", "六", "日"]:
            fg_col = "#D32F2F" if w in ("六", "日") else "#333333"
            tk.Label(week_frame, text=w, width=4, font=("Arial", 9, "bold"), fg=fg_col, bg="#FFFFFF").pack(side="left")

        self.grid_frame = tk.Frame(self, bg="#FFFFFF", padx=5, pady=5)
        self.grid_frame.pack()
        
        btn_frame = tk.Frame(self, pady=5, bg="#F5F5F5")
        btn_frame.pack(fill="x")
        tk.Button(btn_frame, text="今天", command=self.select_today, bg="#E3F2FD", relief="groove", font=("Arial", 9)).pack(side="left", padx=5)
        tk.Button(btn_frame, text="明天", command=self.select_tomorrow, bg="#E8F5E9", relief="groove", font=("Arial", 9)).pack(side="left", padx=5)
        tk.Button(btn_frame, text="清除", command=self.clear_date, bg="#FFEBEE", relief="groove", font=("Arial", 9)).pack(side="right", padx=5)
        
        self.render_calendar()

    def render_calendar(self):
        for widget in self.grid_frame.winfo_children():
            widget.destroy()
            
        self.lbl_month.config(text=f"{self.year} 年 {self.month} 月")
        cal = calendar.monthcalendar(self.year, self.month)
        
        for r_idx, week in enumerate(cal):
            for c_idx, day in enumerate(week):
                if day == 0:
                    tk.Label(self.grid_frame, text="", width=4, bg="#FFFFFF").grid(row=r_idx, column=c_idx)
                else:
                    date_str = f"{self.year}/{self.month:02d}/{day:02d}"
                    fg_col = "#D32F2F" if c_idx >= 5 else "#000000"
                    btn = tk.Button(
                        self.grid_frame, 
                        text=str(day), 
                        width=4, 
                        bg="#F9F9F9",
                        fg=fg_col,
                        relief="flat",
                        command=lambda d=date_str: self.set_date(d)
                    )
                    btn.grid(row=r_idx, column=c_idx, padx=1, pady=1)

    def prev_month(self):
        if self.month == 1:
            self.month = 12
            self.year -= 1
        else:
            self.month -= 1
        self.render_calendar()

    def next_month(self):
        if self.month == 12:
            self.month = 1
            self.year += 1
        else:
            self.month += 1
        self.render_calendar()

    def select_today(self):
        t = datetime.now()
        self.set_date(f"{t.year}/{t.month:02d}/{t.day:02d}")

    def select_tomorrow(self):
        t = datetime.now() + timedelta(days=1)
        self.set_date(f"{t.year}/{t.month:02d}/{t.day:02d}")

    def clear_date(self):
        self.target_var.set("")
        self.destroy()

    def set_date(self, date_str):
        self.target_var.set(date_str)
        self.destroy()

def split_date_and_time(raw_val):
    """
    將 Excel / 貼上讀取的 datetime 或日期時間字串拆分為 (純日期, 純時間)
    例如:
    - datetime.datetime(2026, 8, 14, 0, 0) -> ("2026-08-14", "")
    - "2026-08-14 00:00:00" -> ("2026-08-14", "")
    - "2026-08-14 08:30:00" -> ("2026-08-14", "08:30")
    """
    if raw_val is None:
        return "", ""
    
    if isinstance(raw_val, datetime):
        d_str = f"{raw_val.year:04d}-{raw_val.month:02d}-{raw_val.day:02d}"
        if raw_val.hour == 0 and raw_val.minute == 0 and raw_val.second == 0:
            t_str = ""
        else:
            t_str = f"{raw_val.hour:02d}:{raw_val.minute:02d}"
        return d_str, t_str

    val_str = str(raw_val).strip()
    if not val_str:
        return "", ""

    parts = val_str.split()
    if len(parts) >= 2:
        d_part = parts[0]
        t_part = parts[1]
        if t_part in ("00:00:00", "00:00", "0:00", "00:00:00.000"):
            return d_part, ""
        else:
            t_sub = t_part.split(':')
            if len(t_sub) >= 2:
                return d_part, f"{t_sub[0]:0>2}:{t_sub[1]:0>2}"
            return d_part, t_part

    return val_str, ""

class ImportRangeDialog(tk.Toplevel):
    """
    Excel 匯入筆數與範圍選擇對話框（預設以「當天日期加二天 (今天~後天)」為優先智慧顯示）
    """
    def __init__(self, parent, records, filename=""):
        super().__init__(parent)
        self.title("📥 勝一訂單匯入選擇 (優先顯示當天至+2天排程)")
        self.geometry("820x640")
        self.configure(padx=15, pady=15)
        self.transient(parent)
        self.grab_set()

        self.records = records
        self.total_count = len(records)
        self.selected_records = None
        self.current_filtered_records = []

        # 計算當天日期 + 2 天 (今天、明天、後天)
        now_dt = datetime.now()
        self.d0_str = now_dt.strftime('%Y-%m-%d')
        self.d1_str = (now_dt + timedelta(days=1)).strftime('%Y-%m-%d')
        self.d2_str = (now_dt + timedelta(days=2)).strftime('%Y-%m-%d')
        self.target_window = [self.d0_str, self.d1_str, self.d2_str]

        # 檢查檔案內是否有符合當天至+2天的記錄
        matched_window = [r for r in self.records if self.normalize_date_str(r.get('date')) in self.target_window]
        
        # 若當前日期完全無匹配（如為歷史範例檔），自動尋找檔案中最大日期為基準的 3 天區間
        if not matched_window:
            valid_dates = []
            for r in self.records:
                nd = self.normalize_date_str(r.get('date'))
                if nd:
                    try:
                        valid_dates.append(datetime.strptime(nd, '%Y-%m-%d'))
                    except Exception:
                        pass
            if valid_dates:
                max_d = max(valid_dates)
                self.d0_str = (max_d - timedelta(days=2)).strftime('%Y-%m-%d')
                self.d1_str = (max_d - timedelta(days=1)).strftime('%Y-%m-%d')
                self.d2_str = max_d.strftime('%Y-%m-%d')
                self.target_window = [self.d0_str, self.d1_str, self.d2_str]
                matched_window = [r for r in self.records if self.normalize_date_str(r.get('date')) in self.target_window]

        self.matched_window_records = matched_window

        # 預設選取筆數模式
        self.mode_var = tk.StringVar(value="window" if self.matched_window_records else "count")
        self.count_var = tk.IntVar(value=min(20, self.total_count))

        self.setup_ui(filename)
        self.apply_filter()

    def normalize_date_str(self, d_val):
        if not d_val:
            return ""
        s = str(d_val).strip().replace('/', '-').replace('.', '-')
        parts = s.split(' ')[0].split('-')
        if len(parts) == 3:
            try:
                y, m, d = int(parts[0]), int(parts[1]), int(parts[2])
                return f"{y:04d}-{m:02d}-{d:02d}"
            except Exception:
                pass
        return s

    def setup_ui(self, filename):
        # 頂部提示資訊
        info_frame = tk.LabelFrame(self, text="檔案與排程日期偵測", font=("Arial", 10, "bold"), padx=10, pady=8, fg="#002060")
        info_frame.pack(fill="x", pady=(0, 8))

        tk.Label(info_frame, text=f"📄 檔案名稱：{filename} | 📊 總計台積電排程：{self.total_count} 筆", font=("Arial", 9, "bold")).pack(anchor="w")
        
        window_label = f"🎯 優先篩選區間（當天至+2天）：{self.d0_str} 至 {self.d2_str} (共 {len(self.matched_window_records)} 筆排程)"
        tk.Label(info_frame, text=window_label, fg="#2E7D32", font=("Arial", 10, "bold")).pack(anchor="w", pady=(2, 0))

        # 篩選與快速選擇控制區
        ctrl_frame = tk.LabelFrame(self, text="⚡ 快速選擇匯入範圍 (預設優先顯示當天至+2天)", font=("Arial", 10, "bold"), padx=10, pady=8, fg="#C00000")
        ctrl_frame.pack(fill="x", pady=(0, 8))

        # 第一列：依日期快速篩選
        date_bar = tk.Frame(ctrl_frame)
        date_bar.pack(fill="x", pady=(0, 6))

        tk.Label(date_bar, text="【日期優先】", font=("Arial", 9, "bold"), fg="#1B5E20").pack(side="left")
        
        btn_win = tk.Button(
            date_bar, 
            text=f"🌟 當天至+2天 ({len(self.matched_window_records)} 筆)", 
            command=lambda: self.set_mode("window"), 
            font=("Arial", 9, "bold"), 
            bg="#E8F5E9", 
            fg="#2E7D32", 
            padx=10,
            relief="groove",
            cursor="hand2"
        )
        btn_win.pack(side="left", padx=4)

        tk.Button(
            date_bar, 
            text=f"今天 ({self.d0_str})", 
            command=lambda: self.set_mode("d0"), 
            font=("Arial", 9), 
            bg="#F1F8E9", 
            padx=6,
            cursor="hand2"
        ).pack(side="left", padx=3)

        tk.Button(
            date_bar, 
            text=f"明天 ({self.d1_str})", 
            command=lambda: self.set_mode("d1"), 
            font=("Arial", 9), 
            bg="#F1F8E9", 
            padx=6,
            cursor="hand2"
        ).pack(side="left", padx=3)

        tk.Button(
            date_bar, 
            text=f"後天 ({self.d2_str})", 
            command=lambda: self.set_mode("d2"), 
            font=("Arial", 9), 
            bg="#F1F8E9", 
            padx=6,
            cursor="hand2"
        ).pack(side="left", padx=3)

        # 第二列：依倒數筆數選擇
        count_bar = tk.Frame(ctrl_frame)
        count_bar.pack(fill="x")

        tk.Label(count_bar, text="【筆數擷取】", font=("Arial", 9, "bold"), fg="#0D47A1").pack(side="left")

        for n in [5, 10, 20, self.total_count]:
            lbl = f"最新 {n} 筆" if n < self.total_count else f"全部 ({n} 筆)"
            tk.Button(
                count_bar, 
                text=lbl, 
                command=lambda num=n: self.set_count_mode(num), 
                font=("Arial", 9), 
                bg="#E3F2FD", 
                fg="#0D47A1", 
                padx=6,
                cursor="hand2"
            ).pack(side="left", padx=3)

        tk.Label(count_bar, text="自訂倒數筆數：", font=("Arial", 9)).pack(side="left", padx=(10, 2))
        spin = tk.Spinbox(count_bar, from_=1, to=self.total_count, textvariable=self.count_var, width=6, command=lambda: self.set_count_mode(self.count_var.get()), font=("Arial", 9, "bold"))
        spin.pack(side="left")
        spin.bind("<KeyRelease>", lambda e: self.set_count_mode(self.count_var.get()))

        # 預覽表格區
        preview_frame = tk.LabelFrame(self, text="📋 即將匯入台積電排程資料預覽", font=("Arial", 9, "bold"), padx=8, pady=6)
        preview_frame.pack(fill="both", expand=True, pady=(0, 8))

        self.preview_listbox = tk.Listbox(preview_frame, font=("Consolas", 10), selectmode="none", borderwidth=0)
        scroll = ttk.Scrollbar(preview_frame, orient="vertical", command=self.preview_listbox.yview)
        self.preview_listbox.configure(yscrollcommand=scroll.set)

        self.preview_listbox.pack(side="left", fill="both", expand=True)
        scroll.pack(side="right", fill="y")

        # 底部確定按鈕區
        action_frame = tk.Frame(self)
        action_frame.pack(fill="x")

        self.btn_confirm = tk.Button(
            action_frame, 
            text="🚀 確認匯入資料", 
            command=self.confirm_import, 
            bg="#2E7D32", 
            fg="white", 
            font=("Arial", 11, "bold"), 
            pady=6,
            cursor="hand2"
        )
        self.btn_confirm.pack(side="left", fill="x", expand=True, padx=(0, 5))

        tk.Button(
            action_frame, 
            text="❌ 取消", 
            command=self.destroy, 
            bg="#9E9E9E", 
            fg="white", 
            font=("Arial", 10), 
            pady=6,
            width=10,
            cursor="hand2"
        ).pack(side="right", padx=5)

    def set_mode(self, mode):
        self.mode_var.set(mode)
        self.apply_filter()

    def set_count_mode(self, num):
        try:
            num = int(num)
        except Exception:
            num = 10
        num = max(1, min(num, self.total_count))
        self.count_var.set(num)
        self.mode_var.set("count")
        self.apply_filter()

    def apply_filter(self):
        mode = self.mode_var.get()
        
        if mode == "window":
            if self.matched_window_records:
                self.current_filtered_records = list(self.matched_window_records)
                desc = f"「當天至+2天 ({self.d0_str} ~ {self.d2_str})」"
            else:
                self.current_filtered_records = self.records[-10:]
                desc = "「最新 10 筆」"
        elif mode == "d0":
            matched = [r for r in self.records if self.normalize_date_str(r.get('date')) == self.d0_str]
            self.current_filtered_records = matched if matched else self.records[-5:]
            desc = f"「今天 ({self.d0_str})」"
        elif mode == "d1":
            matched = [r for r in self.records if self.normalize_date_str(r.get('date')) == self.d1_str]
            self.current_filtered_records = matched if matched else self.records[-5:]
            desc = f"「明天 ({self.d1_str})」"
        elif mode == "d2":
            matched = [r for r in self.records if self.normalize_date_str(r.get('date')) == self.d2_str]
            self.current_filtered_records = matched if matched else self.records[-5:]
            desc = f"「後天 ({self.d2_str})」"
        else:
            try:
                cnt = int(self.count_var.get())
            except Exception:
                cnt = 10
            cnt = max(1, min(cnt, self.total_count))
            self.current_filtered_records = self.records[-cnt:]
            desc = f"「最新 {cnt} 筆」"

        # 更新預覽列表
        self.preview_listbox.delete(0, tk.END)
        for idx, rec in enumerate(self.current_filtered_records, 1):
            d_norm = self.normalize_date_str(rec.get("date")) or "無日期"
            t_val = rec.get("time", "") or "無時間"
            b_str = (rec.get("batch") or "").ljust(12)
            l_str = (rec.get("loc") or "").ljust(10)
            tk_str = (get_tank_from_batch(rec.get("batch") or "")).ljust(6)
            
            tag = "★優先" if d_norm in self.target_window else "  "
            self.preview_listbox.insert(tk.END, f"{tag} [{idx:02d}] 日期: {d_norm.ljust(10)} | 時間: {t_val.ljust(6)} | 批號: {b_str} | 槽號: {tk_str} | 地點: {l_str}")

        cnt_res = len(self.current_filtered_records)
        self.btn_confirm.config(text=f"🚀 確認匯入 {desc} 共 {cnt_res} 筆排程資料")

    def confirm_import(self):
        self.selected_records = self.current_filtered_records
        self.destroy()

# ================= 核心邏輯 =================

def normalize_location(loc_raw):
    if not loc_raw:
        return ""
    loc = str(loc_raw).strip()
    if re.match(r'^[A-Za-z0-9]+$', loc):
        return loc
    m_ap = re.search(r'(AP\d+(?:[A-Za-z0-9]+)?)', loc, re.IGNORECASE)
    if m_ap:
        return m_ap.group(1).upper()
    m = re.search(r'(?:[Ff])?(\d+[A-Za-z]?|[A-Za-z]\d+)\s*(?:廠)?\s*[-_]?\s*([A-Za-z0-9]+)', loc)
    if m:
        return f"{m.group(1)}{m.group(2)}"
    cleaned = loc.replace('台積電', '').replace('台積', '').replace('廠', '').replace(' ', '').replace('-', '').replace('_', '')
    return cleaned if cleaned else loc

def get_tank_from_batch(batch):
    batch = batch.strip()
    if not batch:
        return ""
    if len(batch) < 9 or len(batch) > 13:
        return "長度錯誤"
    m = re.search(r'^[0-9]{4,5}([A-Za-z]\d{2,3})', batch)
    if m:
        return m.group(1)
    if batch.endswith('J1') or batch.endswith('T1') or batch.endswith('T2'):
        return batch[5:-2]
    return batch[5:9]

def find_row_by_label(ws, labels):
    for r in range(1, 20):
        val = ws.cell(row=r, column=2).value
        if val and isinstance(val, str):
            for label in labels:
                if label in val:
                    return r
    return None

def generate_transport_notice_file(output_path, items, mat_no="L12C53161"):
    is_append = os.path.exists(output_path)
    if is_append:
        wb = openpyxl.load_workbook(output_path)
        ws = wb.active
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "運輸通知表"
        ws.views.sheetView[0].showGridLines = True
        
        # 欄寬設定 (Left Card: A~F, Spacer: G, Right Card: H~M)
        col_widths = {
            'A': 8, 'B': 18, 'C': 18, 'D': 16, 'E': 14, 'F': 16,
            'G': 4,
            'H': 8, 'I': 18, 'J': 18, 'K': 16, 'L': 14, 'M': 16
        }
        for col, width in col_widths.items():
            ws.column_dimensions[col].width = width
        
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    
    fill_yellow = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    fill_green = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    fill_bright_yellow = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    
    openpyxl_font_dark_blue = Font(name="Microsoft JhengHei", color="002060", size=11)
    openpyxl_font_dark_blue_b13 = Font(name="Microsoft JhengHei", color="002060", size=13, bold=True)
    openpyxl_font_strike_blue_b13 = Font(name="Microsoft JhengHei", color="002060", size=13, bold=True, strike=True)
    openpyxl_font_red_b13 = Font(name="Microsoft JhengHei", color="C00000", size=13, bold=True)
    openpyxl_font_dark_blue_b14 = Font(name="Microsoft JhengHei", color="002060", size=14, bold=True)
    
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    weekdays = ["星期一", "星期二", "星期三", "星期四", "星期五", "星期六", "星期日"]

    def render_notice_card(start_r, start_c, is_modified_card, item):
        r1 = start_r
        r2 = start_r + 1
        r3 = start_r + 2
        r4 = start_r + 3
        r5 = start_r + 4
        r6 = start_r + 5
        
        c1 = start_c
        c2 = start_c + 1
        c3 = start_c + 2
        c4 = start_c + 3
        c5 = start_c + 4
        c6 = start_c + 5
        
        for r in range(r1, r6 + 1):
            ws.row_dimensions[r].height = 24 if r >= r3 else 22
            for c in range(c1, c6 + 1):
                cell = ws.cell(row=r, column=c)
                cell.border = thin_border
                cell.font = openpyxl_font_dark_blue
                cell.alignment = align_center

        # --- Title ---
        prod_name = item.get("product", "").strip() or "IPA"
        weight_val = item.get("weight", "").strip() or get_product_weight(prod_name)
        
        ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c3)
        title_suffix = "出貨排程修正通知" if is_modified_card else "出貨排程通知"
        title_rt = CellRichText([
            TextBlock(InlineFont(color="002060", b=True, sz=12, rFont="Microsoft JhengHei"), f"Shiny {prod_name} Lorry\n"),
            TextBlock(InlineFont(color="002060", sz=10, rFont="Microsoft JhengHei"), f"(料號：{mat_no}) {title_suffix}")
        ])
        cell_a1 = ws.cell(row=r1, column=c1)
        cell_a1.value = title_rt
        cell_a1.fill = fill_yellow
        cell_a1.alignment = align_center
        
        cell_d1 = ws.cell(row=r1, column=c4, value="廠區")
        cell_d1.fill = fill_yellow
        cell_d1.font = Font(name="Microsoft JhengHei", color="002060", size=11, bold=True)
        
        cell_e1 = ws.cell(row=r1, column=c5, value="槽號")
        cell_e1.fill = fill_yellow
        cell_e1.font = Font(name="Microsoft JhengHei", color="002060", size=11, bold=True)
        
        date_raw = item.get("date", "").strip()
        formatted_date = ""
        weekday_str = ""
        if date_raw:
            dt = None
            for fmt in ("%Y/%m/%d", "%Y-%m-%d", "%Y/%M/%d", "%Y.%m.%d"):
                try:
                    dt = datetime.strptime(date_raw, fmt)
                    break
                except ValueError:
                    pass
            if dt:
                weekday_str = weekdays[dt.weekday()]
                formatted_date = f"{dt.year}/{dt.month}/{dt.day}"
            else:
                formatted_date = date_raw
        
        cell_f1 = ws.cell(row=r1, column=c6, value=formatted_date)
        cell_f1.fill = fill_green
        cell_f1.font = Font(name="Microsoft JhengHei", color="002060", size=11, bold=True)
        
        cell_f2 = ws.cell(row=r2, column=c6, value=weekday_str)
        cell_f2.fill = fill_green
        cell_f2.font = Font(name="Microsoft JhengHei", color="002060", size=11, bold=True)

        # --- Body ---
        ws.merge_cells(start_row=r3, start_column=c1, end_row=r6, end_column=c1)
        cell_a3 = ws.cell(row=r3, column=c1, value=prod_name)
        cell_a3.font = openpyxl_font_dark_blue_b14
        
        time_val = item.get("time", "").strip()
        mod_time_val = item.get("mod_time", "").strip()
        
        ws.merge_cells(start_row=r3, start_column=c2, end_row=r3, end_column=c3)
        ws.cell(row=r3, column=c2, value="預計到廠時間")
        cell_f3 = ws.cell(row=r3, column=c6, value=time_val)
        
        ws.merge_cells(start_row=r4, start_column=c2, end_row=r4, end_column=c3)
        ws.cell(row=r4, column=c2, value="修正到廠時間")
        cell_f4 = ws.cell(row=r4, column=c6)
        
        if is_modified_card:
            cell_f3.font = openpyxl_font_strike_blue_b13
            cell_f4.value = mod_time_val
            cell_f4.font = openpyxl_font_red_b13
        else:
            cell_f3.font = openpyxl_font_dark_blue_b13
            cell_f4.value = ""
        
        ws.merge_cells(start_row=r5, start_column=c2, end_row=r5, end_column=c3)
        ws.cell(row=r5, column=c2, value="充填數量(KG)")
        ws.cell(row=r5, column=c6, value=weight_val)
        
        ws.merge_cells(start_row=r2, start_column=c4, end_row=r5, end_column=c4)
        full_loc = item.get("loc", "").strip()
        
        d_rt = CellRichText([
            TextBlock(InlineFont(color="002060", b=True, sz=14, rFont="Microsoft JhengHei"), "台積\n"),
            TextBlock(InlineFont(color="C00000", b=True, sz=14, rFont="Microsoft JhengHei"), full_loc)
        ])
        cell_d2 = ws.cell(row=r2, column=c4)
        cell_d2.value = d_rt
        
        ws.merge_cells(start_row=r2, start_column=c5, end_row=r5, end_column=c5)
        cell_e2 = ws.cell(row=r2, column=c5, value=item.get("tank", "").strip())
        cell_e2.font = openpyxl_font_dark_blue_b14
        cell_e2.fill = fill_bright_yellow
        
        ws.merge_cells(start_row=r6, start_column=c2, end_row=r6, end_column=c5)
        cell_b6 = ws.cell(row=r6, column=c2)
        cell_b6.value = CellRichText([
            TextBlock(InlineFont(color="C00000", b=True, sz=11, rFont="Microsoft JhengHei"), "PFA 500ml"),
            TextBlock(InlineFont(color="002060", sz=11, rFont="Microsoft JhengHei"), " 取樣瓶裝原液 8 分滿放置工具箱內")
        ])
        
        cell_f6 = ws.cell(row=r6, column=c6, value="6 支")
        cell_f6.font = Font(name="Microsoft JhengHei", color="002060", size=11, bold=True)

    curr_row = ws.max_row + 2 if is_append and ws.max_row > 1 else 1
    for item in items:
        # 左側：原始出貨排程通知卡片 (Columns A~F)
        render_notice_card(curr_row, 1, False, item)
        
        # 若該筆有輸入「修正到廠時間」，在右側旁 (Columns H~M) 併排產生「出貨排程修正通知」卡片
        if item.get("mod_time", "").strip():
            render_notice_card(curr_row, 8, True, item)
            
        curr_row += 7

    wb.save(output_path)
    wb.close()

# ================= 介面與操作 =================

class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("勝一三合一單產生系統 & 運輸通知表產生器")
        self.geometry("1180x750")
        self.configure(padx=15, pady=15)
        
        self.base_dir = os.path.dirname(os.path.abspath(__file__))
        self.template_path = os.path.join(self.base_dir, "台積電槽車barcode三合一單-範本.xlsx")
        self.mapping_path = os.path.join(self.base_dir, "地點代號對照表.xlsx")
        
        self.mapping_dict = {}
        self.load_mapping()
        
        self.setup_ui()
        self.entries = []
        self.coa_paths = []
        self.add_input_rows(20)

    def load_mapping(self):
        self.mapping_dict = {}
        if os.path.exists(self.mapping_path):
            try:
                map_wb = openpyxl.load_workbook(self.mapping_path, data_only=True)
                map_ws = map_wb.active
                for row in map_ws.iter_rows(values_only=True):
                    if row and len(row) >= 2 and row[0] and row[1]:
                        loc_key = str(row[0]).strip().upper()
                        loc_val = str(row[1]).strip()
                        # 過濾標題列 (例如 短地點, 長代號, 地點)
                        if any(kw in loc_key for kw in ("地點", "代號", "SHORT", "LOCATION", "KEY", "HEADER")):
                            continue
                        self.mapping_dict[loc_key] = loc_val
                map_wb.close()
            except Exception as e:
                pass

        if hasattr(self, "lbl_mapping_status"):
            m_color = "green" if os.path.exists(self.mapping_path) else "red"
            m_text = f"對照表檔案 (地點代號對照表.xlsx): ✅ 已找到 (載入 {len(self.mapping_dict)} 筆代號)" if os.path.exists(self.mapping_path) else "對照表檔案 (地點代號對照表.xlsx): ❌ 未找到"
            self.lbl_mapping_status.config(text=m_text, fg=m_color)

        if hasattr(self, "entries"):
            for entry in self.entries:
                loc_val = entry["loc_var"].get().strip().upper()
                if loc_val:
                    self.on_loc_change(entry["loc_var"], entry["long_code_var"])

    def reload_mapping_with_msg(self):
        """點擊『🔄 重新載入對照表』時執行"""
        self.load_mapping()
        messagebox.showinfo("對照表已更新", f"已重新讀取『地點代號對照表.xlsx』！\n目前共載入 {len(self.mapping_dict)} 筆地點對照碼。\n表格中的地點長代號已同步更新！")

    def open_calendar_dialog(self, target_var):
        CalendarDialog(self, target_var)

    def setup_ui(self):
        # 檔案狀態區
        status_frame = tk.LabelFrame(self, text="系統狀態", font=("Arial", 10, "bold"), padx=10, pady=8)
        status_frame.pack(fill="x", pady=(0, 8))
        
        t_color = "green" if os.path.exists(self.template_path) else "red"
        t_text = "✅ 已找到" if os.path.exists(self.template_path) else "❌ 未找到 (請將檔案放入資料夾)"
        tk.Label(status_frame, text=f"範本檔案 (台積電槽車barcode三合一單-範本.xlsx): {t_text}", fg=t_color).pack(anchor="w")
        
        # 對照表狀態列 + 重新載入按鈕
        map_status_frame = tk.Frame(status_frame)
        map_status_frame.pack(anchor="w", pady=(2, 0))

        m_color = "green" if os.path.exists(self.mapping_path) else "red"
        m_text = f"對照表檔案 (地點代號對照表.xlsx): ✅ 已找到 (載入 {len(self.mapping_dict)} 筆代號)" if os.path.exists(self.mapping_path) else "對照表檔案 (地點代號對照表.xlsx): ❌ 未找到"
        
        self.lbl_mapping_status = tk.Label(map_status_frame, text=m_text, fg=m_color)
        self.lbl_mapping_status.pack(side="left")

        tk.Button(
            map_status_frame, 
            text="🔄 重新載入對照表", 
            command=self.reload_mapping_with_msg, 
            bg="#607D8B", 
            fg="white", 
            font=("Arial", 8, "bold"), 
            padx=6,
            cursor="hand2"
        ).pack(side="left", padx=10)

        # 頂部快捷批次控制與匯入區
        top_ctrl_frame = tk.Frame(self)
        top_ctrl_frame.pack(fill="x", pady=(0, 8))
        
        tk.Label(top_ctrl_frame, text="提示：可在「批號」貼上多筆資料，或使用右側按鈕匯入 Excel / 還原既有通知表修訂。", fg="#555", justify="left").pack(side="left")
        
        tk.Button(top_ctrl_frame, text="📥 從 Excel 匯入", command=self.import_from_excel, bg="#2196F3", fg="white", font=("Arial", 10, "bold"), padx=10).pack(side="right", padx=(5, 0))
        tk.Button(top_ctrl_frame, text="📂 載入既有『運輸通知表』修訂", command=self.load_existing_transport_notice, bg="#7B1FA2", fg="white", font=("Arial", 10, "bold"), padx=10).pack(side="right", padx=5)
        tk.Button(top_ctrl_frame, text="🖼️ 上傳 COA 截圖", command=self.upload_coa, bg="#FF9800", fg="white", font=("Arial", 10, "bold"), padx=10).pack(side="right", padx=5)
        tk.Button(top_ctrl_frame, text="🗑️ 清除全部資料 (全清)", command=self.clear_all_rows, bg="#D32F2F", fg="white", font=("Arial", 9, "bold"), padx=8).pack(side="right", padx=5)

        # 批次設定預設值列
        batch_setting_frame = tk.LabelFrame(self, text="一鍵批次設定 (日期 / 預計時間 / 修正時間)", font=("Arial", 9, "bold"), padx=8, pady=5)
        batch_setting_frame.pack(fill="x", pady=(0, 8))
        
        # 全選
        self.select_all_var = tk.BooleanVar(value=True)
        tk.Checkbutton(batch_setting_frame, text="全選", variable=self.select_all_var, command=self.toggle_select_all, font=("Arial", 10, "bold")).pack(side="left", padx=(0, 10))
        
        # 日期
        tk.Label(batch_setting_frame, text="批次日期:").pack(side="left")
        self.default_date_var = tk.StringVar(value="")
        date_batch_entry = tk.Entry(batch_setting_frame, textvariable=self.default_date_var, width=12)
        date_batch_entry.pack(side="left", padx=2)
        tk.Button(batch_setting_frame, text="📅", command=lambda: self.open_calendar_dialog(self.default_date_var), font=("Arial", 8), width=3).pack(side="left", padx=(0, 2))
        tk.Button(batch_setting_frame, text="套用至全列", command=self.apply_default_date, bg="#607D8B", fg="white", font=("Arial", 8)).pack(side="left", padx=(2, 12))
        
        # 預計時間
        tk.Label(batch_setting_frame, text="批次預計時間:").pack(side="left")
        self.default_time_var = tk.StringVar(value="")
        tk.Entry(batch_setting_frame, textvariable=self.default_time_var, width=8).pack(side="left", padx=2)
        tk.Button(batch_setting_frame, text="套用至全列", command=self.apply_default_time, bg="#607D8B", fg="white", font=("Arial", 8)).pack(side="left", padx=(2, 12))

        # 修正時間
        tk.Label(batch_setting_frame, text="批次修正時間:").pack(side="left")
        self.default_mod_time_var = tk.StringVar(value="")
        tk.Entry(batch_setting_frame, textvariable=self.default_mod_time_var, width=8).pack(side="left", padx=2)
        tk.Button(batch_setting_frame, text="套用至全列", command=self.apply_default_mod_time, bg="#607D8B", fg="white", font=("Arial", 8)).pack(side="left", padx=(2, 10))

        # 輸出選項
        opt_frame = tk.Frame(batch_setting_frame)
        opt_frame.pack(side="right")
        self.gen_3in1_var = tk.BooleanVar(value=True)
        self.gen_transport_var = tk.BooleanVar(value=True)
        tk.Checkbutton(opt_frame, text="產生三合一單", variable=self.gen_3in1_var, font=("Arial", 9, "bold"), fg="#1B5E20").pack(side="left", padx=5)
        tk.Checkbutton(opt_frame, text="產生運輸通知表", variable=self.gen_transport_var, font=("Arial", 9, "bold"), fg="#0D47A1").pack(side="left", padx=5)

        # 滾動容器
        table_container = tk.Frame(self)
        table_container.pack(fill="both", expand=True)

        self.canvas = tk.Canvas(table_container, borderwidth=0, highlightthickness=0)
        self.scrollbar = ttk.Scrollbar(table_container, orient="vertical", command=self.canvas.yview)
        
        # 單一 Grid 容器 (scrollable_frame)
        self.scrollable_frame = tk.Frame(self.canvas, padx=5, pady=5)

        self.scrollable_frame.bind(
            "<Configure>",
            lambda e: self.canvas.configure(scrollregion=self.canvas.bbox("all"))
        )

        self.canvas.create_window((0, 0), window=self.scrollable_frame, anchor="nw")
        self.canvas.configure(yscrollcommand=self.scrollbar.set)

        self.canvas.pack(side="left", fill="both", expand=True)
        self.scrollbar.pack(side="right", fill="y")

        # --- Row 0: 表格標題列 (置於滾動區最上方，共享 100% 同一 Grid 欄位規格) ---
        headers = [
            (0, "產生"),
            (1, "項次"),
            (2, "批號 (10~11碼)"),
            (3, "槽號 (自動)"),
            (4, "品名 (產品)"),
            (5, "地點 (如 15P5)"),
            (6, "長代號 (自動)"),
            (7, "出貨日期 📅"),
            (8, "預計到廠時間"),
            (9, "修正到廠時間"),
            (10, "單列清空")
        ]
        
        for col_idx, title in headers:
            lbl = tk.Label(
                self.scrollable_frame, 
                text=title, 
                font=("Arial", 10, "bold"), 
                bg="#EAEAEA", 
                fg="#333333",
                padx=6, 
                pady=6,
                relief="groove"
            )
            lbl.grid(row=0, column=col_idx, sticky="ew", padx=1, pady=(0, 6))

        # 產生按鈕
        btn_frame = tk.Frame(self)
        btn_frame.pack(fill="x", pady=10)
        tk.Button(btn_frame, text="🚀 開始批次產生 Excel 報表", command=self.generate_files, bg="#4CAF50", fg="white", font=("Arial", 12, "bold"), pady=8).pack(fill="x")

    def toggle_select_all(self):
        state = self.select_all_var.get()
        for entry in self.entries:
            entry["chk_var"].set(state)

    def apply_default_date(self):
        val = self.default_date_var.get().strip()
        for entry in self.entries:
            if entry["batch_var"].get().strip():
                entry["date_var"].set(val)

    def apply_default_time(self):
        val = self.default_time_var.get().strip()
        for entry in self.entries:
            if entry["batch_var"].get().strip():
                entry["time_var"].set(val)

    def apply_default_mod_time(self):
        val = self.default_mod_time_var.get().strip()
        for entry in self.entries:
            if entry["batch_var"].get().strip():
                entry["mod_time_var"].set(val)

    def add_input_rows(self, count):
        for i in range(count):
            row_idx = len(self.entries) + 1
            row_grid_idx = row_idx
            
            # Col 0: 勾選
            chk_var = tk.BooleanVar(value=True)
            chk = tk.Checkbutton(self.scrollable_frame, variable=chk_var)
            chk.grid(row=row_grid_idx, column=0, padx=2, pady=2)

            # Col 1: 項次
            lbl_num = tk.Label(self.scrollable_frame, text=str(row_idx), font=("Arial", 10), width=3)
            lbl_num.grid(row=row_grid_idx, column=1, padx=2, pady=2)
            
            # Col 2: 批號 (10~11碼)
            batch_var = tk.StringVar()
            batch_entry = tk.Entry(self.scrollable_frame, textvariable=batch_var, width=15, font=("Arial", 10))
            batch_entry.grid(row=row_grid_idx, column=2, padx=2, pady=2, sticky="ew")
            
            # Col 3: 槽號 (自動)
            tank_var = tk.StringVar()
            tank_entry = tk.Entry(self.scrollable_frame, textvariable=tank_var, state="readonly", width=8, font=("Arial", 10), fg="blue")
            tank_entry.grid(row=row_grid_idx, column=3, padx=2, pady=2, sticky="ew")

            # Col 4: 品名 (產品)
            prod_var = tk.StringVar(value="IPA")
            prod_entry = tk.Entry(self.scrollable_frame, textvariable=prod_var, width=11, font=("Arial", 10))
            prod_entry.grid(row=row_grid_idx, column=4, padx=2, pady=2, sticky="ew")
            
            # Col 5: 地點
            loc_var = tk.StringVar()
            loc_entry = tk.Entry(self.scrollable_frame, textvariable=loc_var, width=11, font=("Arial", 10))
            loc_entry.grid(row=row_grid_idx, column=5, padx=2, pady=2, sticky="ew")
            
            # Col 6: 長代號
            long_code_var = tk.StringVar()
            long_code_entry = tk.Entry(self.scrollable_frame, textvariable=long_code_var, state="readonly", width=14, font=("Arial", 9), fg="purple")
            long_code_entry.grid(row=row_grid_idx, column=6, padx=2, pady=2, sticky="ew")
            
            # Col 7: 出貨日期 (Entry + 📅 日曆按鈕)
            date_frame = tk.Frame(self.scrollable_frame)
            date_frame.grid(row=row_grid_idx, column=7, padx=2, pady=2, sticky="ew")
            
            date_var = tk.StringVar(value="")
            date_entry = tk.Entry(date_frame, textvariable=date_var, width=11, font=("Arial", 10))
            date_entry.pack(side="left", fill="x", expand=True)
            
            btn_cal = tk.Button(date_frame, text="📅", command=lambda dv=date_var: self.open_calendar_dialog(dv), font=("Arial", 8), cursor="hand2")
            btn_cal.pack(side="right", padx=(2, 0))
            
            # Col 8: 預計到廠時間
            time_var = tk.StringVar(value="")
            time_entry = tk.Entry(self.scrollable_frame, textvariable=time_var, width=10, font=("Arial", 10))
            time_entry.grid(row=row_grid_idx, column=8, padx=2, pady=2, sticky="ew")

            # Col 9: 修正到廠時間
            mod_time_var = tk.StringVar(value="")
            mod_time_entry = tk.Entry(self.scrollable_frame, textvariable=mod_time_var, width=10, font=("Arial", 10), fg="red")
            mod_time_entry.grid(row=row_grid_idx, column=9, padx=2, pady=2, sticky="ew")

            # Col 10: 單列清空按鈕
            btn_clear_row = tk.Button(
                self.scrollable_frame, 
                text="清空", 
                command=lambda r=row_idx-1: self.clear_single_row(r), 
                font=("Microsoft JhengHei", 9, "bold"), 
                bg="#FFEBEE", 
                fg="#C62828", 
                cursor="hand2", 
                width=6,
                pady=1
            )
            btn_clear_row.grid(row=row_grid_idx, column=10, padx=4, pady=2)

            # 綁定事件
            batch_var.trace_add("write", lambda name, index, mode, bv=batch_var, tv=tank_var: self.on_batch_change(bv, tv))
            loc_var.trace_add("write", lambda name, index, mode, lv=loc_var, lcv=long_code_var: self.on_loc_change(lv, lcv))
            
            for widget in (batch_entry, loc_entry, date_entry, prod_entry):
                widget.bind("<<Paste>>", lambda e, r=row_idx-1, w=widget: self.on_paste(e, r, w))
                widget.bind("<Control-v>", lambda e, r=row_idx-1, w=widget: self.on_paste(e, r, w))
                widget.bind("<Control-V>", lambda e, r=row_idx-1, w=widget: self.on_paste(e, r, w))
            
            self.entries.append({
                "chk_var": chk_var,
                "batch_var": batch_var,
                "tank_var": tank_var,
                "prod_var": prod_var,
                "loc_var": loc_var,
                "long_code_var": long_code_var,
                "date_var": date_var,
                "time_var": time_var,
                "mod_time_var": mod_time_var
            })

    def clear_all_rows(self):
        """一鍵清除全部輸入欄位 (全清)"""
        if messagebox.askyesno("確認清空", "確定要清空所有已填寫的批號、地點與時間資料嗎？"):
            for entry in self.entries:
                entry["batch_var"].set("")
                entry["loc_var"].set("")
                entry["long_code_var"].set("")
                entry["tank_var"].set("")
                entry["prod_var"].set("IPA")
                entry["date_var"].set("")
                entry["time_var"].set("")
                entry["mod_time_var"].set("")

    def clear_single_row(self, r_idx):
        """清空單一列的資料 (單個清)"""
        if 0 <= r_idx < len(self.entries):
            entry = self.entries[r_idx]
            entry["batch_var"].set("")
            entry["loc_var"].set("")
            entry["long_code_var"].set("")
            entry["tank_var"].set("")
            entry["date_var"].set("")
            entry["time_var"].set("")
            entry["mod_time_var"].set("")

    def on_batch_change(self, batch_var, tank_var):
        batch = batch_var.get().upper().strip()
        tank = get_tank_from_batch(batch)
        tank_var.set(tank)

    def on_loc_change(self, loc_var, long_code_var):
        loc = loc_var.get().strip().upper()
        if not loc:
            long_code_var.set("")
            return
            
        if loc in self.mapping_dict:
            long_code_var.set(self.mapping_dict[loc])
        else:
            long_code_var.set("❌ 未知代號")

    def parse_pasted_row_items(self, parts):
        """
        智慧解析貼上行中的元素，自動辨識：出貨日期、批號、槽號、地點、時間。
        支援 Image 1 範例（到貨 + 批號 + 槽號 + 地點）及各式自訂順序！
        """
        res = {"batch": "", "loc": "", "date": "", "time": "", "mod_time": ""}
        clean_parts = [p.strip() for p in parts if p.strip()]
        if not clean_parts:
            return res

        unassigned = []
        for p in clean_parts:
            p_upper = p.upper()
            
            # 1. 日期 (包含 /, -, ., 或月/日，長度 <= 12)
            if not res["date"] and (any(char in p for char in ("/", "-", ".")) or "月" in p or "日" in p):
                if len(p) <= 12 and not p.isalnum():
                    res["date"] = p
                    continue

            # 2. 批號 (8~12 碼英數混合，例如 26817E3051)
            if not res["batch"] and (8 <= len(p_upper) <= 12 and any(c.isdigit() for c in p_upper) and any(c.isalpha() for c in p_upper)):
                res["batch"] = p_upper
                continue

            # 3. 地點 (在 mapping_dict 內或含 18P, 15P, P, 廠等)
            if not res["loc"] and (p_upper in self.mapping_dict or "18P" in p_upper or "15P" in p_upper or "P" in p_upper or "廠" in p):
                res["loc"] = p_upper
                continue

            # 4. 槽號 (3~5 碼以 E/T/P 開頭，例如 E305) -> 可略過，因為寫入批號時微服務會自動算槽號！
            if len(p_upper) <= 5 and p_upper.startswith(("E", "T", "P")):
                continue

            # 其它填入未指派
            unassigned.append(p)

        # 針對缺額自動後補填入
        for item in unassigned:
            if not res["batch"] and len(item) >= 6:
                res["batch"] = item.upper()
            elif not res["loc"] and len(item) <= 10:
                res["loc"] = item.upper()
            elif not res["date"] and ("/" in item or "-" in item or "." in item):
                res["date"] = item
            elif not res["time"]:
                res["time"] = item
            elif not res["mod_time"]:
                res["mod_time"] = item

        return res

    def on_paste(self, event, start_row_idx, widget=None):
        try:
            clipboard = self.clipboard_get()
            lines = clipboard.split('\n')
            
            valid_lines = [l for l in lines if l.strip()]
            if not valid_lines:
                return "break"
                
            needed_rows = start_row_idx + len(valid_lines)
            if needed_rows > len(self.entries):
                self.add_input_rows(needed_rows - len(self.entries))
            
            curr_row = start_row_idx
            for line in lines:
                if not line.strip(): continue
                parts = line.split('\t')
                if len(parts) == 1:
                    parts = line.split()
                    
                if len(parts) >= 2:
                    # 使用智慧元素剖析器！
                    parsed = self.parse_pasted_row_items(parts)
                    
                    self.entries[curr_row]["chk_var"].set(True)
                    if parsed["batch"]:
                        self.entries[curr_row]["batch_var"].set(parsed["batch"])
                    if parsed["loc"]:
                        self.entries[curr_row]["loc_var"].set(parsed["loc"])
                    if parsed["date"]:
                        self.entries[curr_row]["date_var"].set(parsed["date"])
                    if parsed["time"]:
                        self.entries[curr_row]["time_var"].set(parsed["time"])
                    if parsed["mod_time"]:
                        self.entries[curr_row]["mod_time_var"].set(parsed["mod_time"])
                    
                    curr_row += 1
                elif len(parts) == 1:
                    val = parts[0].strip()
                    # 單一欄位貼上：依當前焦點與內容自動指派
                    if val:
                        parsed = self.parse_pasted_row_items([val])
                        if parsed["batch"]:
                            self.entries[curr_row]["batch_var"].set(parsed["batch"])
                        elif parsed["loc"]:
                            self.entries[curr_row]["loc_var"].set(parsed["loc"])
                        elif parsed["date"]:
                            self.entries[curr_row]["date_var"].set(parsed["date"])
                        else:
                            self.entries[curr_row]["batch_var"].set(val)
                    curr_row += 1
                    
            return "break"
        except Exception as e:
            messagebox.showerror("貼上失敗", f"解析貼上內容時發生錯誤:\n{e}")
            return "break"


    def upload_coa(self):
        if not os.path.exists(r'C:\Program Files\Tesseract-OCR	esseract.exe') and not os.path.exists(r'C:\Program Files (x86)\Tesseract-OCR	esseract.exe'):
            ans = messagebox.askyesno(
                "缺少 OCR 引擎", 
                "系統偵測到您尚未安裝『Tesseract OCR 引擎』，無法使用截圖辨識功能！\n\n"
                "是否要立即開啟官方下載網頁？\n\n"
                "(請下載最新的 64 bit 安裝檔，並【一直按下一步】安裝在預設路徑即可)"
            )
            if ans:
                webbrowser.open("https://github.com/UB-Mannheim/tesseract/wiki")
            return
            
        filepaths = filedialog.askopenfilenames(title="選擇 COA 截圖", filetypes=[("Image files", "*.jpg *.jpeg *.png")])
        if filepaths:
            self.coa_paths.extend(filepaths)
            messagebox.showinfo("上傳成功", f"成功上傳 {len(filepaths)} 張截圖！\n目前共 {len(self.coa_paths)} 張待處理。")

    def import_from_excel(self):
        filepaths = filedialog.askopenfilenames(
            title="選擇要匯入的 Excel 或 CSV 檔案 (可多選)",
            filetypes=[("Excel & CSV files", "*.xlsx *.xls *.csv")]
        )
        if not filepaths:
            return
            
        total_imported = 0
        for filepath in filepaths:
            try:
                rows = []
                if filepath.lower().endswith('.csv'):
                    try:
                        with open(filepath, 'r', encoding='utf-8-sig') as f:
                            reader = csv.reader(f)
                            rows = list(reader)
                    except UnicodeDecodeError:
                        with open(filepath, 'r', encoding='cp950') as f:
                            reader = csv.reader(f)
                            rows = list(reader)
                else:
                    wb = openpyxl.load_workbook(filepath, data_only=True)
                    if '空白班表' in wb.sheetnames:
                        ws = wb['空白班表']
                    else:
                        ws = wb.active
                    for r in ws.iter_rows(values_only=True):
                        rows.append(r)
                    wb.close()
                
                records = []
                
                # Check for vertical CSV format (e.g. RawLotId in first column)
                is_vertical_csv = False
                for r_idx, row_vals in enumerate(rows[:20]):
                    if len(row_vals) >= 2:
                        key = str(row_vals[0]).strip()
                        if key == "RawLotId":
                            is_vertical_csv = True
                            break
                            
                if is_vertical_csv:
                    batch_val = ""
                    date_val = ""
                    loc_val = ""
                    
                    for r_idx, row_vals in enumerate(rows[:30]):
                        if len(row_vals) >= 2:
                            key = str(row_vals[0]).strip()
                            val = str(row_vals[1]).strip()
                            if key == "RawLotId":
                                batch_val = val
                            elif key == "DeliverDate":
                                date_val = val
                                
                    fname = os.path.basename(filepath)
                    m = re.search(r'\s([A-Za-z0-9]+)_\d+\.csv$', fname, re.IGNORECASE)
                    if m:
                        loc_val = m.group(1)
                    else:
                        parts = fname.split('_')
                        if len(parts) >= 2:
                            last_part = parts[-2]
                            sub_parts = last_part.split(' ')
                            loc_val = sub_parts[-1]
                            
                    if batch_val:
                        d_pure, t_pure = split_date_and_time(date_val) if date_val else ("", "")
                        records.append({
                            "batch": batch_val,
                            "loc": loc_val,
                            "date": d_pure,
                            "time": "",
                            "mod_time": ""
                        })
                else:
                    # Original horizontal format parsing
                    batch_col = -1
                    loc_col = -1
                    date_col = -1
                    time_col = -1
                    mod_time_col = -1
                    cust_col = -1
                    prod_col = -1
                    start_row = -1
                    
                    for r_idx, row_vals in enumerate(rows[:10]):
                        for c_idx, val in enumerate(row_vals):
                            if val is None: continue
                            v_str = str(val).strip()
                            if "批號" in v_str:
                                batch_col = c_idx
                            if "對象" in v_str or "客戶" in v_str or "廠商" in v_str:
                                cust_col = c_idx
                            if "品名" in v_str or "產品" in v_str:
                                prod_col = c_idx
                            if "地點" in v_str or "交貨地點" in v_str or "指送地" in v_str or "指送" in v_str:
                                loc_col = c_idx
                            if "日期" in v_str or "到貨日期" in v_str or "出車日期" in v_str:
                                date_col = c_idx
                            if "修正" in v_str and "時間" in v_str:
                                mod_time_col = c_idx
                            elif ("時間" in v_str or "到貨時間" in v_str) and mod_time_col == -1:
                                time_col = c_idx
                                
                        if batch_col != -1 and loc_col != -1:
                            start_row = r_idx + 1
                            break
                            
                    if batch_col == -1 or loc_col == -1:
                        batch_col = 1
                        loc_col = 3
                        start_row = 1
                    
                    for r_idx in range(start_row, len(rows)):
                        row = rows[r_idx]
                        if len(row) <= max(batch_col, loc_col): continue
                        
                        b_val = row[batch_col] if batch_col < len(row) else None
                        c_val = row[cust_col] if cust_col != -1 and cust_col < len(row) else None
                        l_val = row[loc_col] if loc_col < len(row) else None
                        d_val = row[date_col] if date_col != -1 and date_col < len(row) else None
                        t_val = row[time_col] if time_col != -1 and time_col < len(row) else None
                        mt_val = row[mod_time_col] if mod_time_col != -1 and mod_time_col < len(row) else None
                        
                        # 嚴格過濾：僅保留台積電 (TSMC) 訂單
                        is_tsmc = False
                        if c_val and "台積" in str(c_val):
                            is_tsmc = True
                        elif l_val and ("台積" in str(l_val) or re.search(r'^(?:12P|14P|15P|18P|AP|F\d+)', str(l_val).strip())):
                            is_tsmc = True
                        elif not c_val and l_val:
                            norm_l = normalize_location(l_val)
                            if norm_l in self.mapping_dict or "台積" in str(l_val) or re.search(r'^(?:12P|14P|15P|18P|AP|F\d+)', str(l_val).strip()):
                                is_tsmc = True
                                
                        if not is_tsmc:
                            continue
                        
                        p_val = row[prod_col] if prod_col != -1 and prod_col < len(row) else None
                        p_str = str(p_val).strip() if p_val else "IPA"
                        
                        if b_val and str(b_val).strip() and not str(b_val).startswith('批號'):
                            d_pure, t_pure = split_date_and_time(d_val)
                            t_final = str(t_val).strip() if t_val else t_pure
                            
                            records.append({
                                "batch": str(b_val).strip(),
                                "product": p_str,
                                "weight": "4300",
                                "loc": normalize_location(l_val) if l_val else "",
                                "date": d_pure,
                                "time": t_final,
                                "mod_time": str(mt_val).strip() if mt_val else ""
                            })
                        
                if not records:
                    messagebox.showinfo("提示", f"在檔案 {os.path.basename(filepath)} 中找不到有效資料！")
                    continue

                dialog = ImportRangeDialog(self, records, os.path.basename(filepath))
                self.wait_window(dialog)

                if not dialog.selected_records:
                    continue

                target_records = dialog.selected_records

                start_idx = 0
                for idx, entry in enumerate(self.entries):
                    if not entry["batch_var"].get().strip():
                        start_idx = idx
                        break
                else:
                    start_idx = len(self.entries)
                    
                needed_rows = start_idx + len(target_records)
                if needed_rows > len(self.entries):
                    self.add_input_rows(needed_rows - len(self.entries))
                    
                for i, rec in enumerate(target_records):
                    row_e = self.entries[start_idx + i]
                    row_e["chk_var"].set(True)
                    row_e["batch_var"].set(rec["batch"])
                    if rec.get("product"): row_e["prod_var"].set(rec["product"])
                    row_e["loc_var"].set(rec["loc"])
                    if rec["date"]: row_e["date_var"].set(rec["date"])
                    if rec["time"]: row_e["time_var"].set(rec["time"])
                    if rec["mod_time"]: row_e["mod_time_var"].set(rec["mod_time"])
                    
                total_imported += len(target_records)
                
            except Exception as e:
                messagebox.showerror("匯入失敗", f"解析檔案 {os.path.basename(filepath)} 時發生錯誤:\n{e}")
                
        if total_imported > 0:
            messagebox.showinfo("匯入完成", f"已成功從所選檔案中匯入共 {total_imported} 筆資料！")

    def load_existing_transport_notice(self):
        """
        讓人員手動選擇要修訂的『運輸通知表.xlsx』或資料夾，預設開啟當天的輸出資料夾。
        還原原本的所有位置 (第1~N列)，讓人員直接於指定列輸入『修正到廠時間』！
        """

        # 預設開啟今天的輸出資料夾
        today_dir_name = f"勝一三合一單輸出_{datetime.now().strftime('%Y%m%d')}"
        today_dir = os.path.join(self.base_dir, today_dir_name)
        initial_dir = today_dir if os.path.exists(today_dir) else self.base_dir

        filepath = filedialog.askopenfilename(
            initialdir=initial_dir,
            title="選擇要修訂的運輸通知表 Excel 檔案",
            filetypes=[("Excel 運輸通知表", "*.xlsx *.xls")]
        )
        if not filepath:
            return

        records = []
        target_dir = os.path.dirname(filepath)
        folder_name = os.path.basename(target_dir)

        # 1. 優先嘗試讀取該輸出資料夾中的 session.json 快取（包含完整 10 碼批號）
        session_file = os.path.join(target_dir, "session.json")
        if not os.path.exists(session_file):
            # 備用讀取根目錄 session 快取
            session_file = os.path.join(self.base_dir, "last_generated_session.json")

        if os.path.exists(session_file):
            try:
                with open(session_file, "r", encoding="utf-8") as f:
                    records = json.load(f)
            except Exception:
                pass

        # 2. 若快取不存在，則直接分析選取的 Excel 運輸通知表卡片
        if not records:
            try:
                wb = openpyxl.load_workbook(filepath, data_only=True)
                ws = wb.active

                curr_r = 1
                while curr_r <= ws.max_row:
                    d_val = ws.cell(row=curr_r, column=6).value
                    t_val = ws.cell(row=curr_r+2, column=6).value
                    mt_val = ws.cell(row=curr_r+3, column=6).value or ws.cell(row=curr_r+3, column=13).value
                    l_val = ws.cell(row=curr_r+1, column=4).value
                    tank_val = ws.cell(row=curr_r+1, column=5).value

                    l_str = str(l_val).replace("台積", "").strip() if l_val else ""
                    d_pure, t_pure = split_date_and_time(d_val)
                    t_final = str(t_val).strip() if t_val else t_pure

                    if d_pure or l_str or tank_val:
                        records.append({
                            "batch": "",
                            "tank": str(tank_val).strip() if tank_val else "",
                            "loc": l_str,
                            "date": d_pure,
                            "time": t_final,
                            "mod_time": str(mt_val).strip() if mt_val else ""
                        })
                    curr_r += 7
                wb.close()
            except Exception as e:
                messagebox.showerror("載入失敗", f"讀取既有運輸通知表失敗:\n{e}")
                return

        if not records:
            messagebox.showwarning("提示", "選取的檔案或資料夾中無任何可還原的排程紀錄！")
            return

        # 清空目前表格
        for entry in self.entries:
            entry["batch_var"].set("")
            entry["loc_var"].set("")
            entry["long_code_var"].set("")
            entry["tank_var"].set("")
            entry["date_var"].set("")
            entry["time_var"].set("")
            entry["mod_time_var"].set("")

        needed_rows = len(records)
        if needed_rows > len(self.entries):
            self.add_input_rows(needed_rows - len(self.entries))

        for idx, rec in enumerate(records):
            row_e = self.entries[idx]
            if rec.get("batch"): row_e["batch_var"].set(rec["batch"])
            if rec.get("loc"): row_e["loc_var"].set(rec["loc"])
            if rec.get("date"): row_e["date_var"].set(rec["date"])
            if rec.get("time"): row_e["time_var"].set(rec["time"])
            if rec.get("mod_time"): row_e["mod_time_var"].set(rec["mod_time"])

        messagebox.showinfo(
            "還原成功", 
            f"已成功從『{folder_name}』載入 {len(records)} 筆既有排程紀錄！\n\n"
            f"所有排程已按原始位置 (第 1~{len(records)} 列) 精準對齊。\n"
            f"請直接在需要修正的排程列填寫【修正到廠時間】即可！"
        )

    def generate_files(self):
        if not os.path.exists(self.template_path):
            messagebox.showerror("錯誤", f"找不到範本檔案:\n{self.template_path}")
            return
        if not os.path.exists(self.mapping_path):
            messagebox.showerror("錯誤", f"找不到對照表檔案:\n{self.mapping_path}")
            return
            
        do_3in1 = self.gen_3in1_var.get()
        do_transport = self.gen_transport_var.get()
        
        if not do_3in1 and not do_transport:
            messagebox.showwarning("提示", "請至少勾選一種報表類型（三合一單 或 運輸通知表）！")
            return

        valid_data = []
        for idx, row in enumerate(self.entries):
            batch = row["batch_var"].get().strip().upper()
            loc = row["loc_var"].get().strip().upper()
            tank = row["tank_var"].get().strip()
            date_str = row["date_var"].get().strip()
            time_str = row["time_var"].get().strip()
            mod_time_str = row["mod_time_var"].get().strip()
            
            if not batch and not loc:
                continue
                
            # 自動勾選有輸入資料的列
            if not row["chk_var"].get():
                row["chk_var"].set(True)
                
            if not batch or not loc:
                messagebox.showerror("錯誤", f"第 {idx+1} 項資料不齊全！請確認已填寫批號與地點。")
                return
            if len(batch) < 9 or len(batch) > 13:
                messagebox.showerror("錯誤", f"第 {idx+1} 項的批號長度錯誤！\n勝一批號通常為 10~11 碼，目前輸入: {batch} (長度 {len(batch)})")
                return
                
            prod_name = row["prod_var"].get().strip() or "IPA"
            weight_val = get_product_weight(prod_name)
            valid_data.append({
                "batch": batch,
                "tank": tank,
                "product": prod_name,
                "weight": weight_val,
                "loc": loc,
                "date": date_str,
                "time": time_str,
                "mod_time": mod_time_str
            })
            
        if not valid_data:
            messagebox.showinfo("提示", "請至少勾選並輸入一筆有效資料！")
            return

        missing_locs = [data["loc"] for data in valid_data if data["loc"] not in self.mapping_dict]
        if missing_locs:
            missing_str = ", ".join(set(missing_locs))
            messagebox.showerror("錯誤", f"地點代號對照表中找不到以下地點：\n{missing_str}\n\n請先更新對照表後再試！")
            return

        output_dir = os.path.join(self.base_dir, f"勝一三合一單輸出_{datetime.now().strftime('%Y%m%d')}")
        os.makedirs(output_dir, exist_ok=True)
        
        success_3in1 = 0
        error_msgs = []
        mat_no = "L12C53161"

        if do_3in1:
            coa_crops = {}
            if os.path.exists(r'C:\Program Files\Tesseract-OCR\tesseract.exe'):
                pytesseract.pytesseract.tesseract_cmd = r'C:\Program Files\Tesseract-OCR\tesseract.exe'
            else:
                pytesseract.pytesseract.tesseract_cmd = r'C:\Program Files (x86)\Tesseract-OCR\tesseract.exe' 
            if self.coa_paths:
                for img_path in self.coa_paths:
                    try:
                        orig_img = PILImage.open(img_path)
                        img = orig_img.resize((orig_img.width * 2, orig_img.height * 2), PILImage.Resampling.LANCZOS)
                        d = pytesseract.image_to_data(img, output_type=Output.DICT)
                        
                        header_bottom = int(img.height * 0.4)
                        for i in range(len(d['text'])):
                            if 'Batch' in d['text'][i] or 'ID' in d['text'][i] or 'No' in d['text'][i]:
                                header_bottom = d['top'][i] + d['height'][i] + 12
                                break
                                
                        img_top = img.crop((0, 0, img.width, header_bottom))
                        
                        for i in range(len(d['text'])):
                            text = d['text'][i].strip()
                            digits = ''.join(c for c in text if c.isdigit())
                            if len(digits) >= 8:
                                batch_digits = digits
                                row_top = max(0, d['top'][i] - 14)
                                row_bottom = min(img.height, d['top'][i] + d['height'][i] + 16)
                                
                                img_row = img.crop((0, row_top, img.width, row_bottom))
                                new_img = PILImage.new('RGB', (img.width, img_top.height + img_row.height))
                                new_img.paste(img_top, (0, 0))
                                new_img.paste(img_row, (0, img_top.height))
                                
                                coa_crops[batch_digits] = new_img
                    except Exception as e:
                        print("OCR error:", e)

            for data in valid_data:
                batch_no = data["batch"]
                tank_no = data["tank"]
                loc = data["loc"]
                loc_code = self.mapping_dict[loc]
                
                try:
                    wb = openpyxl.load_workbook(self.template_path)
                    ws = wb.worksheets[0]
                    
                    tank_row = find_row_by_label(ws, ['槽號']) or 5
                    batch_row = find_row_by_label(ws, ['批號']) or 7
                    loc_row = find_row_by_label(ws, ['送達地點', '地點']) or 11
                    mat_row = find_row_by_label(ws, ['料號']) or 3
                    sup_row = find_row_by_label(ws, ['供應商']) or 9
                    
                    raw_mat = str(ws.cell(row=mat_row, column=3).value or "").strip()
                    if raw_mat.startswith("4"):
                        mat_no = raw_mat[1:]
                    elif raw_mat:
                        mat_no = raw_mat
                    
                    final_tank_no = "5" + tank_no
                    final_batch_no = "6" + batch_no
                    
                    ws.cell(row=tank_row, column=3).value = final_tank_no
                    ws.cell(row=batch_row, column=3).value = final_batch_no
                    ws.cell(row=loc_row, column=3).value = loc_code
                    
                    images_to_keep = []
                    for img in ws._images:
                        if img.height < 150 and img.width > 200:
                            images_to_keep.append(img)
                    ws._images = images_to_keep
                    
                    c3_val = ws.cell(row=mat_row, column=3).value or ""
                    c6_val = ws.cell(row=sup_row, column=3).value or ""
                    qr_str = f"||{c3_val}||{final_tank_no}||{final_batch_no}||{c6_val}||{loc_code}"
                    
                    qr = qrcode.QRCode(box_size=4, border=2)
                    qr.add_data(qr_str)
                    qr.make(fit=True)
                    raw_img = qr.make_image(fill_color="black", back_color="white").convert('RGB')
                    
                    offset_x = 35
                    offset_y = 15
                    new_width = raw_img.width + offset_x
                    new_height = raw_img.height + offset_y
                    img_qr = Image.new('RGBA', (new_width, new_height), (255,255,255,0))
                    img_qr.paste(raw_img, (offset_x, offset_y))
                    
                    img_byte_arr = BytesIO()
                    img_qr.save(img_byte_arr, format='PNG')
                    img_byte_arr.seek(0)
                    
                    new_qr = OpenpyxlImage(img_byte_arr)
                    new_qr.anchor = 'F2'
                    ws.add_image(new_qr)
                    
                    safe_loc = "".join(c for c in loc if c.isalnum() or c in (' ', '_', '-')).rstrip()
                    if not safe_loc:
                        safe_loc = "未命名地點"
                    
                    # 產生檔名規格：[出貨日期]. [地點]台積電槽車barcode三合一單.xlsx (例如: 2026.8.18. 18P3B台積電槽車barcode三合一單.xlsx)
                    date_raw = data.get("date", "").strip()
                    dt_file = None
                    if date_raw:
                        for fmt in ("%Y/%m/%d", "%Y-%m-%d", "%Y/%M/%d", "%Y.%m.%d"):
                            try:
                                dt_file = datetime.strptime(date_raw, fmt)
                                break
                            except ValueError:
                                pass
                    if not dt_file:
                        dt_file = datetime.now()
                        

                    # Insert COA Crop
                    found_coa = False
                    user_digits = ''.join(c for c in batch_no if c.isdigit())
                    for k_batch, crop_img in coa_crops.items():
                        if user_digits in k_batch or k_batch in user_digits:
                            img_byte_arr2 = BytesIO()
                            crop_img.save(img_byte_arr2, format='PNG')
                            img_byte_arr2.seek(0)
                            xl_img = OpenpyxlImage(img_byte_arr2)
                            xl_img.width = 867
                            xl_img.height = 450
                            xl_img.anchor = 'F5'
                            ws.add_image(xl_img)
                            found_coa = True
                            break
                    
                    if not found_coa and self.coa_paths:
                        error_msgs.append(f"⚠️ 警告: 批號 {batch_no} 未在截圖找到，已留白處理！")

                    date_prefix = f"{dt_file.year}.{dt_file.month}.{dt_file.day}. "
                    tank_suffix = f"_{tank_no}" if tank_no else ""
                    base_filename = f"{date_prefix}{safe_loc}{tank_suffix}_台積電槽車barcode三合一單.xlsx"
                    output_path = os.path.join(output_dir, base_filename)
                    counter = 1
                    while os.path.exists(output_path):
                        base_filename = f"{date_prefix}{safe_loc}_{counter}_台積電槽車barcode三合一單.xlsx"
                        output_path = os.path.join(output_dir, base_filename)
                        counter += 1
                        
                    output_filename = base_filename
                    wb.save(output_path)
                    wb.close()
                    success_3in1 += 1
                except Exception as e:
                    error_msgs.append(f"處理三合一單 {loc}_{batch_no} 失敗: {e}")

        success_transport = False
        if do_transport:
            try:
                transport_path = os.path.join(output_dir, "運輸通知表.xlsx")
                generate_transport_notice_file(transport_path, valid_data, mat_no=mat_no)
                success_transport = True
            except Exception as e:
                error_msgs.append(f"產生運輸通知表失敗: {e}")

        # 自動快取當前 Session 資料至輸出資料夾與根目錄，供往後一鍵精準還原修訂
        try:
            for target_path in (os.path.join(output_dir, "session.json"), os.path.join(self.base_dir, "last_generated_session.json")):
                with open(target_path, "w", encoding="utf-8") as f:
                    json.dump(valid_data, f, ensure_ascii=False, indent=2)
        except Exception:
            pass

        msg_parts = []
        if do_3in1:
            msg_parts.append(f"• 三合一單：成功產生 {success_3in1} 份")
        if do_transport:
            status_str = "成功" if success_transport else "失敗"
            msg_parts.append(f"• 運輸通知表：{status_str} (共 {len(valid_data)} 筆排程卡片)")
            
        msg = "\n".join(msg_parts) + f"\n\n檔案已儲存於資料夾：\n{output_dir}"
        
        if error_msgs:
            msg += "\n\n部分錯誤:\n" + "\n".join(error_msgs[:5])
            messagebox.showwarning("完成 (但有部分錯誤)", msg)
        else:
            messagebox.showinfo("成功", msg)
            
        os.startfile(output_dir)

if __name__ == "__main__":
    app = App()
    app.mainloop()

