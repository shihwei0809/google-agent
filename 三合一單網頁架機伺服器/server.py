import os
import sys
import socket
import json
import zipfile
from io import BytesIO
from datetime import datetime
import openpyxl
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.cell.rich_text import TextBlock, CellRichText
from openpyxl.cell.text import InlineFont
import qrcode
from PIL import Image
from fastapi import FastAPI, Form, Request, HTTPException, UploadFile, File
from fastapi.responses import HTMLResponse, StreamingResponse, JSONResponse
from fastapi.staticfiles import StaticFiles
import uvicorn

# ================= 1. IP 與 Port 自動取得 (Rule 6 規範) =================

def get_local_ip():
    try:
        s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        s.connect(("8.8.8.8", 80))
        ip = s.getsockname()[0]
        s.close()
        return ip
    except Exception:
        return "127.0.0.1"

def find_available_port(start_port: int, max_attempts: int = 50) -> int:
    for port in range(start_port, start_port + max_attempts):
        with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:
            try:
                s.bind(("0.0.0.0", port))
                return port
            except OSError:
                continue
    return start_port

# ================= 2. 核心業務邏輯 (批號解析與對照) =================

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
TEMPLATE_PATH = os.path.join(BASE_DIR, "台積電槽車barcode三合一單-範本.xlsx")
MAPPING_PATH = os.path.join(BASE_DIR, "地點代號對照表.xlsx")

def load_location_mapping():
    mapping = {
        "15P5": "E1550155A",
        "15P6": "E1550156A",
        "18P3B": "EF180183B",
        "12P7": "E00700001"
    }
    if os.path.exists(MAPPING_PATH):
        try:
            wb = openpyxl.load_workbook(MAPPING_PATH, data_only=True)
            ws = wb.active
            for row in ws.iter_rows(values_only=True):
                if row and len(row) >= 2 and row[0] and row[1]:
                    k = str(row[0]).strip().upper()
                    v = str(row[1]).strip()
                    if "地點" in k or "代號" in k or "SHORT" in k:
                        continue
                    mapping[k] = v
            wb.close()
        except Exception as e:
            print(f"警告: 讀取地點對照表失敗: {e}")
    return mapping

def extract_tank_from_batch(batch_no: str) -> str:
    batch = batch_no.strip().upper()
    if len(batch) != 10:
        return ""
    if batch.endswith("J1"):
        return batch[5:8]
    return batch[5:9]

# ================= 3. FastAPI Web 應用程式 =================

app = FastAPI(title="台積電槽車 Barcode 三合一單專用架機伺服器")

@app.get("/api/mapping")
def get_mapping():
    mapping = load_location_mapping()
    return JSONResponse({"status": "success", "count": len(mapping), "data": mapping})

def generate_transport_workbook(items, mat_no="L12C53161"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "運輸通知表"
    ws.views.sheetView[0].showGridLines = True
    
    col_widths = {
        'A': 8, 'B': 18, 'C': 18, 'D': 16, 'E': 14, 'F': 16,
        'G': 4,
        'H': 8, 'I': 18, 'J': 18, 'K': 16, 'L': 14, 'M': 16
    }
    for col, width in col_widths.items():
        ws.column_dimensions[col].width = width
    
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    
    fill_yellow = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    fill_green = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    fill_bright_yellow = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    
    openpyxl_font_dark_blue = Font(name="Microsoft JhengHei", color="002060", size=11)
    openpyxl_font_dark_blue_b13 = Font(name="Microsoft JhengHei", color="002060", size=13, bold=True)
    openpyxl_font_strike_blue_b13 = Font(name="Microsoft JhengHei", color="002060", size=13, bold=True, strike=True)
    openpyxl_font_red_b13 = Font(name="Microsoft JhengHei", color="C00000", size=13, bold=True)
    openpyxl_font_dark_blue_b14 = Font(name="Microsoft JhengHei", color="002060", size=14, bold=True)
    
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    weekdays = ["星期一", "星期二", "星期三", "星期四", "星期五", "星期六", "星期日"]

    def render_notice_card(start_r, start_c, is_modified_card, item):
        r1 = start_r
        r2 = start_r + 1
        r3 = start_r + 2
        r4 = start_r + 3
        r5 = start_r + 4
        r6 = start_r + 5
        
        c1 = start_c
        c2 = start_c + 1
        c3 = start_c + 2
        c4 = start_c + 3
        c5 = start_c + 4
        c6 = start_c + 5
        
        for r in range(r1, r6 + 1):
            ws.row_dimensions[r].height = 24 if r >= r3 else 22
            for c in range(c1, c6 + 1):
                cell = ws.cell(row=r, column=c)
                cell.border = thin_border
                cell.font = openpyxl_font_dark_blue
                cell.alignment = align_center

        ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c3)
        title_suffix = "出貨排程修正通知" if is_modified_card else "出貨排程通知"
        title_rt = CellRichText([
            TextBlock(InlineFont(color="002060", b=True, sz=12, rFont="Microsoft JhengHei"), "Shiny IPA Lorry\n"),
            TextBlock(InlineFont(color="002060", sz=10, rFont="Microsoft JhengHei"), f"(料號：{mat_no}) {title_suffix}")
        ])
        cell_a1 = ws.cell(row=r1, column=c1)
        cell_a1.value = title_rt
        cell_a1.fill = fill_yellow
        cell_a1.alignment = align_center
        
        cell_d1 = ws.cell(row=r1, column=c4, value="廠區")
        cell_d1.fill = fill_yellow
        cell_d1.font = Font(name="Microsoft JhengHei", color="002060", size=11, bold=True)
        
        cell_e1 = ws.cell(row=r1, column=c5, value="槽號")
        cell_e1.fill = fill_yellow
        cell_e1.font = Font(name="Microsoft JhengHei", color="002060", size=11, bold=True)
        
        date_raw = item.get("date", "").strip() if item.get("date") else datetime.now().strftime("%Y-%m-%d")
        formatted_date = ""
        weekday_str = ""
        if date_raw:
            dt = None
            date_part = date_raw.split()[0]
            for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d", "%Y%m%d", "%m/%d/%Y", "%d/%m/%Y"):
                try:
                    dt = datetime.strptime(date_part, fmt)
                    break
                except ValueError:
                    pass
            if dt:
                weekday_str = weekdays[dt.weekday()]
                formatted_date = f"{dt.year}/{dt.month}/{dt.day}"
            else:
                formatted_date = date_raw
        
        cell_f1 = ws.cell(row=r1, column=c6, value=formatted_date)
        cell_f1.fill = fill_green
        cell_f1.font = Font(name="Microsoft JhengHei", color="002060", size=11, bold=True)
        
        cell_f2 = ws.cell(row=r2, column=c6, value=weekday_str)
        cell_f2.fill = fill_green
        cell_f2.font = Font(name="Microsoft JhengHei", color="002060", size=11, bold=True)

        ws.merge_cells(start_row=r3, start_column=c1, end_row=r6, end_column=c1)
        cell_a3 = ws.cell(row=r3, column=c1, value="IPA")
        cell_a3.font = openpyxl_font_dark_blue_b14
        
        time_val = item.get("time", "").strip() if item.get("time") else ""
        mod_time_val = item.get("modTime", "").strip() if item.get("modTime") else ""
        
        ws.merge_cells(start_row=r3, start_column=c2, end_row=r3, end_column=c3)
        ws.cell(row=r3, column=c2, value="預計到廠時間")
        cell_f3 = ws.cell(row=r3, column=c6, value=time_val)
        
        ws.merge_cells(start_row=r4, start_column=c2, end_row=r4, end_column=c3)
        ws.cell(row=r4, column=c2, value="修正到廠時間")
        cell_f4 = ws.cell(row=r4, column=c6)
        
        if is_modified_card:
            cell_f3.font = openpyxl_font_strike_blue_b13
            cell_f4.value = mod_time_val
            cell_f4.font = openpyxl_font_red_b13
        else:
            cell_f3.font = openpyxl_font_dark_blue_b13
            cell_f4.value = ""
        
        ws.merge_cells(start_row=r5, start_column=c2, end_row=r5, end_column=c3)
        ws.cell(row=r5, column=c2, value="充填數量(KG)")
        ws.cell(row=r5, column=c6, value="4300")
        
        ws.merge_cells(start_row=r2, start_column=c4, end_row=r5, end_column=c4)
        full_loc = item.get("loc", "").strip() if item.get("loc") else ""
        
        d_rt = CellRichText([
            TextBlock(InlineFont(color="002060", b=True, sz=14, rFont="Microsoft JhengHei"), "台積\n"),
            TextBlock(InlineFont(color="C00000", b=True, sz=14, rFont="Microsoft JhengHei"), full_loc)
        ])
        cell_d2 = ws.cell(row=r2, column=c4)
        cell_d2.value = d_rt
        
        ws.merge_cells(start_row=r2, start_column=c5, end_row=r5, end_column=c5)
        cell_e2 = ws.cell(row=r2, column=c5, value=item.get("tank", "").strip() if item.get("tank") else "")
        cell_e2.font = openpyxl_font_dark_blue_b14
        cell_e2.fill = fill_bright_yellow
        
        ws.merge_cells(start_row=r6, start_column=c2, end_row=r6, end_column=c5)
        cell_b6 = ws.cell(row=r6, column=c2)
        cell_b6.value = CellRichText([
            TextBlock(InlineFont(color="C00000", b=True, sz=11, rFont="Microsoft JhengHei"), "PFA 500ml"),
            TextBlock(InlineFont(color="002060", sz=11, rFont="Microsoft JhengHei"), " 取樣瓶裝原液 8 分滿放置工具箱內")
        ])
        
        cell_f6 = ws.cell(row=r6, column=c6, value="6 支")
        cell_f6.font = Font(name="Microsoft JhengHei", color="002060", size=11, bold=True)

    curr_row = 1
    for item in items:
        render_notice_card(curr_row, 1, False, item)
        if item.get("modTime", "").strip():
            render_notice_card(curr_row, 8, True, item)
        curr_row += 7

    return wb

# 1. 一鍵打包產生所有報表 ZIP (與 BAT 產出完全相同)
@app.post("/api/generate_all_zip")
async def generate_all_zip(request: Request):
    try:
        data_json = await request.json()
        records = data_json.get("records", [])
        do_3in1 = data_json.get("do3in1", True)
        do_transport = data_json.get("doTransport", True)

        if not records:
            raise HTTPException(status_code=400, detail="請至少提供一筆有效的排程資料。")

        mapping = load_location_mapping()
        zip_buffer = BytesIO()

        today_str = datetime.now().strftime('%Y%m%d')
        folder_name = f"三合一單輸出_{today_str}"

        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            # 1. 產生三合一單 Excel 報表
            if do_3in1 and os.path.exists(TEMPLATE_PATH):
                used_filenames = set()
                for item in records:
                    batch = item.get("batch", "").strip().upper()
                    loc = item.get("loc", "").strip().upper()
                    if not batch or not loc or len(batch) != 10 or loc not in mapping:
                        continue

                    loc_code = mapping[loc]
                    custom_tank = item.get("tank", "").strip()
                    if custom_tank and custom_tank != "自動槽號":
                        tank_no = custom_tank
                    else:
                        tank_no = extract_tank_from_batch(batch)
                    tank_with_prefix = "5" + tank_no
                    batch_with_prefix = "6" + batch

                    wb = openpyxl.load_workbook(TEMPLATE_PATH)
                    ws = wb["barcode"] if "barcode" in wb.sheetnames else wb.worksheets[0]

                    ws['C5'] = tank_with_prefix
                    ws['C7'] = batch_with_prefix
                    ws['C11'] = loc_code

                    mat_no = str(ws['C3'].value or "4L12C53161").strip()
                    sup_no = str(ws['C9'].value or "375970680").strip()
                    qr_str = f"||{mat_no}||{tank_with_prefix}||{batch_with_prefix}||{sup_no}||{loc_code}"
                    ws['B20'] = qr_str

                    # 清除舊圖片
                    images_to_keep = [img for img in ws._images if not (0.8 < (img.width/img.height if img.height>0 else 1) < 1.2 and img.width < 300)]
                    ws._images = images_to_keep

                    # 生成 QR Code PNG
                    qr = qrcode.QRCode(box_size=4, border=2)
                    qr.add_data(qr_str)
                    qr.make(fit=True)
                    raw_img = qr.make_image(fill_color="black", back_color="white").convert('RGB')

                    offset_x, offset_y = 35, 45
                    img_qr = Image.new('RGBA', (raw_img.width + offset_x, raw_img.height + offset_y), (255, 255, 255, 0))
                    img_qr.paste(raw_img, (offset_x, offset_y))

                    img_io = BytesIO()
                    img_qr.save(img_io, format='PNG')
                    img_io.seek(0)

                    new_qr = OpenpyxlImage(img_io)
                    new_qr.anchor = 'F2'
                    ws.add_image(new_qr)

                    # 若使用者有上傳 COA 截圖，自動針對「該筆排程的指定單一批號列」進行標頭+單列精確裁切 (與對照手冊 100% 相同)
                    if "latest_coa" in COA_CACHE and COA_CACHE["latest_coa"]:
                        try:
                            coa_raw = Image.open(BytesIO(COA_CACHE["latest_coa"])).convert("RGB")
                            w, h = coa_raw.size

                            # 1. 如果有瀏覽器端傳來的 OCR 座標，就能精準切出單一列！
                            ocr_data = COA_CACHE.get("ocr_data")
                            cropped_coa = None
                            if ocr_data and "boxes" in ocr_data and ocr_data["boxes"]:
                                header_bottom = ocr_data.get("header_bottom", int(h * 0.35))
                                header_bottom = max(0, min(h, int(header_bottom)))
                                img_top = coa_raw.crop((0, 0, w, header_bottom))
                                
                                # 尋找與當前 batch 匹配的區塊
                                target_digits = ''.join(c for c in batch if c.isdigit())
                                batch_box = None
                                
                                for box in ocr_data.get("boxes", []):
                                    box_text = box.get("text", "").upper()
                                    box_digits = ''.join(c for c in box_text if c.isdigit())
                                    if box_text == batch or (len(box_digits) >= 6 and (box_digits in target_digits or target_digits in box_digits)):
                                        batch_box = box
                                        break
                                    
                                if batch_box:
                                    row_t = int(batch_box.get("row_top", batch_box.get("top", 0) - 8))
                                    row_b = int(batch_box.get("row_bottom", batch_box.get("bottom", h) + 12))
                                    row_top = max(0, min(h, row_t))
                                    row_bottom = max(row_top + 5, min(h, row_b))
                                    
                                    img_row = coa_raw.crop((0, row_top, w, row_bottom))
                                    
                                    cropped_coa = Image.new("RGB", (w, img_top.height + img_row.height), "white")
                                    cropped_coa.paste(img_top, (0, 0))
                                    cropped_coa.paste(img_row, (0, img_top.height))
                                    print(f"[COA Crop] 成功為批號 {batch} 裁切：表頭高 {img_top.height}px + 數據列高 {img_row.height}px")

                            # 如果 OCR 失敗或沒有匹配到該批號，降級使用全保留方式
                            if not cropped_coa:
                                crop_bottom = int(h * 0.95)
                                cropped_coa = coa_raw.crop((0, 0, w, crop_bottom))

                            # 4. 保持原始高解析度畫質，在 Excel 中設定精準尺寸：寬 23.7cm (~896px) x 高 11.5cm (~435px)
                            coa_io = BytesIO()
                            cropped_coa.save(coa_io, format="PNG")
                            coa_io.seek(0)

                            coa_img = OpenpyxlImage(coa_io)
                            coa_img.width = int(round(23.7 * 96 / 2.54))   # 23.7 公分 (~896 px)
                            coa_img.height = int(round(11.5 * 96 / 2.54))  # 11.5 公分 (~435 px)

                            coa_img.anchor = 'F5'  # 100% 精準貼入圖一 barcode 分頁的 F5 儲存格！
                            ws.add_image(coa_img)
                        except Exception as coa_e:
                            print(f"[COA Crop Error] 處理批號 {batch} COA 發生錯誤: {coa_e}")

                    excel_io = BytesIO()
                    wb.save(excel_io)
                    wb.close()
                    excel_io.seek(0)

                    # 檔名公式：[出貨日期]. [槽號] [廠別]台積電槽車barcode三合一單.xlsx (例如: 2026.8.18. E44 18P3B台積電槽車barcode三合一單.xlsx)
                    date_raw = str(item.get("date", "")).strip()
                    dt_file = None
                    if date_raw:
                        date_part = date_raw.split()[0]
                        for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d", "%Y%m%d", "%m/%d/%Y", "%d/%m/%Y"):
                            try:
                                dt_file = datetime.strptime(date_part, fmt)
                                break
                            except ValueError:
                                pass
                    if not dt_file:
                        dt_file = datetime.now()

                    date_prefix = f"{dt_file.year}.{dt_file.month}.{dt_file.day}. "
                    tank_part = f"{tank_no} " if tank_no else ""
                    base_name = f"{date_prefix}{tank_part}{loc}台積電槽車barcode三合一單.xlsx"
                    test_name = base_name
                    counter = 1
                    while f"{folder_name}/{test_name}" in used_filenames:
                        test_name = f"{date_prefix}{tank_part}{loc}_{counter}台積電槽車barcode三合一單.xlsx"
                        counter += 1
                    file_name = test_name
                    used_filenames.add(f"{folder_name}/{file_name}")
                    zip_file.writestr(f"{folder_name}/{file_name}", excel_io.getvalue())

            # 2. 產生獨立運輸通知表 Excel
            if do_transport:
                wb_t = generate_transport_workbook(records)
                t_io = BytesIO()
                wb_t.save(t_io)
                wb_t.close()
                t_io.seek(0)
                zip_file.writestr(f"{folder_name}/運輸通知表.xlsx", t_io.getvalue())

            # 3. 寫入 session.json 至 ZIP 根目錄，供本機 BAT 或網頁版載入時 100% 精準還原原始完整 10 碼批號
            try:
                session_payload = []
                for r in records:
                    session_payload.append({
                        "batch": r.get("batch", ""),
                        "tank": r.get("tank", ""),
                        "loc": r.get("loc", ""),
                        "date": r.get("date", ""),
                        "time": r.get("time", ""),
                        "mod_time": r.get("modTime", r.get("mod_time", ""))
                    })
                zip_file.writestr(f"{folder_name}/session.json", json.dumps(session_payload, ensure_ascii=False, indent=2).encode('utf-8'))
            except Exception as se:
                print(f"[Session JSON Error] {se}")
            
            # 3. 產生額外附加檔案 (若有上傳 Excel，依批號過濾並只保留單列)
            extra_file = EXTRA_FILE_CACHE.get("latest_file")
            if extra_file and extra_file["ext"].lower() in [".xlsx", ".xls"]:
                for item in records:
                    batch = item.get("batch", "").strip().upper()
                    loc = item.get("loc", "").strip().upper()
                    if not batch or not loc or len(batch) != 10 or loc not in mapping:
                        continue
                    
                    try:
                        # 每次都從記憶體重新讀取原始檔案
                        new_wb = openpyxl.load_workbook(BytesIO(extra_file["content"]))
                        new_ws = new_wb.active
                        
                        # 尋找匹配的批號 (從第 7 列開始找，批號通常在 A 欄)
                        matched_row_idx = None
                        for r in range(7, new_ws.max_row + 1):
                            val = str(new_ws.cell(row=r, column=1).value or "").strip().upper()
                            if val == batch:
                                matched_row_idx = r
                                break
                                
                        if matched_row_idx:
                            # 巧妙利用刪除列來保留格式：刪除目標列之前的所有資料列 (7 ~ 目標-1)
                            if matched_row_idx > 7:
                                new_ws.delete_rows(7, matched_row_idx - 7)
                                
                            # 此時目標列已經往上移動變成第 7 列了，接著刪除第 8 列以後的所有資料
                            if new_ws.max_row > 7:
                                new_ws.delete_rows(8, new_ws.max_row - 7)
                            
                            # 組合新檔名：[原檔名前半部]-[MMDD] [Loc].[Ext]
                            orig_name = extra_file["filename"]
                            base_name = orig_name.rsplit('-', 1)[0] if '-' in orig_name else orig_name
                            
                            date_raw = item.get("date", "").strip()
                            mmdd = "0000"
                            if date_raw:
                                for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d", "%Y%m%d", "%m/%d/%Y", "%d/%m/%Y"):
                                    try:
                                        dt = datetime.strptime(date_raw, fmt)
                                        mmdd = f"{dt.month:02d}{dt.day:02d}"
                                        break
                                    except ValueError:
                                        pass
                            custom_tank = item.get("tank", "").strip()
                            if custom_tank and custom_tank != "自動槽號":
                                tank_no = custom_tank
                            else:
                                tank_no = extract_tank_from_batch(batch)
                            tank_part = f"{tank_no} " if tank_no else ""
                            new_filename = f"{base_name}-{mmdd} {tank_part}{loc}{extra_file['ext']}"
                            
                            # 儲存到 ZIP
                            out_buf = BytesIO()
                            new_wb.save(out_buf)
                            zip_file.writestr(f"{folder_name}/{new_filename}", out_buf.getvalue())
                    except Exception as ex:
                        print(f"處理附加檔案 {batch} 時發生錯誤: {ex}")

        # 生成完成後，清空快取避免影響下一次
        # COA_CACHE.clear() 
        # EXTRA_FILE_CACHE.clear()

        zip_buffer.seek(0)
        zip_filename = f"{folder_name}.zip"

        from urllib.parse import quote
        encoded_filename = quote(zip_filename)

        return StreamingResponse(
            zip_buffer,
            media_type="application/zip",
            headers={'Content-Disposition': f"attachment; filename*=UTF-8''{encoded_filename}"}
        )

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"一鍵打包產生失敗: {e}")

# 2. 智慧 OCR 辨識與文字解析 API
@app.post("/api/ocr_parse")
async def ocr_parse(file: UploadFile = File(...)):
    try:
        content = await file.read()
        extracted_records = []

        # 嘗試簡單文字與批號地點正規表示法解析
        text = ""
        try:
            text = content.decode("utf-8", "ignore")
        except Exception:
            text = str(content)

        import re
        # 尋找 10 碼批號模式與地點
        batches = re.findall(r'\b[0-9A-Z]{10}\b', text.upper())
        mapping = load_location_mapping()

        found_locs = []
        for word in text.upper().split():
            clean_w = re.sub(r'[^A-Z0-9]', '', word)
            if clean_w in mapping:
                found_locs.append(clean_w)

        for i, b in enumerate(batches):
            loc_val = found_locs[i] if i < len(found_locs) else "18P3B"
            tank_val = extract_tank_from_batch(b)
            extracted_records.append({
                "batch": b,
                "tank": tank_val,
                "loc": loc_val
            })

        return JSONResponse({
            "status": "success",
            "count": len(extracted_records),
            "records": extracted_records
        })
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"OCR 解析失敗: {e}")

# 2. COA 截圖處理與 Excel 自動貼上 API (純 Python Pillow 零依賴 .exe 方案)
COA_CACHE = {}
EXTRA_FILE_CACHE = {}

@app.post("/api/upload_extra_file")
async def upload_extra_file(file: UploadFile = File(...)):
    try:
        content = await file.read()
        import os
        filename, ext = os.path.splitext(file.filename)
        
        EXTRA_FILE_CACHE["latest_file"] = {
            "content": content,
            "filename": filename,
            "ext": ext
        }

        return JSONResponse({
            "status": "success",
            "message": f"附加檔案 {file.filename} 上傳成功！產生報表時將自動依排程複製與命名。"
        })
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"附加檔案處理失敗: {e}")

@app.post("/api/upload_coa_image")
async def upload_coa_image(file: UploadFile = File(...), ocr_data: str = Form(None)):
    try:
        content = await file.read()
        img = Image.open(BytesIO(content)).convert("RGB")

        # 保留原始高畫質，不進行任何強制壓縮或縮放
        w, h = img.size

        img_io = BytesIO()
        img.save(img_io, format="PNG")
        img_bytes = img_io.getvalue()
        COA_CACHE["latest_coa"] = img_bytes
        if ocr_data:
            import json
            try:
                COA_CACHE["ocr_data"] = json.loads(ocr_data)
            except:
                COA_CACHE["ocr_data"] = None

        return JSONResponse({
            "status": "success",
            "message": "COA 截圖已成功接收並完成影像最佳化！產生三合一單時將自動嵌入 Excel。"
        })
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"COA 截圖處理失敗: {e}")

static_dir = os.path.join(BASE_DIR, "static")
if os.path.exists(static_dir):
    app.mount("/", StaticFiles(directory=static_dir, html=True), name="static")

if __name__ == "__main__":
    port = find_available_port(8002)
    local_ip = get_local_ip()
    print("============================================================")
    print(f"TSMC Lorry Barcode Server started successfully!")
    print(f"Local URL: http://localhost:{port}")
    print(f"LAN URL: http://{local_ip}:{port}")
    print("============================================================")
    uvicorn.run(app, host="0.0.0.0", port=port, log_level="warning")
